import gc
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from scripts.invoice_fetch.attachment_handler import build_managed_attachment_name
from scripts.invoice_fetch.excel_export import export_excel


class AttachmentNamingTests(unittest.TestCase):
    def test_naming_with_full_details(self):
        name = build_managed_attachment_name(
            original_name="my_invoice.pdf",
            invoice_date="2026-06-13",
            expense_date="2026-06-12",
            fallback_date="2026-06-11",
            category="餐饮",
            total_amount="123.45",
            invoice_number="98765432",
            role="原件",
        )
        self.assertEqual(name, "2026-06-12_餐饮_123.45_98765432_原件.pdf")

    def test_naming_with_missing_details_uses_original_name(self):
        name = build_managed_attachment_name(
            original_name="test.ofd",
            invoice_date="2026-06-13",
            expense_date=None,
            category=None,
            total_amount=None,
            invoice_number=None,
            role=None,
        )
        self.assertEqual(name, "2026-06-13_test.ofd")

    def test_naming_partial_fields_uses_explicit_placeholders(self):
        name = build_managed_attachment_name(
            original_name="receipt.png",
            invoice_date="2026-06-13",
            category="交通",
            role="证明材料",
        )
        self.assertEqual(
            name,
            "2026-06-13_交通_金额待补全_待补全_证明材料.png",
        )


class ExcelSortingTests(unittest.TestCase):
    def test_export_sorts_by_effective_date_without_mutating_input(self):
        rows = [
            {"id": 1, "expense_date": "", "invoice_date": "", "mail_date": "", "created_at": "", "invoice_number": "1"},
            {"id": 2, "expense_date": "2026-06-12", "invoice_number": "2"},
            {"id": 3, "expense_date": "", "invoice_date": "2026-06-10", "invoice_number": "3"},
            {"id": 4, "expense_date": "", "invoice_date": "", "mail_date": "2026-06-08", "invoice_number": "4"},
            {"id": 5, "expense_date": "", "invoice_date": "", "mail_date": "", "created_at": "2026-06-05", "invoice_number": "5"},
        ]
        original = [dict(row) for row in rows]

        with tempfile.TemporaryDirectory() as td:
            destination = Path(td) / "test.xlsx"
            export_excel(rows, destination)
            workbook = load_workbook(destination)
            sheet = workbook.active
            invoice_numbers = [sheet.cell(row=index, column=1).value for index in range(2, 7)]
            workbook.close()
            del sheet, workbook
            gc.collect()

        self.assertEqual(rows, original)
        self.assertEqual(invoice_numbers, ["5", "4", "3", "2", "1"])


if __name__ == "__main__":
    unittest.main()
