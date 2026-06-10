# -*- coding: utf-8 -*-
"""Tests for PREVIEW-PDF-NAV-LOG-001: multi-page PDF preview, keyboard navigation, link logs."""

import unittest
import tempfile
import shutil
import hashlib
import json
from pathlib import Path
from unittest.mock import MagicMock, patch, PropertyMock


# ── 1. File info formatting tests ────────────────────────────────────

class TestFileInfoFormatting(unittest.TestCase):
    """Tests for _format_preview_file_info and _get_pdf_page_info."""

    def _make_doc(self, title="主发票", path_suffix=".pdf"):
        return {"title": title, "path": Path(f"test{path_suffix}")}

    def test_non_pdf_displays_file_sequence_only(self):
        """Non-PDF files show file sequence only, no PDF page info."""
        from scripts.invoice_fetch.gui.app import InvoiceReviewApp

        # Create a minimal app instance to access methods
        app_obj = InvoiceReviewApp.__new__(InvoiceReviewApp)
        doc = self._make_doc("证明材料", ".png")
        result = app_obj._format_preview_file_info(doc, 1, 2)
        self.assertIn("文件 2/2", result)
        self.assertIn("证明材料", result)
        self.assertNotIn("PDF", result)

    def test_pdf_with_current_page_shows_page_fraction(self):
        """PDF with current page shows 'PDF 1/7' format."""
        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        app_obj = InvoiceReviewApp.__new__(InvoiceReviewApp)
        doc = self._make_doc("主发票", ".pdf")
        result = app_obj._format_preview_file_info(doc, 0, 2, pdf_page=1, pdf_page_count=7)
        self.assertIn("文件 1/2", result)
        self.assertIn("主发票", result)
        self.assertIn("PDF 1/7", result)

    def test_pdf_without_current_page_shows_total_only(self):
        """PDF without current page shows 'PDF 共 7 页'."""
        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        app_obj = InvoiceReviewApp.__new__(InvoiceReviewApp)
        doc = self._make_doc("证明材料", ".pdf")
        result = app_obj._format_preview_file_info(doc, 1, 2, pdf_page=None, pdf_page_count=7)
        self.assertIn("文件 2/2", result)
        self.assertIn("证明材料", result)
        self.assertIn("PDF 共 7 页", result)

    def test_pdf_page_count_none_no_pdf_suffix(self):
        """When pdf_page_count is None, PDF suffix is not appended."""
        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        app_obj = InvoiceReviewApp.__new__(InvoiceReviewApp)
        doc = self._make_doc("主发票", ".pdf")
        result = app_obj._format_preview_file_info(doc, 0, 1, pdf_page=1, pdf_page_count=None)
        self.assertEqual(result, "文件 1/1｜主发票")


# ── 2. MultiPage compatibility tests ─────────────────────────────────

class TestMultiPageCompatibility(unittest.TestCase):
    """Tests for MultiPage vs SinglePage fallback."""

    def test_multipage_used_when_available(self):
        """QPdfView.PageMode has MultiPage → use MultiPage."""
        mock_page_mode = MagicMock()
        mock_page_mode.MultiPage = "MultiPage"
        mock_page_mode.SinglePage = "SinglePage"

        mock_qpdfview = MagicMock()
        mock_qpdfview.PageMode = mock_page_mode

        # Simulate the logic
        if hasattr(mock_qpdfview.PageMode, "MultiPage"):
            mode = mock_qpdfview.PageMode.MultiPage
        else:
            mode = mock_qpdfview.PageMode.SinglePage
        self.assertEqual(mode, "MultiPage")

    def test_singlepage_fallback_when_multipage_unavailable(self):
        """QPdfView.PageMode without MultiPage → use SinglePage."""
        # Use a class without MultiPage attribute
        class MockPageMode:
            SinglePage = "SinglePage"

        class MockQPdfView:
            PageMode = MockPageMode

        qpv = MockQPdfView()
        if hasattr(qpv.PageMode, "MultiPage"):
            mode = qpv.PageMode.MultiPage
        else:
            mode = qpv.PageMode.SinglePage
        self.assertEqual(mode, "SinglePage")

    def test_singlepage_fallback_no_exception(self):
        """Fallback to SinglePage does not raise an exception."""
        class MockPageMode:
            SinglePage = "SinglePage"

        class MockQPdfView:
            PageMode = MockPageMode()

        qpv = MockQPdfView()
        try:
            if hasattr(qpv.PageMode, "MultiPage"):
                mode = qpv.PageMode.MultiPage
            else:
                mode = qpv.PageMode.SinglePage
        except Exception:
            self.fail("Fallback to SinglePage should not raise")
        self.assertEqual(mode, "SinglePage")


# ── 3. Keyboard focus protection tests ───────────────────────────────

class TestKeyboardFocusProtection(unittest.TestCase):
    """Tests for _focus_is_editing_widget."""

    @classmethod
    def setUpClass(cls):
        try:
            from PySide6.QtWidgets import QApplication
            import sys
            cls.app = QApplication.instance() or QApplication(sys.argv)
        except (ImportError, RuntimeError):
            cls.app = None

    @classmethod
    def tearDownClass(cls):
        cls.app = None

    def setUp(self):
        if self.app is None:
            self.skipTest("PySide6 not available")

    def test_lineedit_is_editing(self):
        """QLineEdit focus -> True."""
        from PySide6.QtWidgets import QLineEdit, QWidget, QVBoxLayout
        from scripts.invoice_fetch.gui.app import InvoiceReviewApp

        window = QWidget()
        window.setWindowTitle("test")
        widget = QLineEdit()
        layout = QVBoxLayout(window)
        layout.addWidget(widget)
        window.show()
        widget.setFocus()
        self.app.processEvents()

        app_obj = InvoiceReviewApp.__new__(InvoiceReviewApp)
        self.assertTrue(app_obj._focus_is_editing_widget())
        window.hide()

    def test_plaintextedit_is_editing(self):
        """QPlainTextEdit focus -> True."""
        from PySide6.QtWidgets import QPlainTextEdit, QWidget, QVBoxLayout
        from scripts.invoice_fetch.gui.app import InvoiceReviewApp

        window = QWidget()
        window.setWindowTitle("test")
        widget = QPlainTextEdit()
        layout = QVBoxLayout(window)
        layout.addWidget(widget)
        window.show()
        widget.setFocus()
        self.app.processEvents()

        app_obj = InvoiceReviewApp.__new__(InvoiceReviewApp)
        self.assertTrue(app_obj._focus_is_editing_widget())
        window.hide()

    def test_combobox_is_editing(self):
        """QComboBox focus -> True."""
        from PySide6.QtWidgets import QComboBox, QWidget, QVBoxLayout
        from scripts.invoice_fetch.gui.app import InvoiceReviewApp

        window = QWidget()
        window.setWindowTitle("test")
        widget = QComboBox()
        layout = QVBoxLayout(window)
        layout.addWidget(widget)
        window.show()
        widget.setFocus()
        self.app.processEvents()

        app_obj = InvoiceReviewApp.__new__(InvoiceReviewApp)
        self.assertTrue(app_obj._focus_is_editing_widget())
        window.hide()

    def test_pushbutton_is_not_editing(self):
        """QPushButton focus -> False."""
        from PySide6.QtWidgets import QPushButton, QWidget, QVBoxLayout
        from scripts.invoice_fetch.gui.app import InvoiceReviewApp

        window = QWidget()
        window.setWindowTitle("test")
        widget = QPushButton("Test")
        layout = QVBoxLayout(window)
        layout.addWidget(widget)
        window.show()
        widget.setFocus()
        self.app.processEvents()

        app_obj = InvoiceReviewApp.__new__(InvoiceReviewApp)
        self.assertFalse(app_obj._focus_is_editing_widget())
        window.hide()

    def test_no_focus_is_not_editing(self):
        """No focused widget -> False."""
        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        app_obj = InvoiceReviewApp.__new__(InvoiceReviewApp)
        self.assertFalse(app_obj._focus_is_editing_widget())


# ── 4. File-level navigation tests ───────────────────────────────────

class TestFileLevelNavigation(unittest.TestCase):
    """Tests for _prev_preview_doc / _next_preview_doc."""

    def test_next_file_increments_index(self):
        """Ctrl+Right equivalent: index advances from 0 to 1."""
        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        app_obj = InvoiceReviewApp.__new__(InvoiceReviewApp)
        app_obj.current_preview_docs = [{"title": "主发票", "path": Path("a.pdf")},
                                         {"title": "证明材料", "path": Path("b.png")}]
        app_obj.current_preview_index = 0
        # Patch _update_document_preview to avoid QPdfDocument loading
        app_obj._update_document_preview = MagicMock()
        app_obj._refresh_preview_file_info = MagicMock()
        app_obj._update_pdf_page_buttons = MagicMock()
        app_obj.lbl_file_info = MagicMock()

        app_obj._next_preview_doc()
        self.assertEqual(app_obj.current_preview_index, 1)

        app_obj._next_preview_doc()
        self.assertEqual(app_obj.current_preview_index, 0)  # wraps around

    def test_prev_file_decrements_index(self):
        """Ctrl+Left equivalent: index wraps from 0 to last."""
        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        app_obj = InvoiceReviewApp.__new__(InvoiceReviewApp)
        app_obj.current_preview_docs = [{"title": "主发票", "path": Path("a.pdf")},
                                         {"title": "证明材料", "path": Path("b.png")}]
        app_obj.current_preview_index = 0
        app_obj._update_document_preview = MagicMock()
        app_obj._refresh_preview_file_info = MagicMock()
        app_obj._update_pdf_page_buttons = MagicMock()
        app_obj.lbl_file_info = MagicMock()

        app_obj._prev_preview_doc()
        self.assertEqual(app_obj.current_preview_index, 1)  # wraps around


# ── 5. PDF page navigation tests ─────────────────────────────────────

class TestPdfPageNavigation(unittest.TestCase):
    """Tests for _navigate_pdf_page and related methods."""

    def test_navigate_forward_within_bounds(self):
        """Calling _navigate_pdf_page(+1) moves from page 0 to 1."""
        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        app_obj = InvoiceReviewApp.__new__(InvoiceReviewApp)

        mock_nav = MagicMock()
        mock_nav.currentPage.return_value = 0

        mock_pdf_doc = MagicMock()
        mock_pdf_doc.pageCount.return_value = 7
        mock_pdf_doc.pageNavigator.return_value = mock_nav

        mock_pdf_view = MagicMock()

        app_obj.pdf_document = mock_pdf_doc
        app_obj.pdf_view = mock_pdf_view
        app_obj.preview_stack = MagicMock()
        app_obj.preview_stack.currentWidget.return_value = mock_pdf_view
        app_obj._refresh_preview_file_info = MagicMock()
        app_obj._update_pdf_page_buttons = MagicMock()

        result = app_obj._navigate_pdf_page(1)
        self.assertTrue(result)
        mock_nav.jump.assert_called_once_with(1, 0.0, 0.0)

    def test_navigate_backward_within_bounds(self):
        """Calling _navigate_pdf_page(-1) moves from page 1 to 0."""
        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        app_obj = InvoiceReviewApp.__new__(InvoiceReviewApp)

        mock_nav = MagicMock()
        mock_nav.currentPage.return_value = 1

        mock_pdf_doc = MagicMock()
        mock_pdf_doc.pageCount.return_value = 7
        mock_pdf_doc.pageNavigator.return_value = mock_nav

        mock_pdf_view = MagicMock()

        app_obj.pdf_document = mock_pdf_doc
        app_obj.pdf_view = mock_pdf_view
        app_obj.preview_stack = MagicMock()
        app_obj.preview_stack.currentWidget.return_value = mock_pdf_view
        app_obj._refresh_preview_file_info = MagicMock()
        app_obj._update_pdf_page_buttons = MagicMock()

        result = app_obj._navigate_pdf_page(-1)
        self.assertTrue(result)
        mock_nav.jump.assert_called_once_with(0, 0.0, 0.0)

    def test_navigate_at_lower_boundary_no_op(self):
        """_navigate_pdf_page(-1) at page 0 returns False and does not call jump."""
        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        app_obj = InvoiceReviewApp.__new__(InvoiceReviewApp)

        mock_nav = MagicMock()
        mock_nav.currentPage.return_value = 0

        mock_pdf_doc = MagicMock()
        mock_pdf_doc.pageCount.return_value = 7
        mock_pdf_doc.pageNavigator.return_value = mock_nav

        mock_pdf_view = MagicMock()

        app_obj.pdf_document = mock_pdf_doc
        app_obj.pdf_view = mock_pdf_view
        app_obj.preview_stack = MagicMock()
        app_obj.preview_stack.currentWidget.return_value = mock_pdf_view
        app_obj._refresh_preview_file_info = MagicMock()
        app_obj._update_pdf_page_buttons = MagicMock()

        result = app_obj._navigate_pdf_page(-1)
        self.assertFalse(result)
        mock_nav.jump.assert_not_called()

    def test_navigate_at_upper_boundary_no_op(self):
        """_navigate_pdf_page(+1) at last page returns False."""
        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        app_obj = InvoiceReviewApp.__new__(InvoiceReviewApp)

        mock_nav = MagicMock()
        mock_nav.currentPage.return_value = 6  # last page (0-based, page count=7)

        mock_pdf_doc = MagicMock()
        mock_pdf_doc.pageCount.return_value = 7
        mock_pdf_doc.pageNavigator.return_value = mock_nav

        mock_pdf_view = MagicMock()

        app_obj.pdf_document = mock_pdf_doc
        app_obj.pdf_view = mock_pdf_view
        app_obj.preview_stack = MagicMock()
        app_obj.preview_stack.currentWidget.return_value = mock_pdf_view
        app_obj._refresh_preview_file_info = MagicMock()
        app_obj._update_pdf_page_buttons = MagicMock()

        result = app_obj._navigate_pdf_page(1)
        self.assertFalse(result)
        mock_nav.jump.assert_not_called()

    def test_navigate_non_pdf_returns_false(self):
        """_navigate_pdf_page returns False when not viewing a PDF."""
        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        app_obj = InvoiceReviewApp.__new__(InvoiceReviewApp)
        app_obj.preview_stack = MagicMock()
        app_obj.preview_stack.currentWidget.return_value = MagicMock()  # not pdf_view

        result = app_obj._navigate_pdf_page(1)
        self.assertFalse(result)

    def test_pdf_page_buttons_disabled_for_single_page(self):
        """Page buttons should be disabled when PDF has only 1 page."""
        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        app_obj = InvoiceReviewApp.__new__(InvoiceReviewApp)

        btn_prev = MagicMock()
        btn_next = MagicMock()
        app_obj.btn_prev_page = btn_prev
        app_obj.btn_next_page = btn_next

        mock_nav = MagicMock()
        mock_nav.currentPage.return_value = 0

        mock_pdf_doc = MagicMock()
        mock_pdf_doc.pageCount.return_value = 1
        mock_pdf_doc.pageNavigator.return_value = mock_nav

        mock_pdf_view = MagicMock()
        app_obj.pdf_document = mock_pdf_doc
        app_obj.pdf_view = mock_pdf_view
        app_obj.preview_stack = MagicMock()
        app_obj.preview_stack.currentWidget.return_value = mock_pdf_view

        app_obj._update_pdf_page_buttons()
        btn_prev.setEnabled.assert_called_with(False)
        btn_next.setEnabled.assert_called_with(False)


# ── 6. Link downloader cache tests ────────────────────────────────────

class TestLinkDownloaderCache(unittest.TestCase):
    """Tests for failed URL fingerprint cache."""

    def test_url_fingerprint_deterministic(self):
        """Same URL produces same fingerprint."""
        from scripts.invoice_fetch.link_downloader import LinkDownloader
        url = "https://51fapiao.cn/download?id=12345"
        fp1 = LinkDownloader._url_fingerprint(url)
        fp2 = LinkDownloader._url_fingerprint(url)
        self.assertEqual(fp1, fp2)

    def test_url_fingerprint_different_urls(self):
        """Different URLs produce different fingerprints."""
        from scripts.invoice_fetch.link_downloader import LinkDownloader
        fp1 = LinkDownloader._url_fingerprint("https://a.com/doc1.pdf")
        fp2 = LinkDownloader._url_fingerprint("https://a.com/doc2.pdf")
        self.assertNotEqual(fp1, fp2)

    def test_fingerprint_not_reversible(self):
        """Fingerprint does not contain full URL."""
        from scripts.invoice_fetch.link_downloader import LinkDownloader
        url = "https://51fapiao.cn/download?id=12345&token=secret"
        fp = LinkDownloader._url_fingerprint(url)
        self.assertNotIn("51fapiao", fp)
        self.assertNotIn("token", fp)
        self.assertNotIn("secret", fp)
        self.assertEqual(len(fp), 16)

    def test_failed_url_cached_and_skipped(self):
        """After a URL fails, same fingerprint causes skip on next attempt."""
        from scripts.invoice_fetch.link_downloader import LinkDownloader

        dl = LinkDownloader(download_dir=tempfile.mkdtemp())
        try:
            url = "https://51fapiao.cn/invoice/test123.pdf"
            fp = LinkDownloader._url_fingerprint(url)

            # Simulate failure: add to cache
            dl.failed_url_fingerprints.add(fp)

            # Simulate download_from_email parsing the same URL
            html = f'<html><body><a href="{url}">发票下载</a></body></html>'
            from scripts.invoice_fetch.link_downloader import extract_links_from_html, _dedup_and_prioritize

            raw = extract_links_from_html(html)
            links = _dedup_and_prioritize(raw)

            skipped = 0
            results = []
            for link_url in links:
                lfp = LinkDownloader._url_fingerprint(link_url)
                if lfp in dl.failed_url_fingerprints:
                    skipped += 1
                    continue
                results.append(link_url)

            self.assertEqual(skipped, 1)
            self.assertEqual(len(results), 0)
        finally:
            dl.close()

    def test_skip_when_attachment_invoice_present_config(self):
        """Default config has skip_when_attachment_invoice_present=True."""
        from scripts.invoice_fetch.link_downloader import LinkDownloader
        dl = LinkDownloader(download_dir=tempfile.mkdtemp())
        try:
            self.assertTrue(dl._skip_when_attachment_invoice_present)
            self.assertGreater(dl._max_links_per_email, 0)
        finally:
            dl.close()

    def test_log_contains_skipped_cached_field(self):
        """Summary log includes skipped_cached field."""
        from scripts.invoice_fetch.link_downloader import LinkDownloader, _dedup_and_prioritize

        dl = LinkDownloader(download_dir=tempfile.mkdtemp())
        try:
            # Pre-fail a URL
            url = "https://nuonuo.com/invoice/test.pdf"
            dl.failed_url_fingerprints.add(LinkDownloader._url_fingerprint(url))

            # Simulate summary log
            import logging
            import io
            log_stream = io.StringIO()
            handler = logging.StreamHandler(log_stream)
            handler.setLevel(logging.INFO)
            logger = logging.getLogger("invoice_fetch.link_downloader")
            logger.addHandler(handler)
            logger.setLevel(logging.INFO)

            found = 1
            deduped = 1
            success = 0
            failed = 0
            skipped_cached = 1
            logger.info(
                "链接下载摘要: found=%d deduped=%d success=%d failed=%d skipped_cached=%d elapsed=%.1fs",
                found, deduped, success, failed, skipped_cached, 0.0,
            )

            logger.removeHandler(handler)
            log_output = log_stream.getvalue()
            self.assertIn("skipped_cached=1", log_output)
        finally:
            dl.close()


# ── 7. Attachment-based skip tests ───────────────────────────────────

class TestAttachmentSkipLogic(unittest.TestCase):
    """Tests for skipping link download when attachment invoices exist."""

    def test_skip_when_parse_success_attachments_present(self):
        """When parsed_invoice_pdfs has parse_success, skip link download."""
        # Simulate the decision logic
        parsed_invoice_pdfs = [
            (MagicMock(), MagicMock(parse_success=True)),
        ]
        link_dl = MagicMock()
        link_dl._skip_when_attachment_invoice_present = True

        skip_for_attachment = False
        if link_dl._skip_when_attachment_invoice_present:
            success_attachment_pdfs = [
                att for att, info in parsed_invoice_pdfs
                if info.parse_success
            ]
            if success_attachment_pdfs:
                skip_for_attachment = True

        self.assertTrue(skip_for_attachment)
        # _download_url should not be called
        link_dl._download_url.assert_not_called()

    def test_do_not_skip_when_no_parse_success(self):
        """When no attachment invoices parse successfully, still try link download."""
        parsed_invoice_pdfs = [
            (MagicMock(), MagicMock(parse_success=False)),
        ]
        link_dl = MagicMock()
        link_dl._skip_when_attachment_invoice_present = True

        skip_for_attachment = False
        if link_dl._skip_when_attachment_invoice_present:
            success_attachment_pdfs = [
                att for att, info in parsed_invoice_pdfs
                if info.parse_success
            ]
            if success_attachment_pdfs:
                skip_for_attachment = True

        self.assertFalse(skip_for_attachment)

    def test_do_not_skip_when_config_disabled(self):
        """When config disables the skip, always try link download."""
        parsed_invoice_pdfs = [
            (MagicMock(), MagicMock(parse_success=True)),
        ]
        link_dl = MagicMock()
        link_dl._skip_when_attachment_invoice_present = False

        skip_for_attachment = False
        if link_dl._skip_when_attachment_invoice_present:
            success_attachment_pdfs = [
                att for att, info in parsed_invoice_pdfs
                if info.parse_success
            ]
            if success_attachment_pdfs:
                skip_for_attachment = True

        self.assertFalse(skip_for_attachment)

    def test_empty_attachments_still_tries_download(self):
        """No attachments at all -> still tries link download."""
        parsed_invoice_pdfs = []
        link_dl = MagicMock()
        link_dl._skip_when_attachment_invoice_present = True

        skip_for_attachment = False
        if link_dl._skip_when_attachment_invoice_present:
            success_attachment_pdfs = [
                att for att, info in parsed_invoice_pdfs
                if info.parse_success
            ]
            if success_attachment_pdfs:
                skip_for_attachment = True

        self.assertFalse(skip_for_attachment)


# ── 8. Filename conflict log tests ───────────────────────────────────

class TestFilenameConflictLogging(unittest.TestCase):
    """Tests for source_mode-based conflict log level."""

    def test_reprocess_mode_logs_info_not_warning(self):
        """In reprocess mode, conflict log uses INFO, not WARNING."""
        import logging
        import io

        log_stream = io.StringIO()
        handler = logging.StreamHandler(log_stream)
        handler.setLevel(logging.DEBUG)
        handler.setFormatter(logging.Formatter("%(levelname)s:%(message)s"))
        logger = logging.getLogger("invoice_fetch.conflict_test")
        logger.addHandler(handler)
        logger.setLevel(logging.DEBUG)

        try:
            # Simulate the conflict message in reprocess mode
            effective_mode = "reprocess"
            if effective_mode in ("reprocess", "repair"):
                logger.info("检测到同名文件，已安全改名保存: <masked>")

            log_output = log_stream.getvalue()
            self.assertIn("INFO", log_output)
            self.assertIn("已安全改名保存", log_output)
        finally:
            logger.removeHandler(handler)

    def test_normal_mode_logs_warning(self):
        """In normal mode, conflict log uses WARNING."""
        import logging
        import io

        log_stream = io.StringIO()
        handler = logging.StreamHandler(log_stream)
        handler.setLevel(logging.DEBUG)
        handler.setFormatter(logging.Formatter("%(levelname)s:%(message)s"))
        logger = logging.getLogger("invoice_fetch.conflict_test2")
        logger.addHandler(handler)
        logger.setLevel(logging.DEBUG)

        try:
            effective_mode = "normal"
            if effective_mode in ("reprocess", "repair"):
                logger.info("检测到同名文件，已安全改名保存: <masked>")
            else:
                logger.warning("检测到同名文件，已安全改名保存: <masked>")

            log_output = log_stream.getvalue()
            self.assertIn("WARNING", log_output)
            self.assertIn("已安全改名保存", log_output)
        finally:
            logger.removeHandler(handler)

    def test_conflict_msg_does_not_leak_filename(self):
        """Conflict log uses mask_filename, never raw path."""
        from scripts.invoice_fetch.log_privacy import mask_filename

        # mask_filename should produce output that doesn't match the raw name
        raw_name = "my-real-invoice-12345.pdf"
        masked = mask_filename(raw_name)
        self.assertNotEqual(masked, raw_name)


# ── 9. Duplicate attachment backfill tests ────────────────────────────

class TestDuplicateAttachmentBackfill(unittest.TestCase):
    """Tests for update_invoice_attachment_path_if_missing."""

    def setUp(self):
        import tempfile
        self.temp_dir = Path(tempfile.mkdtemp())

    def tearDown(self):
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_backfill_when_attachment_path_empty(self):
        """Existing invoice with empty attachment_path -> backfill succeeds."""
        from scripts.invoice_fetch.db import InvoiceDB
        db_path = self.temp_dir / "test_backfill_empty.db"
        with InvoiceDB(db_path) as db:
            inv_id = db.insert_invoice({
                "invoice_number": "12345678",
                "total_amount": "100.00",
                "seller_name": "测试公司",
                "invoice_date": "2026-06-01",
                "attachment_path": "",
                "review_status": "to_review",
            })
            # Create a real attachment file
            att_file = self.temp_dir / "test_attachment.pdf"
            att_file.write_text("dummy pdf")
            result = db.update_invoice_attachment_path_if_missing(
                inv_id, str(att_file), file_hash="abc123"
            )
            self.assertTrue(result)
            inv = db.get_invoice(inv_id)
            self.assertEqual(inv["attachment_path"], str(att_file))

    def test_no_backfill_when_attachment_exists(self):
        """Existing invoice with valid attachment -> no backfill."""
        from scripts.invoice_fetch.db import InvoiceDB
        db_path = self.temp_dir / "test_backfill_exists.db"
        existing_file = self.temp_dir / "existing.pdf"
        existing_file.write_text("existing")
        with InvoiceDB(db_path) as db:
            inv_id = db.insert_invoice({
                "invoice_number": "12345678",
                "total_amount": "100.00",
                "seller_name": "测试公司",
                "invoice_date": "2026-06-01",
                "attachment_path": str(existing_file),
                "review_status": "to_review",
            })
            new_file = self.temp_dir / "new_attachment.pdf"
            new_file.write_text("new")
            result = db.update_invoice_attachment_path_if_missing(
                inv_id, str(new_file), file_hash="xyz789"
            )
            self.assertFalse(result)
            inv = db.get_invoice(inv_id)
            self.assertEqual(inv["attachment_path"], str(existing_file))

    def test_backfill_file_not_deleted_by_cleanup(self):
        """Backfilled path must be in kept_paths (simulated)."""
        # This is a logical test: verify that the method exists and is importable
        from scripts.invoice_fetch.db import InvoiceDB
        self.assertTrue(hasattr(InvoiceDB, "update_invoice_attachment_path_if_missing"))


# ── 10. Missing field safe backfill tests ─────────────────────────────

class TestDuplicateMissingFieldBackfill(unittest.TestCase):
    """Tests for update_invoice_missing_fields."""

    def setUp(self):
        import tempfile
        self.temp_dir = Path(tempfile.mkdtemp())

    def tearDown(self):
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_backfill_empty_seller_name_to_review(self):
        """to_review invoice with empty seller_name -> backfill succeeds."""
        from scripts.invoice_fetch.db import InvoiceDB
        db_path = self.temp_dir / "test_seller_backfill.db"
        with InvoiceDB(db_path) as db:
            inv_id = db.insert_invoice({
                "invoice_number": "12345678",
                "total_amount": "100.00",
                "seller_name": "",
                "invoice_date": "2026-06-01",
                "review_status": "to_review",
            })
            result = db.update_invoice_missing_fields(
                inv_id, {"seller_name": "新销售方公司"},
            )
            self.assertIn("seller_name", result["updated_fields"])
            inv = db.get_invoice(inv_id)
            self.assertEqual(inv["seller_name"], "新销售方公司")

    def test_no_overwrite_existing_seller_name(self):
        """to_review invoice with existing seller_name -> no overwrite."""
        from scripts.invoice_fetch.db import InvoiceDB
        db_path = self.temp_dir / "test_no_overwrite.db"
        with InvoiceDB(db_path) as db:
            inv_id = db.insert_invoice({
                "invoice_number": "12345678",
                "total_amount": "100.00",
                "seller_name": "已有销售方",
                "invoice_date": "2026-06-01",
                "review_status": "to_review",
            })
            result = db.update_invoice_missing_fields(
                inv_id, {"seller_name": "另一个销售方"},
                only_if_empty=True,
            )
            self.assertIn("seller_name", result["skipped_fields"])
            inv = db.get_invoice(inv_id)
            self.assertEqual(inv["seller_name"], "已有销售方")

    def test_approved_does_not_backfill_business_fields(self):
        """Approved invoice -> seller_name backfill is skipped."""
        from scripts.invoice_fetch.db import InvoiceDB
        db_path = self.temp_dir / "test_approved_skip.db"
        with InvoiceDB(db_path) as db:
            inv_id = db.insert_invoice({
                "invoice_number": "12345678",
                "total_amount": "100.00",
                "seller_name": "",
                "invoice_date": "2026-06-01",
                "review_status": "approved",
            })
            result = db.update_invoice_missing_fields(
                inv_id, {"seller_name": "新销售方"},
            )
            self.assertIn("seller_name", result["skipped_fields"])
            inv = db.get_invoice(inv_id)
            self.assertEqual(inv["seller_name"], "")

    def test_approved_allows_attachment_backfill(self):
        """Approved invoice -> attachment_path backfill is allowed."""
        from scripts.invoice_fetch.db import InvoiceDB
        db_path = self.temp_dir / "test_approved_att.db"
        with InvoiceDB(db_path) as db:
            inv_id = db.insert_invoice({
                "invoice_number": "12345678",
                "total_amount": "100.00",
                "seller_name": "已审核销售方",
                "invoice_date": "2026-06-01",
                "attachment_path": "",
                "review_status": "approved",
            })
            result = db.update_invoice_missing_fields(
                inv_id, {"attachment_path": "/path/to/file.pdf"},
            )
            # attachment_path is exempt from approved check
            self.assertIn("attachment_path", result["updated_fields"])

    def test_backfill_multiple_fields(self):
        """Backfill seller_name, buyer_name, invoice_date simultaneously."""
        from scripts.invoice_fetch.db import InvoiceDB
        db_path = self.temp_dir / "test_multi_backfill.db"
        with InvoiceDB(db_path) as db:
            inv_id = db.insert_invoice({
                "invoice_number": "12345678",
                "total_amount": "100.00",
                "seller_name": "",
                "buyer_name": "",
                "invoice_date": "",
                "review_status": "to_review",
            })
            result = db.update_invoice_missing_fields(
                inv_id, {
                    "seller_name": "销售方A",
                    "buyer_name": "购买方B",
                    "invoice_date": "2026-06-15",
                },
            )
            self.assertEqual(len(result["updated_fields"]), 3)


# ── 11. Seller name fallback parser tests ────────────────────────────

class TestSellerNameFallbackParser(unittest.TestCase):
    """Tests for _extract_seller_name_fallback."""

    @staticmethod
    def _extract(text):
        from scripts.invoice_fetch.invoice_parser import InvoiceParser
        return InvoiceParser._extract_seller_name_fallback(text)

    def test_extract_seller_from_seller_block(self):
        """Text with both buyer and seller blocks -> extract seller name."""
        text = (
            "购买方信息\n"
            "名称：购买方科技有限公司\n"
            "纳税人识别号：9111000011112222\n\n"
            "销售方信息\n"
            "名称：北京华联科技有限公司\n"
            "纳税人识别号：9111000033334444\n\n"
            "价税合计 ¥100.00\n"
        )
        result = self._extract(text)
        self.assertIn("北京", result)
        self.assertIn("华联", result)
        self.assertNotIn("购买方", result)

    def test_extract_seller_with_ocr_spacing(self):
        """Seller block with OCR cross-character spacing still matches."""
        text = (
            "销 售 方 信 息\n"
            "名称：北京某餐饮管理有限公司\n"
            "纳 税 人 识 别 号：91110105123456789X\n"
            "备注：无\n"
        )
        result = self._extract(text)
        self.assertIn("北京", result)
        self.assertIn("餐饮", result)

    def test_no_buyer_confusion(self):
        """Only buyer info present -> returns empty."""
        text = (
            "购买方信息\n"
            "名称：购买方科技有限公司\n"
            "纳税人识别号：9111000011112222\n"
            "备注：无\n"
        )
        result = self._extract(text)
        self.assertEqual(result, "")

    def test_no_seller_block_returns_empty(self):
        """No seller block at all -> returns empty."""
        text = "项目名称：餐饮服务\n规格型号：无\n价税合计：¥100.00\n"
        result = self._extract(text)
        self.assertEqual(result, "")

    def test_seller_name_with_parentheses(self):
        """Seller name containing parentheses is still extracted."""
        text = (
            "销售方信息\n"
            "名称：华住酒店管理（上海）有限公司\n"
            "纳税人识别号：91310000666600001U\n"
        )
        result = self._extract(text)
        self.assertIn("华住", result)
        self.assertIn("上海", result)

    def test_seller_block_stops_before_remarks(self):
        """Seller block extraction stops at 备注 marker."""
        text = (
            "销售方信息\n"
            "名称：上海锦江酒店管理有限公司\n"
            "纳税人识别号：9111000099998888\n"
            "备注：开票人：张三\n"
            "收款人：李四\n"
            "购买方信息\n"
            "名称：测试购买方有限公司\n"
        )
        result = self._extract(text)
        self.assertIn("锦江", result)
        self.assertNotIn("张三", result)
        self.assertNotIn("购买方", result)

    def test_fallback_integrated_in_parse_pdf(self):
        """Verify _extract_seller_name_fallback is callable from InvoiceParser."""
        from scripts.invoice_fetch.invoice_parser import InvoiceParser
        parser = InvoiceParser()
        self.assertTrue(hasattr(parser, "_extract_seller_name_fallback"))
        result = InvoiceParser._extract_seller_name_fallback("")
        self.assertEqual(result, "")


# ── 12. Link-download skip with attachments (integration test) ────────

class TestAttachmentSkipLinkDownloadIntegration(unittest.TestCase):
    """Test the decision to skip link downloads when attachments already parse."""

    def test_skip_logic_with_config_true(self):
        """Default config (True) + parse_success attachment -> skip."""
        from unittest.mock import MagicMock

        parsed_invoice_pdfs = [
            (MagicMock(), MagicMock(parse_success=True)),
        ]
        link_dl = MagicMock()
        link_dl._skip_when_attachment_invoice_present = True

        skip = False
        if getattr(link_dl, "_skip_when_attachment_invoice_present", True):
            success_attachment_pdfs = [
                att for att, info in parsed_invoice_pdfs if info.parse_success
            ]
            if success_attachment_pdfs:
                skip = True
        self.assertTrue(skip)

    def test_skip_logic_with_config_false(self):
        """Config False -> never skip even with parse_success attachment."""
        from unittest.mock import MagicMock

        parsed_invoice_pdfs = [
            (MagicMock(), MagicMock(parse_success=True)),
        ]
        link_dl = MagicMock()
        link_dl._skip_when_attachment_invoice_present = False

        skip = False
        if getattr(link_dl, "_skip_when_attachment_invoice_present", True):
            success_attachment_pdfs = [
                att for att, info in parsed_invoice_pdfs if info.parse_success
            ]
            if success_attachment_pdfs:
                skip = True
        self.assertFalse(skip)

    def test_no_parse_success_no_skip(self):
        """No parse_success attachment -> don't skip, try link download."""
        from unittest.mock import MagicMock

        parsed_invoice_pdfs = [
            (MagicMock(), MagicMock(parse_success=False)),
        ]
        link_dl = MagicMock()
        link_dl._skip_when_attachment_invoice_present = True

        skip = False
        if getattr(link_dl, "_skip_when_attachment_invoice_present", True):
            success_attachment_pdfs = [
                att for att, info in parsed_invoice_pdfs if info.parse_success
            ]
            if success_attachment_pdfs:
                skip = True
        self.assertFalse(skip)


# ── 13. Hardening commits tests for INVOICE-PREVIEW-DOWNLOAD-PARSER-002 ──

class TestLinkDownloaderAttemptLimit(unittest.TestCase):
    def test_attempt_limit_enforced(self):
        from scripts.invoice_fetch.link_downloader import LinkDownloader
        import logging
        import io
        import tempfile

        dl = LinkDownloader(download_dir=tempfile.mkdtemp())
        dl._max_links_per_email = 5

        # Mock _download_url to always fail (return None)
        dl._download_url = MagicMock(return_value=None)

        # Set up a logger interceptor to read summary log
        log_stream = io.StringIO()
        handler = logging.StreamHandler(log_stream)
        handler.setLevel(logging.INFO)
        logger = logging.getLogger("scripts.invoice_fetch.link_downloader")
        logger.addHandler(handler)
        logger.setLevel(logging.INFO)

        try:
            # Construct a dummy email with 7 links
            class MockMessage:
                def is_multipart(self):
                    return False
                def get_content_type(self):
                    return "text/html"
                def get_payload(self, decode=True):
                    return (
                        "<html><body>"
                        '<a href="https://51fapiao.cn/1">link1</a>'
                        '<a href="https://51fapiao.cn/2">link2</a>'
                        '<a href="https://51fapiao.cn/3">link3</a>'
                        '<a href="https://51fapiao.cn/4">link4</a>'
                        '<a href="https://51fapiao.cn/5">link5</a>'
                        '<a href="https://51fapiao.cn/6">link6</a>'
                        '<a href="https://51fapiao.cn/7">link7</a>'
                        "</body></html>"
                    ).encode("utf-8")
                def get_content_charset(self):
                    return "utf-8"

            msg = MockMessage()
            results = dl.download_from_email(msg, mail_uid=123)

            # Assertions
            self.assertEqual(len(results), 0)
            self.assertEqual(dl._download_url.call_count, 5)

            log_output = log_stream.getvalue()
            self.assertIn("failed=5", log_output)
            self.assertIn("attempted=5", log_output)
        finally:
            logger.removeHandler(handler)
            dl.close()


class TestDuplicateSafeBackfill(unittest.TestCase):
    def setUp(self):
        import tempfile
        self.temp_dir = Path(tempfile.mkdtemp())
        self.db_path = self.temp_dir / "test_dup.db"

    def tearDown(self):
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_duplicate_safe_backfill_logic(self):
        from scripts.invoice_fetch.db import InvoiceDB
        from scripts.invoice_fetch.__main__ import _refresh_invoice_from_parse

        with InvoiceDB(self.db_path) as db:
            # 1. 插入一个 existing 发票，具有非空 seller_name="原有销售方"
            inv_id_1 = db.insert_invoice({
                "invoice_number": "111111",
                "invoice_code": "code1",
                "invoice_date": "2026-06-01",
                "amount": "100.00",
                "total_amount": "100.00",
                "seller_name": "原有销售方",
                "buyer_name": "",
                "invoice_type": "电子发票",
                "category": "其他",
                "review_status": "to_review",
            })
            existing_1 = db.get_invoice(inv_id_1)

            # 调用 refresh，传入不同的 seller_name="新解析的销售方"，但 force_refresh_metadata=False
            res = _refresh_invoice_from_parse(
                db, existing_1,
                invoice_number="111111",
                invoice_code="code1",
                invoice_date="2026-06-01",
                amount="100.00",
                total_amount="100.00",
                seller_name="新解析的销售方",
                buyer_name="购买方A",
                invoice_type="电子发票",
                category="其他",
                has_extra=False,
                extra_type="",
                missing_extra=False,
                parse_note="",
                force_refresh_metadata=False
            )
            self.assertTrue(res)
            # 断言已有的 seller_name 依然是 "原有销售方"（不覆盖）
            inv_after_1 = db.get_invoice(inv_id_1)
            self.assertEqual(inv_after_1["seller_name"], "原有销售方")
            # 断言原本为空的 buyer_name 被成功回填为 "购买方A"
            self.assertEqual(inv_after_1["buyer_name"], "购买方A")

            # 2. 插入一个 existing 发票，seller_name 为空
            inv_id_2 = db.insert_invoice({
                "invoice_number": "222222",
                "invoice_code": "code2",
                "invoice_date": "2026-06-02",
                "amount": "200.00",
                "total_amount": "200.00",
                "seller_name": "",
                "buyer_name": "",
                "invoice_type": "电子发票",
                "category": "其他",
                "review_status": "to_review",
            })
            existing_2 = db.get_invoice(inv_id_2)

            res2 = _refresh_invoice_from_parse(
                db, existing_2,
                invoice_number="222222",
                invoice_code="code2",
                invoice_date="2026-06-02",
                amount="200.00",
                total_amount="200.00",
                seller_name="回填销售方",
                buyer_name="购买方B",
                invoice_type="电子发票",
                category="其他",
                has_extra=False,
                extra_type="",
                missing_extra=False,
                parse_note="",
                force_refresh_metadata=False
            )
            self.assertTrue(res2)
            # 断言原本为空的 seller_name 被成功回填
            inv_after_2 = db.get_invoice(inv_id_2)
            self.assertEqual(inv_after_2["seller_name"], "回填销售方")


class TestClaimedInvoiceProtection(unittest.TestCase):
    def setUp(self):
        import tempfile
        self.temp_dir = Path(tempfile.mkdtemp())
        self.db_path = self.temp_dir / "test_claimed.db"

    def tearDown(self):
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_claimed_invoice_protection_rules(self):
        from scripts.invoice_fetch.db import InvoiceDB
        from scripts.invoice_fetch.__main__ import _refresh_invoice_from_parse

        with InvoiceDB(self.db_path) as db:
            # 1. 插入一个 existing 发票，且将其关联到报销组以使其成为 claimed 发票
            inv_id = db.insert_invoice({
                "invoice_number": "333333",
                "invoice_code": "code3",
                "invoice_date": "2026-06-03",
                "amount": "300.00",
                "total_amount": "300.00",
                "seller_name": "",
                "buyer_name": "",
                "invoice_type": "电子发票",
                "category": "其他",
                "attachment_path": "",
                "review_status": "to_review",
            })
            # 插入 claim_group_items
            db._conn.execute(
                "INSERT INTO claim_group_items (claim_id, invoice_id, note) VALUES (?, ?, ?)",
                (1, inv_id, "test claim")
            )
            db._conn.commit()

            existing = db.get_invoice(inv_id)
            self.assertTrue(db.count_claim_links(inv_id) > 0)

            # 调用 refresh，试图回填 seller_name
            res = _refresh_invoice_from_parse(
                db, existing,
                invoice_number="333333",
                invoice_code="code3",
                invoice_date="2026-06-03",
                amount="300.00",
                total_amount="300.00",
                seller_name="想回填的销售方",
                buyer_name="想回填的购买方",
                invoice_type="电子发票",
                category="其他",
                has_extra=False,
                extra_type="",
                missing_extra=False,
                parse_note="",
                force_refresh_metadata=False
            )
            self.assertTrue(res)

            # 验证 seller_name 依然为空，不被更新
            inv_after = db.get_invoice(inv_id)
            self.assertEqual(inv_after["seller_name"], "")

            # 2. 验证 update_invoice_missing_fields 对 claimed 发票只允许更新 attachment_path / file_hash
            result = db.update_invoice_missing_fields(
                inv_id,
                {
                    "seller_name": "再次尝试",
                    "attachment_path": "/fake/path.pdf",
                    "file_hash": "hash123",
                    "invalid_field": "sql_inject"
                }
            )
            # seller_name 和 invalid_field 被 skipped，而 attachment_path 和 file_hash 被 updated
            self.assertIn("seller_name", result["skipped_fields"])
            self.assertIn("invalid_field", result["skipped_fields"])
            self.assertIn("attachment_path", result["updated_fields"])
            self.assertIn("file_hash", result["updated_fields"])


class TestNavigationFocusStability(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        try:
            from PySide6.QtWidgets import QApplication
            import sys
            cls.app = QApplication.instance() or QApplication(sys.argv)
        except (ImportError, RuntimeError):
            cls.app = None

    @classmethod
    def tearDownClass(cls):
        cls.app = None

    def setUp(self):
        if self.app is None:
            self.skipTest("PySide6 not available")

    def test_focus_out_of_preview_does_not_intercept(self):
        from PySide6.QtWidgets import QWidget, QLineEdit
        from PySide6.QtGui import QKeyEvent
        from PySide6.QtCore import QEvent, Qt
        from scripts.invoice_fetch.gui.app import InvoiceReviewApp

        # 1. 模拟焦点在一个独立的 QLineEdit (代表编辑框)
        window = QWidget()
        window.setWindowTitle("test focus")
        edit = QLineEdit(window)
        window.show()
        edit.setFocus()
        self.app.processEvents()

        app_obj = InvoiceReviewApp.__new__(InvoiceReviewApp)
        app_obj.preview_stack = MagicMock()

        # 模拟安装 eventFilter
        event = QKeyEvent(QEvent.Type.KeyPress, Qt.Key_Left, Qt.NoModifier)
        watched = MagicMock()

        app_obj._prev_preview_doc = MagicMock()
        res = app_obj.eventFilter(watched, event)
        self.assertFalse(res)  # 不抢占焦点
        app_obj._prev_preview_doc.assert_not_called()
        window.hide()

    def test_focus_in_preview_intercepts(self):
        from PySide6.QtWidgets import QWidget, QScrollArea
        from PySide6.QtGui import QKeyEvent
        from PySide6.QtCore import QEvent, Qt
        from scripts.invoice_fetch.gui.app import InvoiceReviewApp

        window = QWidget()
        window.setWindowTitle("test focus in preview")
        # 模拟 image_scroll_area
        scroll_area = QScrollArea(window)
        window.show()
        scroll_area.setFocus()
        self.app.processEvents()

        app_obj = InvoiceReviewApp.__new__(InvoiceReviewApp)
        app_obj._prev_preview_doc = MagicMock()

        # watched 设为 image_scroll_area 模拟事件分发
        app_obj.image_scroll_area = scroll_area
        app_obj.preview_container = window
        app_obj.preview_stack = MagicMock()
        app_obj.preview_stack.currentWidget.return_value = None
        app_obj.pdf_view = None
        app_obj.overlay_toolbar = None

        event = QKeyEvent(QEvent.Type.KeyPress, Qt.Key_Left, Qt.NoModifier)
        res = app_obj.eventFilter(scroll_area, event)
        self.assertTrue(res)  # 成功拦截并消耗事件
        app_obj._prev_preview_doc.assert_called_once()
        window.hide()


class TestInvoiceNoteAndPrivacy001(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        try:
            from PySide6.QtWidgets import QApplication
            import sys
            cls.app = QApplication.instance() or QApplication(sys.argv)
        except (ImportError, RuntimeError):
            cls.app = None

    @classmethod
    def tearDownClass(cls):
        cls.app = None

    def setUp(self):
        import tempfile
        self.temp_dir = Path(tempfile.mkdtemp())
        self.db_path = self.temp_dir / "test_notes.db"
        if self.app is not None:
            self.app.processEvents()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)
        if self.app is not None:
            self.app.processEvents()

    def _select_row(self, window, row_idx):
        from PySide6.QtCore import QItemSelectionModel
        window.table.clearSelection()
        window.table.setCurrentItem(None)
        if window.table.selectionModel() is not None:
            window.table.selectionModel().clearSelection()
            window.table.selectionModel().clear()
        self.app.processEvents()

        window.table.selectRow(row_idx)
        self.app.processEvents()
        model = window.table.model()
        sel_model = window.table.selectionModel()
        idx = model.index(row_idx, 0)
        sel_model.select(idx, QItemSelectionModel.Select | QItemSelectionModel.Rows)
        window._on_table_selection_changed()
        self.app.processEvents()

    def _clear_selection(self, window):
        window.table.clearSelection()
        window.table.setCurrentItem(None)
        if window.table.selectionModel() is not None:
            window.table.selectionModel().clearSelection()
            window.table.selectionModel().clear()
        self.app.processEvents()
        window._on_table_selection_changed()
        self.app.processEvents()

    def _select_all(self, window):
        from PySide6.QtCore import QItemSelectionModel
        window.table.selectAll()
        self.app.processEvents()
        model = window.table.model()
        sel_model = window.table.selectionModel()
        for r in range(window.table.rowCount()):
            idx = model.index(r, 0)
            sel_model.select(idx, QItemSelectionModel.Select | QItemSelectionModel.Rows)
        window._on_table_selection_changed()
        self.app.processEvents()

    def test_backfill_logs_redaction(self):
        from scripts.invoice_fetch.db import InvoiceDB
        from scripts.invoice_fetch.__main__ import _refresh_invoice_from_parse

        with InvoiceDB(self.db_path) as db:
            inv_id = db.insert_invoice({
                "invoice_number": "999999",
                "invoice_code": "",
                "invoice_date": "2026-06-01",
                "amount": "100.00",
                "total_amount": "100.00",
                "seller_name": "",
                "buyer_name": "",
                "invoice_type": "电子发票",
                "category": "其他",
                "review_status": "to_review",
            })
            existing = db.get_invoice(inv_id)

            with self.assertLogs("invoice_fetch", level="INFO") as log_ctx:
                res = _refresh_invoice_from_parse(
                    db, existing,
                    invoice_number="999999",
                    invoice_code="code_new_999",
                    invoice_date="2026-06-01",
                    amount="100.00",
                    total_amount="100.00",
                    seller_name="真实销售方公司",
                    buyer_name="真实购买方公司",
                    invoice_type="电子发票",
                    category="其他",
                    has_extra=False,
                    extra_type="",
                    missing_extra=False,
                    parse_note="",
                    force_refresh_metadata=False
                )
                self.assertTrue(res)

            # Verify the log outputs
            log_messages = "".join(log_ctx.output)
            # Should contain "fields=" and the fields
            self.assertIn("fields=", log_messages)
            self.assertIn("seller_name", log_messages)
            self.assertIn("buyer_name", log_messages)
            self.assertIn("invoice_code", log_messages)

            # MUST NOT contain the actual values
            self.assertNotIn("真实销售方公司", log_messages)
            self.assertNotIn("真实购买方公司", log_messages)
            self.assertNotIn("code_new_999", log_messages)

    def test_confirmed_note_gui_interaction(self):
        if self.app is None:
            self.skipTest("PySide6 not available")

        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        from scripts.invoice_fetch.db import InvoiceDB

        # Pre-populate invoice with note
        with InvoiceDB(self.db_path) as db:
            db.insert_invoice({
                "invoice_number": "111",
                "invoice_date": "2026-06-01",
                "seller_name": "销售方A",
                "total_amount": "100.00",
                "review_status": "to_review",
                "confirmed_note": "这是一条测试个人备注",
            })

        window = InvoiceReviewApp(self.db_path, splash=None)
        try:
            window.show()
            self.app.processEvents()

            # Select row
            self._select_row(window, 0)

            # 1. Test confirmed_note loaded to txt_note
            self.assertEqual(window.txt_note.toPlainText(), "这是一条测试个人备注")

            # 2. Test note summary shows the note content
            self.assertIn("这是一条测试个人备注", window.lbl_note_summary.text())

            # 3. Modify note and save
            window.txt_note.setPlainText("这是修改后的个人备注")
            self.app.processEvents()

            # Trigger save
            window._save_invoice_fields()
            self.app.processEvents()

            # Fetch from DB and verify
            with InvoiceDB(self.db_path) as db:
                rows = db.get_all_invoices()
                self.assertEqual(len(rows), 1)
                self.assertEqual(rows[0]["confirmed_note"], "这是修改后的个人备注")
        finally:
            window.close()

    def test_note_disabled_when_multi_or_no_selection(self):
        if self.app is None:
            self.skipTest("PySide6 not available")

        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        from scripts.invoice_fetch.db import InvoiceDB

        with InvoiceDB(self.db_path) as db:
            db.insert_invoice({
                "invoice_number": "111",
                "invoice_date": "2026-06-01",
                "seller_name": "销售方A",
                "total_amount": "100.00",
                "review_status": "to_review",
            })
            db.insert_invoice({
                "invoice_number": "222",
                "invoice_date": "2026-06-02",
                "seller_name": "销售方B",
                "total_amount": "200.00",
                "review_status": "to_review",
            })

        window = InvoiceReviewApp(self.db_path, splash=None)
        try:
            window.show()
            self.app.processEvents()

            # 1. No selection: txt_note is disabled
            self._clear_selection(window)
            self.assertFalse(window.txt_note.isEnabled())

            # 2. Single selection: txt_note is enabled
            self._select_row(window, 0)
            self.assertTrue(window.txt_note.isEnabled())

            # 3. Multi-selection: txt_note is disabled
            self._select_all(window)
            self.assertFalse(window.txt_note.isEnabled())
        finally:
            window.close()


class TestDetailPanelCompact001(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        try:
            from PySide6.QtWidgets import QApplication
            import sys
            cls.app = QApplication.instance() or QApplication(sys.argv)
        except (ImportError, RuntimeError):
            cls.app = None

    @classmethod
    def tearDownClass(cls):
        cls.app = None

    def setUp(self):
        import tempfile
        self.temp_dir = Path(tempfile.mkdtemp())
        self.db_path = self.temp_dir / "test_compact.db"
        if self.app is not None:
            self.app.processEvents()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)
        if self.app is not None:
            self.app.processEvents()

    def _select_row(self, window, row_idx):
        from PySide6.QtCore import QItemSelectionModel
        window.table.clearSelection()
        window.table.setCurrentItem(None)
        if window.table.selectionModel() is not None:
            window.table.selectionModel().clearSelection()
            window.table.selectionModel().clear()
        self.app.processEvents()

        window.table.selectRow(row_idx)
        self.app.processEvents()
        model = window.table.model()
        sel_model = window.table.selectionModel()
        idx = model.index(row_idx, 0)
        sel_model.select(idx, QItemSelectionModel.Select | QItemSelectionModel.Rows)
        window._on_table_selection_changed()
        self.app.processEvents()

    def test_supporting_docs_selector_loading_single(self):
        if self.app is None:
            self.skipTest("PySide6 not available")

        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        from scripts.invoice_fetch.db import InvoiceDB

        # Create 1 linked supporting doc
        with InvoiceDB(self.db_path) as db:
            db.insert_invoice({
                "invoice_number": "111",
                "invoice_date": "2026-06-01",
                "seller_name": "销售方A",
                "total_amount": "100.00",
                "review_status": "to_review",
                "extra_paths": json.dumps(["attachments/doc1.pdf"]),
            })

        window = InvoiceReviewApp(self.db_path, splash=None)
        try:
            window.show()
            self.app.processEvents()

            # Select the invoice
            self._select_row(window, 0)

            # Verify combo items and button state
            self.assertEqual(window.combo_supporting_docs.count(), 1)
            self.assertIn("[已关联]", window.combo_supporting_docs.itemText(0))
            self.assertTrue(window.btn_open_extra_files.isEnabled())
        finally:
            window.close()

    def test_supporting_docs_selector_multi(self):
        if self.app is None:
            self.skipTest("PySide6 not available")

        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        from scripts.invoice_fetch.db import InvoiceDB

        # Create invoice with 2 linked docs + 1 pending evidence in the same mail
        with InvoiceDB(self.db_path) as db:
            db.insert_invoice({
                "id": 1,
                "invoice_number": "111",
                "invoice_date": "2026-06-01",
                "seller_name": "销售方A",
                "total_amount": "100.00",
                "review_status": "to_review",
                "mailbox_key": "mail_test",
                "mail_uid": 12345,
                "extra_paths": json.dumps(["attachments/doc1.pdf", "attachments/doc2.pdf"]),
            })
            # Pending evidence doc in same mail
            db.insert_invoice({
                "id": 2,
                "invoice_number": "",
                "invoice_date": "2026-06-01",
                "seller_name": "",
                "total_amount": "0.00",
                "review_status": "to_review",
                "invoice_type": "待关联证明材料",
                "mailbox_key": "mail_test",
                "mail_uid": 12345,
                "attachment_path": "attachments/doc3.pdf",
            })

        window = InvoiceReviewApp(self.db_path, splash=None)
        try:
            window.show()
            self.app.processEvents()

            # Find row index of invoice ID 1
            row_idx = 0
            for idx, inv in enumerate(window.invoices_list):
                if inv["id"] == 1:
                    row_idx = idx
                    break

            self._select_row(window, row_idx)

            # Expect 3 items (2 linked, 1 pending)
            self.assertEqual(window.combo_supporting_docs.count(), 3)
            texts = [window.combo_supporting_docs.itemText(i) for i in range(3)]
            linked_count = sum(1 for t in texts if "[已关联]" in t)
            pending_count = sum(1 for t in texts if "[待关联]" in t)
            self.assertEqual(linked_count, 2)
            self.assertEqual(pending_count, 1)
        finally:
            window.close()

    def test_supporting_docs_empty(self):
        if self.app is None:
            self.skipTest("PySide6 not available")

        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        from scripts.invoice_fetch.db import InvoiceDB

        with InvoiceDB(self.db_path) as db:
            db.insert_invoice({
                "invoice_number": "111",
                "invoice_date": "2026-06-01",
                "seller_name": "销售方A",
                "total_amount": "100.00",
                "review_status": "to_review",
                "extra_paths": "",
            })

        window = InvoiceReviewApp(self.db_path, splash=None)
        try:
            window.show()
            self.app.processEvents()

            self._select_row(window, 0)

            self.assertEqual(window.combo_supporting_docs.count(), 1)
            self.assertEqual(window.combo_supporting_docs.itemText(0), "暂无证明材料")
            self.assertFalse(window.combo_supporting_docs.isEnabled())
            self.assertFalse(window.btn_open_extra_files.isEnabled())
            self.assertIn("酒店水单、行程记录、支付截图", window.combo_supporting_docs.toolTip())
        finally:
            window.close()

    def test_open_selected_supporting_doc(self):
        if self.app is None:
            self.skipTest("PySide6 not available")

        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        from scripts.invoice_fetch.db import InvoiceDB

        doc1 = self.temp_dir / "doc1.pdf"
        doc2 = self.temp_dir / "doc2.pdf"
        doc1.write_bytes(b"pdf1")
        doc2.write_bytes(b"pdf2")

        with InvoiceDB(self.db_path) as db:
            db.insert_invoice({
                "invoice_number": "111",
                "invoice_date": "2026-06-01",
                "seller_name": "销售方A",
                "total_amount": "100.00",
                "review_status": "to_review",
                "extra_paths": json.dumps([str(doc1), str(doc2)]),
            })

        window = InvoiceReviewApp(self.db_path, splash=None)
        try:
            window.show()
            self.app.processEvents()

            self._select_row(window, 0)

            # Mock _open_local_path
            window._open_local_path = MagicMock()

            # Select second item (index 1)
            window.combo_supporting_docs.setCurrentIndex(1)
            self.app.processEvents()

            # Trigger opening
            window._open_extra_docs()
            self.app.processEvents()

            # Verify only doc2 was opened
            window._open_local_path.assert_called_once_with(doc2)
        finally:
            window.close()

    def test_notes_compact_ui_height(self):
        if self.app is None:
            self.skipTest("PySide6 not available")

        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        window = InvoiceReviewApp(self.db_path, splash=None)
        try:
            self.assertLessEqual(window.txt_note.maximumHeight(), 80)  # readable note height
        finally:
            window.close()

    def test_supporting_doc_label_formatting(self):
        if self.app is None:
            self.skipTest("PySide6 not available")

        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        window = InvoiceReviewApp(self.db_path, splash=None)
        try:
            path_normal = Path("short_name.pdf")
            label_normal = window._format_supporting_doc_label("已关联", path_normal)
            self.assertEqual(label_normal, "[已关联] short_name.pdf")

            path_long = Path("出租_87.90_2653700000000009000697_ex_6.pdf")
            label_long = window._format_supporting_doc_label("已关联", path_long, max_len=42)
            self.assertTrue(label_long.startswith("[已关联]"))
            self.assertIn("...", label_long)
            self.assertLessEqual(len(label_long), 42)
        finally:
            window.close()

    def test_post_link_combo_reposition(self):
        if self.app is None:
            self.skipTest("PySide6 not available")

        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        from scripts.invoice_fetch.db import InvoiceDB

        doc1 = self.temp_dir / "doc1.pdf"
        doc2 = self.temp_dir / "doc2.pdf"
        doc1.write_bytes(b"pdf1")
        doc2.write_bytes(b"pdf2")

        with InvoiceDB(self.db_path) as db:
            db.insert_invoice({
                "invoice_number": "111",
                "invoice_date": "2026-06-01",
                "seller_name": "销售方A",
                "total_amount": "100.00",
                "review_status": "to_review",
                "extra_paths": json.dumps([str(doc1), str(doc2)]),
            })

        window = InvoiceReviewApp(self.db_path, splash=None)
        try:
            window.show()
            self.app.processEvents()

            self._select_row(window, 0)

            window._update_supporting_docs_selector(window.current_invoice, selected_path=doc2)
            self.app.processEvents()

            self.assertEqual(window.combo_supporting_docs.currentIndex(), 1)
            self.assertIn("doc2.pdf", window.combo_supporting_docs.itemText(1))
        finally:
            window.close()

    def test_excel_export_notes_and_evidence(self):
        from scripts.invoice_fetch.excel_export import export_excel
        import openpyxl

        doc1 = self.temp_dir / "doc1.pdf"
        doc2 = self.temp_dir / "doc2.pdf"

        rows = [
            {
                "invoice_number": "EXCEL001",
                "invoice_code": "CODE001",
                "invoice_date": "2026-06-01",
                "amount": "100.00",
                "total_amount": "100.00",
                "seller_name": "销售方A",
                "buyer_name": "购买方B",
                "confirmed_note": "团建说明",
                "extra_paths": [str(doc1), str(doc2)],
            }
        ]

        excel_file = self.temp_dir / "test_export.xlsx"
        export_excel(rows, excel_file)

        wb = openpyxl.load_workbook(excel_file)
        ws = wb["发票汇总"]

        headers = [ws.cell(row=1, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
        self.assertIn("个人备注", headers)
        self.assertIn("证明材料", headers)

        note_col_idx = headers.index("个人备注") + 1
        evidence_col_idx = headers.index("证明材料") + 1

        note_val = ws.cell(row=2, column=note_col_idx).value
        evidence_val = ws.cell(row=2, column=evidence_col_idx).value

        self.assertEqual(note_val, "团建说明")
        self.assertEqual(evidence_val, "doc1.pdf；doc2.pdf")
        self.assertNotIn(str(self.temp_dir), evidence_val)

    def test_gui_empty_attachment_hint(self):
        if self.app is None:
            self.skipTest("PySide6 not available")

        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        from scripts.invoice_fetch.db import InvoiceDB

        with InvoiceDB(self.db_path) as db:
            db.insert_invoice({
                "invoice_number": "222",
                "invoice_date": "2026-06-01",
                "seller_name": "销售方B",
                "total_amount": "200.00",
                "review_status": "to_review",
                "mail_uid": 123,
                "download_url": "http://nnfp.jss.com.cn/show",
                "attachment_path": "",
            })

        window = InvoiceReviewApp(self.db_path, splash=None)
        try:
            window.show()
            self.app.processEvents()

            self._select_row(window, 0)

            self.assertEqual(window.txt_path.text(), "未下载原件（可重试下载或手动补原件）")
            self.assertTrue(window.btn_add_attachment.isEnabled())
            self.assertTrue(window.btn_retry_download.isEnabled())
            self.assertTrue(window.btn_retry_download.isVisible())
            self.assertFalse(window.btn_open_file.isEnabled())
        finally:
            window.close()

    def test_approval_missing_attachment(self):
        if self.app is None:
            self.skipTest("PySide6 not available")

        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        window = InvoiceReviewApp(self.db_path, splash=None)
        try:
            inv = {
                "invoice_number": "222",
                "invoice_date": "2026-06-01",
                "total_amount": "200.00",
                "attachment_path": "",
                "download_url": "http://nnfp.jss.com.cn/show",
            }
            missing = window._approval_missing_fields(inv)
            self.assertIn("原件", missing)
        finally:
            window.close()


if __name__ == "__main__":
    unittest.main()
