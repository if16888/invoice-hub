import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from scripts.invoice_fetch.claim_export import export_claim_package
from scripts.invoice_fetch.db import InvoiceDB
from scripts.invoice_fetch.reimbursement import buyer_warning


class ReimbursementValidationTests(unittest.TestCase):
    def test_buyer_warning_matches_missing_and_mismatch(self):
        cfg = {
            "reimbursement": {
                "buyer_name": "示例科技有限公司",
                "strict_buyer_check": True,
            }
        }

        self.assertEqual(buyer_warning({"buyer_name": "示例科技有限公司"}, cfg), "")
        self.assertEqual(buyer_warning({"buyer_name": ""}, cfg), "购方抬头待核对")
        self.assertEqual(
            buyer_warning({"buyer_name": "其他公司"}, cfg),
            "购买方与默认开票主体不一致：当前发票：其他公司；默认主体：示例科技有限公司",
        )

    def test_buyer_warning_without_expected_preset_only_warns_when_missing(self):
        cfg = {"reimbursement": {"strict_buyer_check": True}}

        self.assertEqual(buyer_warning({"buyer_name": "任意公司"}, cfg), "")
        self.assertEqual(buyer_warning({"buyer_name": ""}, cfg), "购方抬头待核对")

    def test_claim_export_records_buyer_warning_in_manifest_and_excel(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            runtime = root / "runtime"
            attachment = runtime / "attachments" / "invoice.pdf"
            attachment.parent.mkdir(parents=True)
            attachment.write_bytes(b"pdf")
            db_path = runtime / "invoices.db"

            with InvoiceDB(db_path) as db:
                invoice_id = db.insert_invoice({
                    "invoice_number": "BUYER001",
                    "invoice_date": "2026-05-30",
                    "total_amount": "18.00",
                    "seller_name": "Seller",
                    "buyer_name": "其他公司",
                    "category": "餐饮",
                    "attachment_path": "attachments/invoice.pdf",
                    "review_status": "approved",
                })
                claim_id = db.create_claim_group("buyer warning claim")
                db.add_invoice_to_claim(claim_id, invoice_id)

                export_dir = export_claim_package(
                    db,
                    claim_id,
                    project_root=root,
                    runtime_dir=runtime,
                    reimbursement_config={
                        "buyer_name": "示例科技有限公司",
                        "strict_buyer_check": True,
                    },
                )

            manifest = json.loads((export_dir / "manifest.json").read_text(encoding="utf-8"))
            warning = "购买方与默认开票主体不一致：当前发票：其他公司；默认主体：示例科技有限公司"
            self.assertEqual(manifest["items"][0]["warning"], warning)

            wb = load_workbook(export_dir / "reimbursement.xlsx")
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            self.assertIn("校验提示", headers)
            warning_col = headers.index("校验提示") + 1
            self.assertEqual(ws.cell(row=2, column=warning_col).value, warning)


if __name__ == "__main__":
    unittest.main()
