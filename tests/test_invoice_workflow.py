import email
import json
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest.mock import Mock, patch

import requests
from openpyxl import load_workbook

from scripts.invoice_fetch import __main__ as cli
from scripts.invoice_fetch import link_downloader as link_dl
from scripts.invoice_fetch.ai_classifier import AIClassifier
from scripts.invoice_fetch.attachment_handler import Attachment
from scripts.invoice_fetch.attachment_handler import AttachmentHandler
from scripts.invoice_fetch.config import load_config
from scripts.invoice_fetch.db import InvoiceDB
from scripts.invoice_fetch.excel_export import export_excel
from scripts.invoice_fetch.invoice_parser import InvoiceInfo, InvoiceParser
from scripts.invoice_fetch.link_downloader import DownloadedFile, LinkDownloader
from scripts.invoice_fetch.rule_classifier import classify as rule_classify


class FakeDB:
    def __init__(self):
        self.marked = []

    def mark_downloaded(self, uid):
        self.marked.append(uid)


class FakeFetcher:
    def fetch_by_uid(self, uid, folder="INBOX"):
        msg = email.message.EmailMessage()
        msg["Subject"] = "发票"
        msg["From"] = "billing@example.com"
        msg["Date"] = "Mon, 18 May 2026 10:00:00 +0800"
        return cli.MailMessage(uid=uid, raw_msg=msg)


class NoopLinkDownloader:
    def download_from_email(self, *args, **kwargs):
        return []


class StaticAttachmentHandler:
    def __init__(self, base, attachments):
        self._base = Path(base)
        self._attachments = attachments

    def extract(self, *args, **kwargs):
        return self._attachments


class StaticParser:
    def __init__(self, info):
        self._info = info

    def parse_pdf(self, path):
        return self._info


class MultiLinkDownloader(LinkDownloader):
    def __init__(self):
        super().__init__(tempfile.mkdtemp())

    def _download_url(self, url, mail_uid, idx, date_str, disable_fallback=False):
        return DownloadedFile(
            url=url,
            file_path=f"/tmp/invoice_{idx}.pdf",
            filename=f"invoice_{idx}.pdf",
            size=1024,
            is_invoice=True,
        )


class _FakePdfPage:
    def __init__(self, text):
        self._text = text

    def extract_text(self):
        return self._text


class _FakePdfDocument:
    def __init__(self, text):
        self.pages = [_FakePdfPage(text)]

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        return False


class _FakePdfPlumber:
    def __init__(self, text):
        self._text = text

    def open(self, path):
        return _FakePdfDocument(self._text)


class InvoiceWorkflowTests(unittest.TestCase):
    def test_import_local_directory_records_parsed_pdf_without_moving_source(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            import_dir = base / "local_invoices"
            import_dir.mkdir()
            src = import_dir / "train_invoice.pdf"
            src.write_bytes(b"%PDF- local")
            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                count = cli._import_local_directory(
                    import_dir=import_dir,
                    db=db,
                    parser=StaticParser(InvoiceInfo(
                        invoice_number="12345678",
                        invoice_code="031002500111",
                        invoice_date="2026-05-18",
                        total_amount="42.00",
                        seller_name="12306",
                        invoice_type="电子发票",
                        parse_success=True,
                    )),
                    categories={"transport": {"keywords": ["12306"], "extra_name": ""}},
                    att_dir=runtime / "attachments",
                )
                rows = db.get_all_invoices()

            self.assertEqual(count.get("added"), 1)
            self.assertTrue(src.exists())
            self.assertEqual(rows[0]["invoice_number"], "12345678")
            self.assertEqual(rows[0]["mail_subject"], "本地导入: train_invoice.pdf")
            self.assertTrue((runtime / rows[0]["attachment_path"]).exists())

    def test_import_local_directory_updates_existing_invoice_metadata_on_reimport(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            import_dir = base / "local_invoices"
            import_dir.mkdir()
            src = import_dir / "train_invoice.pdf"
            src.write_bytes(b"%PDF- local")
            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                count1 = cli._import_local_directory(
                    import_dir=import_dir,
                    db=db,
                    parser=StaticParser(InvoiceInfo(
                        invoice_number="12345678",
                        invoice_code="031002500111",
                        invoice_date="2026-05-18",
                        total_amount="42.00",
                        seller_name="旧销售方",
                        invoice_type="电子发票",
                        parse_success=True,
                    )),
                    categories={"transport": {"keywords": ["12306"], "extra_name": ""}},
                    att_dir=runtime / "attachments",
                )
                # Write different bytes to simulate a different file (different file_hash) with conflicting metadata
                src.write_bytes(b"%PDF- local diff bytes for conflict")
                count2 = cli._import_local_directory(
                    import_dir=import_dir,
                    db=db,
                    parser=StaticParser(InvoiceInfo(
                        invoice_number="12345678",
                        invoice_code="031002500111",
                        invoice_date="2026-05-18",
                        total_amount="42.00",
                        seller_name="新销售方",
                        invoice_type="电子发票",
                        parse_success=True,
                    )),
                    categories={"transport": {"keywords": ["12306"], "extra_name": ""}},
                    att_dir=runtime / "attachments",
                )
                rows = db.get_all_invoices()

            self.assertEqual(count1.get("added"), 1)
            self.assertEqual(count2.get("conflicts"), 1)
            self.assertEqual(len(rows), 2)
            self.assertEqual(rows[0]["seller_name"], "新销售方")
            self.assertEqual(rows[0]["invoice_type"], "本地导入冲突")
            self.assertEqual(rows[0]["review_status"], "error")
            self.assertEqual(rows[1]["seller_name"], "旧销售方")

    def test_import_local_directory_reuses_runtime_files_without_copying_again(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td).resolve()
            runtime = base / "runtime"
            attachments_root = runtime / "attachments"
            invoice_dir = attachments_root / "2026-05-18"
            invoice_dir.mkdir(parents=True)
            src = invoice_dir / "runtime_invoice.pdf"
            src.write_bytes(b"%PDF- runtime")

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                count1 = cli._import_local_directory(
                    import_dir=attachments_root,
                    db=db,
                    parser=StaticParser(InvoiceInfo(
                        invoice_number="77777777",
                        invoice_code="031002500111",
                        invoice_date="2026-05-18",
                        total_amount="42.00",
                        seller_name="旧销售方",
                        invoice_type="电子发票",
                        parse_success=True,
                    )),
                    categories={"transport": {"keywords": ["12306"], "extra_name": ""}},
                    att_dir=attachments_root,
                )
                # Write different bytes to simulate a different file (different file_hash) with conflicting metadata
                src.write_bytes(b"%PDF- runtime diff bytes for conflict")
                count2 = cli._import_local_directory(
                    import_dir=attachments_root,
                    db=db,
                    parser=StaticParser(InvoiceInfo(
                        invoice_number="77777777",
                        invoice_code="031002500111",
                        invoice_date="2026-05-18",
                        total_amount="42.00",
                        seller_name="新销售方",
                        invoice_type="电子发票",
                        parse_success=True,
                    )),
                    categories={"transport": {"keywords": ["12306"], "extra_name": ""}},
                    att_dir=attachments_root,
                )
                rows = db.get_all_invoices()

            self.assertEqual(count1.get("added"), 1)
            self.assertEqual(count2.get("conflicts"), 1)
            self.assertEqual(len(rows), 2)
            self.assertEqual(rows[0]["seller_name"], "新销售方")
            self.assertEqual(rows[0]["invoice_type"], "本地导入冲突")
            self.assertEqual(rows[0]["review_status"], "error")
            self.assertEqual(rows[1]["seller_name"], "旧销售方")
            self.assertEqual(Path(rows[0]["attachment_path"]).as_posix(), "attachments/2026-05-18/runtime_invoice.pdf")
            self.assertEqual(sum(1 for _ in attachments_root.rglob("runtime_invoice*.pdf")), 1)

    def test_import_local_directory_restores_soft_deleted_duplicate_pdf(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            import_dir = base / "local_invoices"
            import_dir.mkdir()
            src = import_dir / "train_invoice.pdf"
            src.write_bytes(b"%PDF- local")
            info = InvoiceInfo(
                invoice_number="12345678",
                invoice_code="031002500111",
                invoice_date="2026-05-18",
                total_amount="42.00",
                seller_name="Synthetic Seller",
                buyer_name="Synthetic Buyer",
                invoice_type="电子发票",
                parse_success=True,
            )

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                count1 = cli._import_local_directory(
                    import_dir=import_dir,
                    db=db,
                    parser=StaticParser(info),
                    categories={},
                    att_dir=runtime / "attachments",
                )
                row_id = db.get_all_invoices()[0]["id"]
                self.assertTrue(db.soft_delete_invoice(row_id))

                count2 = cli._import_local_directory(
                    import_dir=import_dir,
                    db=db,
                    parser=StaticParser(info),
                    categories={},
                    att_dir=runtime / "attachments",
                )
                rows = db.get_all_invoices()

            self.assertEqual(count1.get("added"), 1)
            self.assertEqual(count2.get("added"), 1)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["invoice_number"], "12345678")
            self.assertEqual(rows[0]["is_deleted"], 0)
            self.assertTrue(rows[0]["attachment_path"])

    def test_import_local_directory_does_not_restore_soft_deleted_different_seller_conflict(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            import_dir = base / "local_invoices"
            import_dir.mkdir()
            src = import_dir / "train_invoice.pdf"
            src.write_bytes(b"%PDF- local different seller")
            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                row_id = db.insert_invoice({
                    "invoice_number": "12345678",
                    "invoice_code": "031002500111",
                    "invoice_date": "2026-05-18",
                    "total_amount": "42.00",
                    "seller_name": "Old Synthetic Seller",
                    "buyer_name": "Synthetic Buyer",
                    "invoice_type": "电子发票",
                    "attachment_path": "",
                })
                self.assertTrue(db.soft_delete_invoice(row_id))

                count = cli._import_local_directory(
                    import_dir=import_dir,
                    db=db,
                    parser=StaticParser(InvoiceInfo(
                        invoice_number="12345678",
                        invoice_code="031002500111",
                        invoice_date="2026-05-18",
                        total_amount="42.00",
                        seller_name="New Synthetic Seller",
                        buyer_name="Synthetic Buyer",
                        invoice_type="电子发票",
                        parse_success=True,
                    )),
                    categories={},
                    att_dir=runtime / "attachments",
                )
                visible_rows = db.get_all_invoices()
                all_rows = db.get_all_invoices(include_deleted=True)

            self.assertEqual(count.get("added"), 1)
            self.assertEqual(len(visible_rows), 1)
            self.assertEqual(visible_rows[0]["seller_name"], "New Synthetic Seller")
            deleted_rows = [row for row in all_rows if row["seller_name"] == "Old Synthetic Seller"]
            self.assertEqual(len(deleted_rows), 1)
            self.assertEqual(deleted_rows[0]["is_deleted"], 1)

    def test_import_local_directory_skips_duplicate_file_hash_for_unparsed_files(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            import_dir = base / "local_invoices"
            import_dir.mkdir()
            (import_dir / "receipt-a.jpg").write_bytes(b"same image bytes")
            (import_dir / "receipt-b.jpg").write_bytes(b"same image bytes")

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                count = cli._import_local_directory(
                    import_dir=import_dir,
                    db=db,
                    parser=InvoiceParser(),
                    categories={},
                    att_dir=runtime / "attachments",
                )
                rows = db.get_all_invoices()

            self.assertEqual(count.get("added"), 0)
            self.assertEqual(count.get("pending_manual"), 1)
            self.assertEqual(count.get("failed"), 0)
            self.assertEqual(count.get("duplicates"), 1)
            self.assertEqual(len(rows), 1)
            self.assertTrue(rows[0]["file_hash"])

    def test_local_evidence_matching_existing_invoice_attaches_without_new_row(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            import_dir = base / "local_evidence"
            import_dir.mkdir()
            evidence = import_dir / "电子票据行程单_05879011.pdf"
            evidence.write_bytes(b"%PDF- synthetic trip evidence")
            info = InvoiceInfo(
                invoice_number="05879011",
                invoice_type="电子票据行程单",
                parse_success=True,
                raw_text="江苏省车辆通行费电子票据行程单 发票号码 05879011",
            )

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                invoice_id = db.insert_invoice({
                    "invoice_number": "05879011",
                    "invoice_date": "2026-06-01",
                    "total_amount": "30.00",
                    "seller_name": "示例道路服务有限公司",
                    "invoice_type": "电子发票",
                    "review_status": "to_review",
                    "extra_paths": [],
                })
                stats = cli._import_local_directory(
                    import_dir=import_dir,
                    db=db,
                    parser=StaticParser(info),
                    categories={},
                    att_dir=runtime / "attachments",
                )
                rows = db.get_all_invoices()
                updated = db.get_invoice(invoice_id)

            self.assertEqual(stats["added"], 1)
            self.assertEqual(len(rows), 1)
            extra_paths = json.loads(updated["extra_paths"])
            self.assertEqual(len(extra_paths), 1)
            self.assertTrue((runtime / extra_paths[0]).exists())
            self.assertTrue(updated["has_extra"])
            self.assertFalse(updated["missing_extra"])

    def test_local_unmatched_evidence_is_kept_as_pending_link_record(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            import_dir = base / "local_evidence"
            import_dir.mkdir()
            evidence = import_dir / "电子票据行程单_待关联.pdf"
            evidence.write_bytes(b"%PDF- synthetic unmatched evidence")
            info = InvoiceInfo(
                invoice_type="电子票据行程单",
                parse_success=True,
                raw_text="江苏省车辆通行费电子票据行程单 2026-06-01 合计 35 元",
            )

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                stats = cli._import_local_directory(
                    import_dir=import_dir,
                    db=db,
                    parser=StaticParser(info),
                    categories={},
                    att_dir=runtime / "attachments",
                )
                rows = db.get_all_invoices()

            self.assertEqual(stats["pending_manual"], 1)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["invoice_type"], "待关联证明材料")
            self.assertIn(
                "发现疑似证明材料，但没有唯一匹配的主发票，请人工关联",
                rows[0]["parse_note"],
            )
            self.assertIn(rows[0]["review_status"], ("to_review", "error"))

    def test_reimported_evidence_does_not_duplicate_extra_paths(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            import_dir = base / "local_evidence"
            import_dir.mkdir()
            evidence = import_dir / "通行费行程单_05879011.pdf"
            evidence.write_bytes(b"%PDF- same synthetic evidence")
            info = InvoiceInfo(
                invoice_number="05879011",
                invoice_type="通行费行程单",
                parse_success=True,
                raw_text="通行费行程单 05879011",
            )

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                invoice_id = db.insert_invoice({
                    "invoice_number": "05879011",
                    "total_amount": "30.00",
                    "seller_name": "示例道路服务有限公司",
                    "invoice_type": "电子发票",
                    "extra_paths": [],
                })
                first = cli._import_local_directory(
                    import_dir, db, StaticParser(info), {}, runtime / "attachments"
                )
                second = cli._import_local_directory(
                    import_dir, db, StaticParser(info), {}, runtime / "attachments"
                )
                updated = db.get_invoice(invoice_id)

            self.assertEqual(first["added"], 1)
            self.assertEqual(second["duplicates"], 1)
            self.assertEqual(len(json.loads(updated["extra_paths"])), 1)

    def test_standard_electronic_invoice_is_not_treated_as_evidence(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            import_dir = base / "local_invoices"
            import_dir.mkdir()
            invoice = import_dir / "电子发票_05879012.pdf"
            invoice.write_bytes(b"%PDF- synthetic invoice")
            info = InvoiceInfo(
                invoice_number="05879012",
                invoice_date="2026-06-01",
                total_amount="30.00",
                seller_name="示例科技有限公司",
                invoice_type="增值税电子普通发票",
                parse_success=True,
                raw_text="增值税电子普通发票 发票号码 05879012",
            )

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                stats = cli._import_local_directory(
                    import_dir, db, StaticParser(info), {}, runtime / "attachments"
                )
                rows = db.get_all_invoices()

            self.assertEqual(stats["added"], 1)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["invoice_type"], "增值税电子普通发票")
            self.assertEqual(json.loads(rows[0]["extra_paths"]), [])

    def test_local_payment_screenshot_filename_attaches_to_existing_invoice(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            import_dir = base / "local_evidence"
            import_dir.mkdir()
            screenshot = import_dir / "支付截图_05879013.png"
            screenshot.write_bytes(b"synthetic image bytes")

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                invoice_id = db.insert_invoice({
                    "invoice_number": "05879013",
                    "total_amount": "30.00",
                    "seller_name": "示例科技有限公司",
                    "invoice_type": "电子发票",
                    "extra_paths": [],
                })
                stats = cli._import_local_directory(
                    import_dir, db, InvoiceParser(), {}, runtime / "attachments"
                )
                rows = db.get_all_invoices()
                updated = db.get_invoice(invoice_id)

            self.assertEqual(stats["added"], 1)
            self.assertEqual(len(rows), 1)
            self.assertEqual(len(json.loads(updated["extra_paths"])), 1)

    def test_import_local_directory_records_unparsed_ofd_as_exception(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            import_dir = base / "local_invoices"
            import_dir.mkdir()
            src = import_dir / "invoice.ofd"
            src.write_bytes(b"OFD")
            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                count = cli._import_local_directory(
                    import_dir=import_dir,
                    db=db,
                    parser=StaticParser(InvoiceInfo(parse_success=False)),
                    categories={},
                    att_dir=runtime / "attachments",
                )
                rows = db.get_all_invoices()

            self.assertEqual(count.get("pending_manual"), 1)
            self.assertEqual(count.get("failed"), 0)
            self.assertEqual(rows[0]["parse_success"], 0)
            self.assertIn("OFD", rows[0]["parse_note"])
            self.assertTrue(src.exists())

    def test_attachment_handler_extracts_single_part_pdf_attachment(self):
        msg = email.message.Message()
        msg["Content-Type"] = 'application/pdf; name="invoice.pdf"'
        msg["Content-Disposition"] = 'attachment; filename="invoice.pdf"'
        msg.set_payload(b"%PDF- invoice")

        with tempfile.TemporaryDirectory() as td:
            handler = AttachmentHandler(Path(td))
            attachments = handler.extract(msg, 7, "2026-05-18")

        self.assertEqual(len(attachments), 1)
        self.assertTrue(attachments[0].file_path.endswith("invoice.pdf"))

    def test_attachment_handler_rejects_executable_disguised_as_pdf(self):
        msg = email.message.Message()
        msg["Content-Type"] = 'application/pdf; name="invoice.pdf"'
        msg["Content-Disposition"] = 'attachment; filename="invoice.pdf"'
        msg.set_payload(b"MZ" + b"\x00" * 256)

        with tempfile.TemporaryDirectory() as td:
            out_dir = Path(td) / "out"
            handler = AttachmentHandler(out_dir)
            attachments = handler.extract(msg, 7, "2026-05-18")

            self.assertEqual(attachments, [])
            self.assertEqual(list(out_dir.rglob("*.pdf")), [])

    def test_attachment_handler_unzips_invoice_files_inside_zip(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            zip_path = base / "bundle.zip"
            inner_pdf = b"%PDF- invoice"
            with zipfile.ZipFile(zip_path, "w") as zf:
                zf.writestr("invoice.pdf", inner_pdf)
                zf.writestr("readme.txt", "ignore me")

            msg = email.message.Message()
            msg["Content-Type"] = 'application/zip; name="bundle.zip"'
            msg["Content-Disposition"] = 'attachment; filename="bundle.zip"'
            msg.set_payload(zip_path.read_bytes())

            handler = AttachmentHandler(base / "out")
            attachments = handler.extract(msg, 8, "2026-05-18")
            self.assertEqual(len(attachments), 1)
            self.assertTrue(attachments[0].file_path.endswith(".pdf"))
            self.assertTrue(Path(attachments[0].file_path).exists())

    def test_attachment_handler_rejects_disguised_file_inside_zip(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            zip_path = base / "bundle.zip"
            with zipfile.ZipFile(zip_path, "w") as zf:
                zf.writestr("invoice.pdf", b"MZ" + b"\x00" * 256)

            msg = email.message.Message()
            msg["Content-Type"] = 'application/zip; name="bundle.zip"'
            msg["Content-Disposition"] = 'attachment; filename="bundle.zip"'
            msg.set_payload(zip_path.read_bytes())

            out_dir = base / "out"
            handler = AttachmentHandler(out_dir)
            attachments = handler.extract(msg, 8, "2026-05-18")

            self.assertEqual(attachments, [])
            self.assertEqual(list(out_dir.rglob("*.pdf")), [])

    def test_invoice_parser_prefers_sales_section_over_generic_company_fallback(self):
        text = "\n".join([
            "电子发票（普通发票）",
            "销售方信息",
            "南京市秦淮区福瑞得餐饮管理合伙企业（有限合伙）",
            "统一社会信用代码/纳税人识别号: 91320104MAEAX8RN3T",
            "购买方名称: 远景智能零碳（江苏）科技有限公司",
            "统一社会信用代码/纳税人识别号: 91320292MA1YQ1NN9R",
            "发票号码: 2532200000445223017",
            "开票日期: 2025年09月24日",
            "合计 ¥30.00",
        ])

        parser = InvoiceParser()
        fake_plumber = _FakePdfPlumber(text)

        with tempfile.TemporaryDirectory() as td, patch.object(parser, "_plumber", return_value=fake_plumber):
            pdf_path = Path(td) / "invoice.pdf"
            pdf_path.write_bytes(b"%PDF-1.4 synthetic")
            info = parser.parse_pdf(str(pdf_path))

        self.assertTrue(info.parse_success)
        self.assertEqual(info.seller_name, "南京市秦淮区福瑞得餐饮管理合伙企业（有限合伙）")
        self.assertEqual(info.buyer_name, "远景智能零碳（江苏）科技有限公司")
        self.assertEqual(info.invoice_number, "2532200000445223017")
        self.assertEqual(info.invoice_date, "2025-09-24")

    def test_invoice_parser_handles_name_line_inside_sales_section(self):
        text = "\n".join([
            "电子发票（普通发票）",
            "销售方信息",
            "名称: 南京市秦淮区福瑞得餐饮管理合伙企业（有限合伙）",
            "统一社会信用代码/纳税人识别号: 91320104MAEAX8RN3T",
            "购买方名称: 远景智能零碳（江苏）科技有限公司",
            "统一社会信用代码/纳税人识别号: 91320292MA1YQ1NN9R",
            "发票号码: 2532200000445223017",
            "开票日期: 2025年09月24日",
            "合计 ¥30.00",
        ])

        parser = InvoiceParser()
        fake_plumber = _FakePdfPlumber(text)

        with tempfile.TemporaryDirectory() as td, patch.object(parser, "_plumber", return_value=fake_plumber):
            pdf_path = Path(td) / "invoice.pdf"
            pdf_path.write_bytes(b"%PDF-1.4 synthetic")
            info = parser.parse_pdf(str(pdf_path))

        self.assertTrue(info.parse_success)
        self.assertEqual(info.seller_name, "南京市秦淮区福瑞得餐饮管理合伙企业（有限合伙）")
        self.assertEqual(info.buyer_name, "远景智能零碳（江苏）科技有限公司")

    def test_invoice_parser_ignores_stamp_noise_in_sales_section(self):
        text = "\n".join([
            "电子发票（普通发票）",
            "购买方信息",
            "名称: 远景智能零碳（江苏）科技有限公司",
            "销售方信息",
            "（章）",
            "销售方名称: 南京市秦淮区福瑞得餐饮管理合伙企业（有限合伙）",
            "统一社会信用代码/纳税人识别号: 91320104MAEAX8RN3T",
            "发票号码: 2532200000445223017",
            "开票日期: 2025年09月24日",
            "合计 ¥30.00",
        ])

        parser = InvoiceParser()
        fake_plumber = _FakePdfPlumber(text)

        with tempfile.TemporaryDirectory() as td, patch.object(parser, "_plumber", return_value=fake_plumber):
            pdf_path = Path(td) / "invoice.pdf"
            pdf_path.write_bytes(b"%PDF-1.4 synthetic")
            info = parser.parse_pdf(str(pdf_path))

        self.assertTrue(info.parse_success)
        self.assertEqual(info.seller_name, "南京市秦淮区福瑞得餐饮管理合伙企业（有限合伙）")
        self.assertEqual(info.buyer_name, "远景智能零碳（江苏）科技有限公司")

    def test_invoice_parser_handles_interleaved_buyer_seller_line(self):
        text = "\n".join([
            "电子发票（普通发票）",
            "发票号码：26322000003477340276",
            "开票日期：2026年05月03日",
            "购 名称：远景智能零碳（江苏）科技有限公司 销 名称：小菜园南京餐饮管理有限责任公司",
            "买 售",
            "方 方",
            "信 统一社会信用代码/纳税人识别号：91320292MA1YQ1NN9R 信 统一社会信用代码/纳税人识别号：91320104MA223LC10G",
            "息 息",
            "项目名称 规格型号 单 位 数 量 单 价 金 额 税率/征收率 税 额",
            "*餐饮服务*餐饮服务 1 159.51 159.51 6% 9.57",
            "合 计 ¥159.51 ¥9.57",
            "价税合计（大写） 壹佰陆拾玖圆零捌分 （小写）¥169.08",
        ])

        parser = InvoiceParser()
        fake_plumber = _FakePdfPlumber(text)

        with tempfile.TemporaryDirectory() as td, patch.object(parser, "_plumber", return_value=fake_plumber):
            pdf_path = Path(td) / "invoice.pdf"
            pdf_path.write_bytes(b"%PDF-1.4 synthetic")
            info = parser.parse_pdf(str(pdf_path))

        self.assertTrue(info.parse_success)
        self.assertEqual(info.seller_name, "小菜园南京餐饮管理有限责任公司")
        self.assertEqual(info.buyer_name, "远景智能零碳（江苏）科技有限公司")

    def test_railway_ticket_parser(self):
        text = "\n".join([
            "电子发票（铁路电子客票）",
            "发票号码:26329116804004609069 江苏省税务局 开票日期:2026年04月30日",
            "南京南 G47 杭州西",
            "￥129.00",
            "票价:",
            "交款人 : 李飞",
        ])
        parser = InvoiceParser()
        fake_plumber = _FakePdfPlumber(text)

        with tempfile.TemporaryDirectory() as td, patch.object(parser, "_plumber", return_value=fake_plumber):
            pdf_path = Path(td) / "rail_ticket.pdf"
            pdf_path.write_bytes(b"%PDF-1.4 synthetic")
            info = parser.parse_pdf(str(pdf_path))

        self.assertTrue(info.parse_success)
        self.assertEqual(info.invoice_number, "26329116804004609069")
        self.assertEqual(info.invoice_date, "2026-04-30")
        self.assertEqual(info.total_amount, "129.00")
        self.assertEqual(info.amount, "129.00")
        self.assertEqual(info.seller_name, "中国国家铁路集团有限公司")
        self.assertEqual(info.buyer_name, "李飞")
        self.assertEqual(info.invoice_type, "铁路电子客票")

        # Verify CLI _classify maps to '交通' when seller is standard railway company
        cat, _, _ = cli._classify("本地导入: rail.pdf", "", "中国国家铁路集团有限公司", {})
        self.assertEqual(cat, "交通")

    def test_railway_ticket_type_is_not_overwritten_by_generic_e_invoice_type(self):
        text = "\n".join([
            "12306 电子发票",
            "发票号码:26329116804004609069 开票日期:2026年04月30日",
            "南京南 G47 杭州西",
            "￥129.00",
        ])
        parser = InvoiceParser()
        fake_plumber = _FakePdfPlumber(text)

        with tempfile.TemporaryDirectory() as td, patch.object(parser, "_plumber", return_value=fake_plumber):
            pdf_path = Path(td) / "rail_ticket.pdf"
            pdf_path.write_bytes(b"%PDF-1.4 synthetic")
            info = parser.parse_pdf(str(pdf_path))

        self.assertEqual(info.invoice_type, "铁路电子客票")

    def test_transport_receipt_parser_extracts_english_and_chinese_fields(self):
        samples = [
            "\n".join([
                "Electronic Receipt",
                "Order Details",
                "Order ID: RIDE-ORDER-20260101",
                "Supplier Order ID: SUPPLIER-001",
                "Ride City: Singapore",
                "Itinerary Info",
                "Time of Departure(Local time): [2026-01-01 09:30]",
                "Drop-off Time(Local time): [2026-01-01 10:00]",
                "Grand Total (Paid via Alipay): SGD 18.40",
            ]),
            "\n".join([
                "电子收据",
                "订单详情",
                "订单号：RIDE-ORDER-20260101",
                "供应商订单号：SUPPLIER-001",
                "用车城市：新加坡",
                "行程信息",
                "用车时间（当地时间）：[2026-01-01 09:30]",
                "下车时间（当地时间）：[2026-01-01 10:00]",
                "总计（通过支付宝支付）：SGD 18.40",
            ]),
        ]

        for text in samples:
            with self.subTest(language=text.splitlines()[0]):
                parser = InvoiceParser()
                fake_plumber = _FakePdfPlumber(text)
                with tempfile.TemporaryDirectory() as td, patch.object(
                    parser, "_plumber", return_value=fake_plumber
                ):
                    pdf_path = Path(td) / "receipt.pdf"
                    pdf_path.write_bytes(b"%PDF-1.4 synthetic")
                    info = parser.parse_pdf(str(pdf_path))

                self.assertTrue(info.parse_success)
                self.assertEqual(info.invoice_number, "RIDE-ORDER-20260101")
                self.assertEqual(info.invoice_date, "2026-01-01")
                self.assertEqual(info.total_amount, "18.40")
                self.assertEqual(info.invoice_type, "网约车电子收据")

    def test_transport_receipt_language_variants_share_one_record(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            import_dir = base / "receipts"
            import_dir.mkdir(parents=True)
            (import_dir / "receipt-en.pdf").write_bytes(b"%PDF English variant")
            (import_dir / "receipt-zh.pdf").write_bytes(b"%PDF Chinese variant")

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                first_stats = cli._import_local_directory(
                    import_dir=import_dir,
                    db=db,
                    parser=StaticParser(InvoiceInfo(
                        invoice_number="RIDE-ORDER-20260101",
                        invoice_date="2026-01-01",
                        total_amount="18.40",
                        invoice_type="网约车电子收据",
                        parse_success=True,
                        parse_note="synthetic English transport receipt",
                    )),
                    categories={},
                    att_dir=runtime / "attachments",
                    file_paths=[import_dir / "receipt-en.pdf"],
                )
                second_stats = cli._import_local_directory(
                    import_dir=import_dir,
                    db=db,
                    parser=StaticParser(InvoiceInfo(
                        invoice_number="RIDE-ORDER-20260101",
                        invoice_date="2026-01-01",
                        total_amount="",
                        invoice_type="网约车电子收据",
                        parse_success=True,
                        parse_note="synthetic Chinese transport receipt",
                    )),
                    categories={},
                    att_dir=runtime / "attachments",
                    file_paths=[import_dir / "receipt-zh.pdf"],
                )
                rows = db.get_all_invoices()

            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["invoice_number"], "RIDE-ORDER-20260101")
            self.assertEqual(rows[0]["invoice_type"], "网约车电子收据")
            self.assertEqual(rows[0]["total_amount"], "18.40")
            self.assertEqual(rows[0]["category"], "出租车")
            self.assertEqual(first_stats["added"], 1)
            self.assertEqual(second_stats["added"], 1)
            self.assertEqual(second_stats["conflicts"], 0)
            self.assertEqual(len(json.loads(rows[0]["extra_paths"])), 1)

    def test_invoice_parser_extracts_seller_from_two_column_telecom_layout(self):
        """Regression: pdfplumber merges buyer/seller columns into one line.

        When the extracted text contains "名称李飞名称中国电信股份有限公司南京分公司"
        the parser must still pick out the seller (second 名称 field).
        """
        text = "\n".join([
            "电子发票（普通发票）",
            "发票号码：26327000000970967926",
            "开票日期：2026年05月23日",
            # Two-column merge: buyer name + seller name on one line
            "名称李飞名称中国电信股份有限公司南京分公司",
            "统一社会信用代码/纳税人识别号统一社会信用代码/纳税人识别号：91320100748211212",
            "地址、电话地址、电话：南京市建邺区江东中路301号",
            "开户行及账号开户行及账号：工商银行南京兴隆支行",
            "项目名称规格型号数量单价金额税率/征收率税额",
            "电信服务*电信服务费项1 945.50 945.50 *",
            "价税合计（小写）¥945.50",
        ])

        parser = InvoiceParser()
        fake_plumber = _FakePdfPlumber(text)

        with tempfile.TemporaryDirectory() as td, patch.object(parser, "_plumber", return_value=fake_plumber):
            pdf_path = Path(td) / "telecom_invoice.pdf"
            pdf_path.write_bytes(b"%PDF-1.4 synthetic")
            info = parser.parse_pdf(str(pdf_path))

        self.assertTrue(info.parse_success)
        self.assertEqual(info.seller_name, "中国电信股份有限公司南京分公司")
        self.assertEqual(info.invoice_number, "26327000000970967926")
        self.assertEqual(info.invoice_date, "2026-05-23")
        self.assertEqual(info.total_amount, "945.50")

    def test_normalize_date_kangxi_radicals(self):
        from scripts.invoice_fetch.invoice_parser import normalize_date
        # Test standard radicals and Kangxi radical variants
        self.assertEqual(normalize_date("2025年12\u2f4926\u2f47"), "2025-12-26")
        self.assertEqual(normalize_date("2026年05\u2f2601\u2f3c"), "2026-05-01")
        self.assertEqual(normalize_date("2024年08\u2f4915\u2f52"), "2024-08-15")

    def test_link_dedup_keeps_distinct_invoice_query_parameters(self):
        links = [
            "https://example.com/download?invoiceId=123",
            "https://example.com/download?invoiceId=456",
        ]

        self.assertEqual(link_dl._dedup_and_prioritize(links), links)

    def test_link_dedup_drops_tracking_query_parameters(self):
        links = [
            "https://example.com/download?invoiceId=123&utm_source=newsletter",
            "https://example.com/download?invoiceId=123&utm_campaign=spring",
        ]

        self.assertEqual(
            link_dl._dedup_and_prioritize(links),
            ["https://example.com/download?invoiceId=123&utm_source=newsletter"],
        )

    def test_rule_classifier_defers_when_invoice_and_exclusion_keywords_conflict(self):
        result, reason = rule_classify("您的航空电子发票及登机凭证", "airline@example.com")

        self.assertEqual(result, -1)
        self.assertIn("冲突", reason)

    def test_rule_classifier_excludes_github_invoice_hub_notification(self):
        result, reason = rule_classify(
            "[if16888/invoice-hub] Workflow run failed",
            "notifications@github.com",
        )

        self.assertEqual(result, 0)
        self.assertIn("技术通知", reason)

    def test_link_downloader_blocks_technical_hosts_with_invoice_text(self):
        html = """
        <html><body>
          <a href="https://github.com/if16888/invoice-hub/actions/runs/123">
            Invoice Hub workflow notification
          </a>
          <a href="https://billing.example.com/download?invoiceId=123">
            下载发票
          </a>
          <a href="https://portal.51fapiao.cn/document/123">查看</a>
          <a href="https://invoice.nuonuo.com/document/456">查看</a>
          <a href="https://billing.baiwang.com/document/789">查看</a>
        </body></html>
        """

        self.assertEqual(
            link_dl.extract_links_from_html(html),
            [
                "https://billing.example.com/download?invoiceId=123",
                "https://portal.51fapiao.cn/document/123",
                "https://invoice.nuonuo.com/document/456",
                "https://billing.baiwang.com/document/789",
            ],
        )

    def test_ai_classifier_keeps_failed_api_batch_unclassified(self):
        ai = object.__new__(AIClassifier)
        ai.provider = "deepseek"
        ai.model = "deepseek-chat"
        ai.batch_size = 20
        ai.api_key = "test-key"

        with patch("scripts.invoice_fetch.ai_classifier.requests.post",
                   side_effect=requests.Timeout("boom")), \
              patch("scripts.invoice_fetch.ai_classifier.time.sleep"):
            results = ai.classify_batch([
                {"uid": 1, "subject": "maybe invoice", "sender": "a@example.com"},
                {"uid": 2, "subject": "maybe invoice too", "sender": "b@example.com"},
            ])

        self.assertEqual([r["uid"] for r in results], [1, 2])
        self.assertTrue(all(r["is_invoice"] is None for r in results))
        self.assertTrue(all("重试" in r["reason"] for r in results))

    def test_run_classify_does_not_turn_ai_failures_into_non_invoice(self):
        class FailedAI:
            def __init__(self, *args, **kwargs):
                pass

            def classify_batch(self, emails):
                return [
                    {"uid": row["uid"], "is_invoice": None, "reason": "AI 分类 API 失败，将在下次运行时重试"}
                    for row in emails
                ]

        with tempfile.TemporaryDirectory() as td:
            db = InvoiceDB(Path(td) / "invoices.db")
            db.upsert_email(42, "plain subject", "sender@example.com", "2026-05-18")
            with patch.object(cli, "AIClassifier", FailedAI):
                cli._run_classify(db, {"provider": "deepseek", "model": "deepseek-chat"}, no_ai=False)
            stats = db.get_email_stats()
            db.close()

        self.assertEqual(stats["unclassified"], 1)
        self.assertEqual(stats["not_invoice"], 0)

    def test_run_classify_skips_ai_when_provider_is_none(self):
        with tempfile.TemporaryDirectory() as td:
            db = InvoiceDB(Path(td) / "invoices.db")
            db.upsert_email(43, "plain subject", "sender@example.com", "2026-05-18")
            with patch("scripts.invoice_fetch.ai_classifier.AIClassifier") as ai_class:
                cli._run_classify(db, {"provider": "none"}, no_ai=False)
            stats = db.get_email_stats()
            db.close()

        ai_class.assert_not_called()
        self.assertEqual(stats["unclassified"], 1)

    def test_run_classify_contains_ai_initialization_failure(self):
        class BrokenAI:
            def __init__(self, *args, **kwargs):
                raise RuntimeError("synthetic AI initialization failure")

        with tempfile.TemporaryDirectory() as td:
            db = InvoiceDB(Path(td) / "invoices.db")
            db.upsert_email(44, "plain subject", "sender@example.com", "2026-05-18")
            with patch("scripts.invoice_fetch.ai_classifier.AIClassifier", BrokenAI):
                cli._run_classify(
                    db,
                    {"provider": "deepseek", "model": "deepseek-chat"},
                    no_ai=False,
                )
            stats = db.get_email_stats()
            db.close()

        self.assertEqual(stats["unclassified"], 1)
        self.assertEqual(stats["not_invoice"], 0)

    def test_load_config_defaults_empty_ai_model_by_provider(self):
        with tempfile.TemporaryDirectory() as td:
            cfg_path = Path(td) / "config.json"
            cfg_path.write_text(json.dumps({
                "email": {"address": "user@example.com"},
                "ai": {"provider": "deepseek", "model": ""},
            }), encoding="utf-8")

            cfg = load_config(cfg_path)

        self.assertEqual(cfg["ai"]["model"], "deepseek-chat")

    def test_load_config_reports_json_syntax_errors_without_traceback(self):
        with tempfile.TemporaryDirectory() as td:
            cfg_path = Path(td) / "config.json"
            cfg_path.write_text('{"email": {"address": "user@example.com",}', encoding="utf-8")

            with self.assertRaises(SystemExit) as cm:
                load_config(cfg_path)

        self.assertIn("配置文件格式错误", str(cm.exception))

    def test_invoice_rename_collision_uses_numbered_filename(self):
        with tempfile.TemporaryDirectory() as td:
            runtime = Path(td) / "runtime"
            att_dir = runtime / "attachments"
            src_dir = runtime / "incoming"
            src_dir.mkdir(parents=True)
            src = src_dir / "new.pdf"
            src.write_bytes(b"new")
            date_dir = att_dir / "2026-05-18"
            date_dir.mkdir(parents=True)
            existing = date_dir / "meal_20_1001.pdf"
            existing.write_bytes(b"old")

            with patch.object(cli, "RUNTIME_DIR", runtime):
                rel = cli._rename_by_invoice_code(
                    str(src), "1001", "2026-05-18", att_dir,
                    category="meal", total_amount="20", invoice_number="1001")

            self.assertEqual(rel, str(Path("attachments") / "2026-05-18" / "meal_20_1001_1.pdf"))
            self.assertTrue((runtime / rel).exists())
            self.assertTrue(existing.exists())

    def test_invoice_rename_sanitizes_date_directory(self):
        with tempfile.TemporaryDirectory() as td:
            runtime = Path(td) / "runtime"
            att_dir = runtime / "attachments"
            src_dir = runtime / "incoming"
            src_dir.mkdir(parents=True)
            src = src_dir / "new.pdf"
            src.write_bytes(b"new")

            with patch.object(cli, "RUNTIME_DIR", runtime):
                rel = cli._rename_by_invoice_code(
                    str(src), "1001", "../../etc", att_dir,
                    category="meal", total_amount="20", invoice_number="1001")

            self.assertEqual(rel, str(Path("attachments") / "unknown_date" / "meal_20_1001.pdf"))
            self.assertTrue((runtime / rel).exists())
            self.assertNotIn("..", Path(rel).parts)

    def test_link_downloader_ignores_unsafe_url_schemes(self):
        html = """
        <html><body>
          <a href="file:///etc/passwd">鍙戠エ</a>
          <a href="javascript:alert(1)">鍙戠エ</a>
          <a href="data:text/plain,abc">鍙戠エ</a>
          <a href="https://example.com/download?invoiceId=123">鍙戠エ</a>
        </body></html>
        """

        links = link_dl.extract_links_from_html(html)
        self.assertEqual(links, ["https://example.com/download?invoiceId=123"])

    def test_link_downloader_ignores_local_and_private_network_urls(self):
        html = """
        <html><body>
          <a href="http://127.0.0.1/download?invoiceId=123">下载发票</a>
          <a href="http://192.168.1.5/download?invoiceId=123">下载发票</a>
          <a href="http://localhost/download?invoiceId=123">下载发票</a>
          <a href="https://user:pass@example.com/download?invoiceId=123">下载发票</a>
          <a href="https://example.com/download?invoiceId=123">下载发票</a>
        </body></html>
        """

        links = link_dl.extract_links_from_html(html)
        self.assertEqual(links, ["https://example.com/download?invoiceId=123"])

    def test_link_downloader_auto_prefers_edge_before_chrome(self):
        class FakeBrowser:
            def close(self):
                pass

        class FakeChromium:
            def __init__(self):
                self.calls = []

            def launch(self, channel=None, headless=False, args=None):
                self.calls.append(channel or "chromium")
                if channel == "msedge":
                    return FakeBrowser()
                raise RuntimeError(f"{channel or 'chromium'} unavailable")

        class FakePlaywright:
            def __init__(self, chromium):
                self.chromium = chromium
                self.stopped = False

            def stop(self):
                self.stopped = True

        class FakeSyncPlaywright:
            def __init__(self, chromium):
                self._chromium = chromium

            def start(self):
                return FakePlaywright(self._chromium)

        chromium = FakeChromium()
        with patch("scripts.invoice_fetch.config.load_config_safe", return_value={"playwright": {"channel": "auto"}}), \
                patch("playwright.sync_api.sync_playwright", return_value=FakeSyncPlaywright(chromium)):
            downloader = LinkDownloader(tempfile.mkdtemp())
            downloader._ensure_browser()

        self.assertEqual(chromium.calls, ["msedge"])
        downloader.close()

    def test_save_download_to_path_swallows_download_errors(self):
        class FakeDownload:
            def save_as(self, _):
                raise RuntimeError("Target page, context or browser has been closed")

        with tempfile.TemporaryDirectory() as td:
            dest = Path(td) / "invoice.pdf"
            ok = link_dl._save_download_to_path(FakeDownload(), dest)

        self.assertFalse(ok)
        self.assertFalse(dest.exists())

    def test_link_downloader_does_not_treat_printed_webpage_as_invoice_pdf(self):
        class FakePage:
            url = "https://example.com/invoice/view"

            def title(self):
                return "Invoice detail"

            def pdf(self, path, format="A4", print_background=True):
                Path(path).write_bytes(b"%PDF-" + b"x" * 6000)

        with tempfile.TemporaryDirectory() as td:
            downloader = LinkDownloader(td)
            result = downloader._try_page_print_pdf(FakePage(), Path(td), 123, 0)

            self.assertIsNone(result)
            self.assertEqual(list(Path(td).glob("*.pdf")), [])

    def test_transport_categories_are_split_for_download_names(self):
        categories = {
            "taxi": {"keywords": ["打车", "滴滴", "grab"], "extra_name": "行程记录"},
            "transport": {"keywords": ["火车", "高铁", "通行费", "高速", "过路费"], "extra_name": ""},
        }

        self.assertEqual(
            cli._classify("高铁火车票报销凭证", "rail@example.com", "", categories)[0],
            "火车票",
        )
        self.assertEqual(
            cli._classify("滴滴打车电子发票", "billing@example.com", "", categories)[0],
            "出租车",
        )
        self.assertEqual(
            cli._classify("高速公路通行费电子发票", "billing@example.com", "", categories)[0],
            "过路费",
        )

    def test_default_rules_reduce_other_category_for_common_receipts(self):
        self.assertEqual(
            cli._classify("【海底捞电子发票】您有一张电子发票", "", "", {})[0],
            "餐饮",
        )
        self.assertEqual(
            cli._classify("用车服务电子收据", "支付宝提醒 <service@mail.alipay.com>", "", {})[0],
            "出租车",
        )
        self.assertEqual(
            cli._classify("e-receipt", "NO REPLY <ereceipt@changiairport.com>", "", {})[0],
            "交通",
        )

    def test_failed_processing_does_not_mark_email_downloaded(self):
        row = {"uid": 123}
        db = FakeDB()

        with patch.object(cli, "_process_email", return_value=0):
            marked = cli._handle_pending_email(
                row=row,
                fetcher=FakeFetcher(),
                folder="INBOX",
                att_handler=object(),
                parser=object(),
                link_dl=object(),
                db=db,
                categories={},
            )

        self.assertFalse(marked)
        self.assertEqual(db.marked, [])

    def test_handle_pending_email_passes_row_mailbox_key_to_process_email(self):
        row = {
            "uid": 123,
            "mail_date": "2026-06-07",
            "mailbox_key": "train@example.com",
        }
        db = Mock()

        with patch.object(cli, "_process_email", return_value=1) as mock_process:
            marked = cli._handle_pending_email(
                row=row,
                fetcher=FakeFetcher(),
                folder="INBOX",
                att_handler=object(),
                parser=object(),
                link_dl=object(),
                db=db,
                categories={},
            )

        self.assertTrue(marked)
        self.assertEqual(mock_process.call_args.kwargs["mailbox_key"], "train@example.com")

    def test_duplicate_subject_fallback_marks_email_downloaded(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            msg = email.message.EmailMessage()
            msg["Subject"] = "synthetic invoice subject"
            msg["From"] = "billing@example.com"
            msg["Date"] = "Mon, 18 May 2026 10:00:00 +0800"
            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                db.bulk_upsert_emails([{
                    "uid": 457,
                    "subject": "synthetic invoice subject",
                    "sender": "billing@example.com",
                    "date": "2026-05-18",
                }])
                db.bulk_classify([{
                    "uid": 457,
                    "is_invoice": True,
                    "by": "test",
                    "reason": "synthetic",
                }])
                db.insert_invoice({
                    "invoice_number": "1234567890",
                    "invoice_date": "2026-05-18",
                    "total_amount": "12.30",
                    "seller_name": "Synthetic Seller",
                })

                with (
                    patch.object(cli, "parse_subject", return_value={
                        "invoice_number": "1234567890",
                        "total_amount": "12.30",
                        "seller_name": "Synthetic Seller",
                        "invoice_date": "2026-05-18",
                        "invoice_type": "synthetic invoice",
                    }),
                    patch.object(cli, "extract_html_from_message", return_value=""),
                    patch.object(cli, "parse_html_body", return_value={}),
                ):
                    marked = cli._handle_pending_email(
                        row={"uid": 457, "mail_date": "2026-05-18"},
                        fetcher=FakeFetcher(),
                        folder="INBOX",
                        att_handler=StaticAttachmentHandler(runtime / "attachments", []),
                        parser=StaticParser(InvoiceInfo(parse_success=False)),
                        link_dl=NoopLinkDownloader(),
                        db=db,
                        categories={},
                    )

                pending = db.get_invoice_emails_to_download()

            self.assertTrue(marked)
            self.assertEqual(pending, [])

    def test_duplicate_subject_fallback_without_number_marks_email_downloaded(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            msg = email.message.EmailMessage()
            msg["Subject"] = "synthetic receipt subject"
            msg["From"] = "billing@example.com"
            msg["Date"] = "Mon, 18 May 2026 10:00:00 +0800"
            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                db.bulk_upsert_emails([{
                    "uid": 459,
                    "subject": "synthetic receipt subject",
                    "sender": "billing@example.com",
                    "date": "2026-05-18",
                }])
                db.bulk_classify([{
                    "uid": 459,
                    "is_invoice": True,
                    "by": "test",
                    "reason": "synthetic",
                }])
                db.insert_invoice({
                    "invoice_number": "",
                    "invoice_date": "2026-05-18",
                    "total_amount": "12.30",
                    "seller_name": "Synthetic Seller",
                })

                with (
                    patch.object(cli, "parse_subject", return_value={
                        "total_amount": "12.30",
                        "seller_name": "Synthetic Seller",
                        "invoice_date": "2026-05-18",
                        "invoice_type": "synthetic receipt",
                    }),
                    patch.object(cli, "extract_html_from_message", return_value=""),
                    patch.object(cli, "parse_html_body", return_value={}),
                ):
                    marked = cli._handle_pending_email(
                        row={"uid": 459, "mail_date": "2026-05-18"},
                        fetcher=FakeFetcher(),
                        folder="INBOX",
                        att_handler=StaticAttachmentHandler(runtime / "attachments", []),
                        parser=StaticParser(InvoiceInfo(parse_success=False)),
                        link_dl=NoopLinkDownloader(),
                        db=db,
                        categories={},
                    )

                pending = db.get_invoice_emails_to_download()

            self.assertTrue(marked)
            self.assertEqual(pending, [])

    def test_rescanning_soft_deleted_attachment_restores_invoice(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            att_dir = runtime / "attachments"
            att_dir.mkdir(parents=True)
            src = att_dir / "invoice.pdf"
            src.write_bytes(b"%PDF- synthetic invoice")
            attachment = Attachment(
                file_path=str(src),
                original_name="invoice.pdf",
                content_type="application/pdf",
                size=src.stat().st_size,
                is_invoice=True,
                is_extra=False,
            )
            info = InvoiceInfo(
                invoice_number="SOFT-DEL-001",
                invoice_code="CODE001",
                invoice_date="2026-05-18",
                amount="11.00",
                total_amount="12.30",
                seller_name="Synthetic Seller",
                buyer_name="Synthetic Buyer",
                invoice_type="synthetic invoice",
                parse_success=True,
            )
            msg = email.message.EmailMessage()
            msg["Subject"] = "synthetic invoice subject"
            msg["From"] = "billing@example.com"
            msg["Date"] = "Mon, 18 May 2026 10:00:00 +0800"

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                row_id = db.insert_invoice({
                    "invoice_number": "SOFT-DEL-001",
                    "invoice_code": "CODE001",
                    "invoice_date": "2026-05-18",
                    "amount": "11.00",
                    "total_amount": "12.30",
                    "seller_name": "Synthetic Seller",
                    "buyer_name": "Synthetic Buyer",
                    "invoice_type": "synthetic invoice",
                    "attachment_path": "",
                })
                self.assertTrue(db.soft_delete_invoice(row_id))

                recorded = cli._process_email(
                    cli.MailMessage(uid=458, raw_msg=msg),
                    StaticAttachmentHandler(att_dir, [attachment]),
                    StaticParser(info),
                    NoopLinkDownloader(),
                    db,
                    {},
                )
                rows = db.get_all_invoices()

            self.assertEqual(recorded, 1)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["invoice_number"], "SOFT-DEL-001")
            self.assertEqual(rows[0]["is_deleted"], 0)
            self.assertTrue(rows[0]["attachment_path"])

    def test_email_rescan_does_not_restore_soft_deleted_different_seller(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            att_dir = runtime / "attachments"
            att_dir.mkdir(parents=True)
            src = att_dir / "invoice.pdf"
            src.write_bytes(b"%PDF- synthetic invoice")
            attachment = Attachment(
                file_path=str(src),
                original_name="invoice.pdf",
                content_type="application/pdf",
                size=src.stat().st_size,
                is_invoice=True,
                is_extra=False,
            )
            info = InvoiceInfo(
                invoice_number="SOFT-DEL-SELLER-001",
                invoice_code="CODE001",
                invoice_date="2026-05-18",
                amount="11.00",
                total_amount="12.30",
                seller_name="Different Synthetic Seller",
                buyer_name="Synthetic Buyer",
                invoice_type="synthetic invoice",
                parse_success=True,
            )
            msg = email.message.EmailMessage()
            msg["Subject"] = "synthetic invoice subject"
            msg["From"] = "billing@example.com"
            msg["Date"] = "Mon, 18 May 2026 10:00:00 +0800"

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                row_id = db.insert_invoice({
                    "invoice_number": "SOFT-DEL-SELLER-001",
                    "invoice_code": "CODE001",
                    "invoice_date": "2026-05-18",
                    "amount": "11.00",
                    "total_amount": "12.30",
                    "seller_name": "Original Synthetic Seller",
                    "buyer_name": "Synthetic Buyer",
                    "invoice_type": "synthetic invoice",
                    "attachment_path": "",
                })
                self.assertTrue(db.soft_delete_invoice(row_id))

                with self.assertLogs("invoice_fetch", level="INFO") as logs:
                    recorded = cli._process_email(
                        cli.MailMessage(uid=459, raw_msg=msg),
                        StaticAttachmentHandler(att_dir, [attachment]),
                        StaticParser(info),
                        NoopLinkDownloader(),
                        db,
                        {},
                    )
                all_rows = db.get_all_invoices(include_deleted=True)
                active_rows = db.get_all_invoices()

            self.assertEqual(recorded, 1)
            self.assertEqual(len(all_rows), 2)
            self.assertEqual(len(active_rows), 1)
            self.assertEqual(active_rows[0]["seller_name"], "Different Synthetic Seller")
            self.assertNotIn("已恢复已删除的重复发票", "\n".join(logs.output))

    def test_email_rescan_skips_existing_duplicate_with_diagnostic_log(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            att_dir = runtime / "attachments"
            att_dir.mkdir(parents=True)
            src = att_dir / "invoice.pdf"
            src.write_bytes(b"%PDF- synthetic invoice")
            attachment = Attachment(
                file_path=str(src),
                original_name="invoice.pdf",
                content_type="application/pdf",
                size=src.stat().st_size,
                is_invoice=True,
                is_extra=False,
            )
            info = InvoiceInfo(
                invoice_number="DUP-EXIST-001",
                invoice_code="CODE001",
                invoice_date="2026-05-18",
                amount="11.00",
                total_amount="12.30",
                seller_name="Synthetic Seller",
                buyer_name="Synthetic Buyer",
                invoice_type="synthetic invoice",
                parse_success=True,
            )
            msg = email.message.EmailMessage()
            msg["Subject"] = "synthetic invoice subject"
            msg["From"] = "billing@example.com"
            msg["Date"] = "Mon, 18 May 2026 10:00:00 +0800"

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                row_id = db.insert_invoice({
                    "invoice_number": "DUP-EXIST-001",
                    "invoice_code": "CODE001",
                    "invoice_date": "2026-05-18",
                    "amount": "11.00",
                    "total_amount": "12.30",
                    "seller_name": "Synthetic Seller",
                    "buyer_name": "Synthetic Buyer",
                    "invoice_type": "synthetic invoice",
                    "review_status": "approved",
                    "attachment_path": "",
                })

                with self.assertLogs("invoice_fetch", level="INFO") as logs:
                    recorded = cli._process_email(
                        cli.MailMessage(uid=460, raw_msg=msg),
                        StaticAttachmentHandler(att_dir, [attachment]),
                        StaticParser(info),
                        NoopLinkDownloader(),
                        db,
                        {},
                    )
                rows = db.get_all_invoices(include_deleted=True)

            self.assertEqual(recorded, 1)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["id"], row_id)
            log_text = "\n".join(logs.output)
            self.assertIn("duplicate_reason", log_text)
            self.assertIn("existing_id", log_text)
            self.assertIn("review_status=approved", log_text)
            self.assertIn("is_deleted=0", log_text)

    def test_subject_body_dedup_restores_soft_deleted_without_invoice_number(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            msg = email.message.EmailMessage()
            msg["Subject"] = "synthetic receipt subject"
            msg["From"] = "billing@example.com"
            msg["Date"] = "Mon, 18 May 2026 10:00:00 +0800"
            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                row_id = db.insert_invoice({
                    "invoice_number": "",
                    "invoice_date": "2026-05-18",
                    "total_amount": "12.30",
                    "seller_name": "Synthetic Seller",
                    "parse_note": "old subject fallback",
                })
                self.assertTrue(db.soft_delete_invoice(row_id))

                with (
                    patch.object(cli, "parse_subject", return_value={
                        "total_amount": "12.30",
                        "seller_name": "Synthetic Seller",
                        "invoice_date": "2026-05-18",
                        "invoice_type": "synthetic receipt",
                    }),
                    patch.object(cli, "extract_html_from_message", return_value=""),
                    patch.object(cli, "parse_html_body", return_value={}),
                    self.assertLogs("invoice_fetch", level="INFO") as logs,
                ):
                    recorded = cli._process_email(
                        cli.MailMessage(uid=461, raw_msg=msg),
                        StaticAttachmentHandler(runtime / "attachments", []),
                        StaticParser(InvoiceInfo(parse_success=False)),
                        NoopLinkDownloader(),
                        db,
                        {},
                    )
                rows = db.get_all_invoices()

            self.assertEqual(recorded, 1)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["id"], row_id)
            self.assertEqual(rows[0]["is_deleted"], 0)
            self.assertIn("已恢复已删除的重复发票", "\n".join(logs.output))

    def test_subject_body_unmatched_dedup_stays_pending_without_false_success(self):
        with tempfile.TemporaryDirectory() as td:
            runtime = Path(td) / "runtime"
            msg = email.message.EmailMessage()
            msg["Subject"] = "synthetic invoice link"
            msg["From"] = "billing@example.com"
            msg["Date"] = "Mon, 18 May 2026 10:00:00 +0800"
            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                with (
                    patch.object(cli, "parse_subject", return_value={}),
                    patch.object(cli, "extract_html_from_message", return_value=""),
                    patch.object(cli, "parse_html_body", return_value={}),
                    self.assertLogs("invoice_fetch", level="WARNING") as logs,
                ):
                    recorded = cli._process_email(
                        cli.MailMessage(uid=462, raw_msg=msg),
                        StaticAttachmentHandler(runtime / "attachments", []),
                        StaticParser(InvoiceInfo(parse_success=False)),
                        NoopLinkDownloader(),
                        db,
                        {},
                    )
                rows = db.get_all_invoices()

            self.assertEqual(recorded, 0)
            self.assertEqual(rows, [])
            self.assertIn("链接下载未获得官方 PDF/OFD", "\n".join(logs.output))

    def test_process_email_fallback_subject_parser_is_available(self):
        """Regression: re-read fallback must not fail with missing parser imports."""
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            db = InvoiceDB(base / "invoices.db")
            msg = email.message.EmailMessage()
            msg["Subject"] = "synthetic invoice subject"
            msg["From"] = "billing@example.com"
            msg["Date"] = "Mon, 18 May 2026 10:00:00 +0800"

            with (
                patch.object(cli, "parse_subject", return_value={
                    "invoice_number": "1234567890",
                    "total_amount": "12.30",
                    "seller_name": "Synthetic Seller",
                    "invoice_date": "2026-05-18",
                    "invoice_type": "synthetic invoice",
                }),
                patch.object(cli, "extract_html_from_message", return_value=""),
                patch.object(cli, "parse_html_body", return_value={}),
            ):
                recorded = cli._process_email(
                    cli.MailMessage(uid=456, raw_msg=msg),
                    StaticAttachmentHandler(base, []),
                    StaticParser(InvoiceInfo(parse_success=False)),
                    NoopLinkDownloader(),
                    db,
                    {},
                )

            rows = db.get_all_invoices()
            db.close()

            self.assertEqual(recorded, 1)
            self.assertEqual(rows[0]["invoice_number"], "1234567890")

    def test_process_email_subject_fallback_preserves_mailbox_key(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            db = InvoiceDB(base / "invoices.db")
            msg = email.message.EmailMessage()
            msg["Subject"] = "synthetic invoice subject"
            msg["From"] = "billing@example.com"
            msg["Date"] = "Mon, 18 May 2026 10:00:00 +0800"

            with (
                patch.object(cli, "parse_subject", return_value={
                    "invoice_number": "1234567891",
                    "total_amount": "12.30",
                    "seller_name": "Synthetic Seller",
                    "invoice_date": "2026-05-18",
                    "invoice_type": "synthetic invoice",
                }),
                patch.object(cli, "extract_html_from_message", return_value=""),
                patch.object(cli, "parse_html_body", return_value={}),
            ):
                cli._process_email(
                    cli.MailMessage(uid=457, raw_msg=msg),
                    StaticAttachmentHandler(base, []),
                    StaticParser(InvoiceInfo(parse_success=False)),
                    NoopLinkDownloader(),
                    db,
                    {},
                    mailbox_key="account_b",
                )

            rows = db.get_all_invoices()
            db.close()

            self.assertEqual(rows[0]["mailbox_key"], "account_b")

    def test_link_downloader_keeps_multiple_invoice_pdfs_from_one_email(self):
        msg = email.message.EmailMessage()
        msg.set_content(
            """
            <html><body>
              <a href="https://example.com/download/invoice-1">下载发票</a>
              <a href="https://example.com/download/invoice-2">下载发票</a>
            </body></html>
            """,
            subtype="html",
        )

        files = MultiLinkDownloader().download_from_email(msg, 99, "2026-05-18")

        self.assertEqual([f.filename for f in files], ["invoice_0.pdf", "invoice_1.pdf"])

    def test_excel_export_adds_summary_exception_sheet_and_file_links(self):
        rows = [
            {
                "invoice_number": "1001",
                "invoice_date": "2026-05-18",
                "total_amount": "20.50",
                "category": "餐饮",
                "missing_extra": 0,
                "attachment_path": "attachments/2026-05-18/a.pdf",
            },
            {
                "invoice_number": "1002",
                "invoice_date": "2026-05-19",
                "total_amount": "88.00",
                "category": "打车出行",
                "missing_extra": 1,
                "parse_note": "缺少行程记录",
                "attachment_path": "",
            },
        ]

        with tempfile.TemporaryDirectory() as td:
            dest = Path(td) / "发票汇总.xlsx"
            export_excel(rows, dest)
            wb = load_workbook(dest)

        self.assertIn("发票汇总", wb.sheetnames)
        self.assertIn("分类汇总", wb.sheetnames)
        self.assertIn("异常待处理", wb.sheetnames)
        self.assertEqual(wb["分类汇总"]["A2"].value, "餐饮")
        self.assertEqual(wb["异常待处理"]["A2"].value, "1002")
        self.assertIsNotNone(wb["发票汇总"]["O2"].hyperlink)

    def test_overseas_receipt_pdf_is_preserved_without_invoice_number(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            src = base / "changi_receipt.pdf"
            src.write_bytes(b"%PDF- receipt")
            msg = email.message.EmailMessage()
            msg["Subject"] = "e-receipt"
            msg["From"] = "NO REPLY <ereceipt@changiairport.com>"
            msg["Date"] = "Fri, 16 Oct 2025 10:00:00 +0800"
            attachment = Attachment(
                file_path=str(src),
                original_name="changi_receipt.pdf",
                content_type="application/pdf",
                size=src.stat().st_size,
                is_invoice=True,
                is_extra=False,
            )
            db = InvoiceDB(base / "invoices.db")

            recorded = cli._process_email(
                cli.MailMessage(uid=3842, raw_msg=msg),
                StaticAttachmentHandler(base, [attachment]),
                StaticParser(InvoiceInfo(parse_success=False, parse_note="内容不像发票")),
                NoopLinkDownloader(),
                db,
                {},
            )
            rows = db.get_all_invoices()
            db.close()

            self.assertEqual(recorded, 1)
            self.assertEqual(rows[0]["invoice_type"], "海外凭证/收据")
            self.assertEqual(rows[0]["parse_success"], 1)
            self.assertTrue((base.parent / rows[0]["attachment_path"]).exists())

    def test_non_receipt_bad_pdf_is_not_recorded_as_downloaded(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            src = base / "random.pdf"
            src.write_bytes(b"%PDF- random")
            msg = email.message.EmailMessage()
            msg["Subject"] = "weekly report"
            msg["From"] = "noreply@example.com"
            msg["Date"] = "Fri, 16 Oct 2025 10:00:00 +0800"
            attachment = Attachment(
                file_path=str(src),
                original_name="random.pdf",
                content_type="application/pdf",
                size=src.stat().st_size,
                is_invoice=True,
                is_extra=False,
            )
            db = InvoiceDB(base / "invoices.db")

            recorded = cli._process_email(
                cli.MailMessage(uid=99, raw_msg=msg),
                StaticAttachmentHandler(base, [attachment]),
                StaticParser(InvoiceInfo(parse_success=False, parse_note="内容不像发票")),
                NoopLinkDownloader(),
                db,
                {},
            )
            rows = db.get_all_invoices()
            db.close()

            self.assertEqual(recorded, 0)
            self.assertEqual(rows, [])

    def test_process_email_repairs_missing_paths_on_duplicate_invoice(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            attachments_root = runtime / "attachments"
            attachments_root.mkdir(parents=True, exist_ok=True)

            invoice_file = base / "invoice.pdf"
            invoice_file.write_bytes(b"%PDF- invoice")
            extra_file = base / "trip_record.pdf"
            extra_file.write_bytes(b"%PDF- trip")

            msg = email.message.EmailMessage()
            msg["Subject"] = "江苏省收费公路通行费电子票据"
            msg["From"] = "jsgs96777fapiao@mail.fapiao.js96777.com"
            msg["Date"] = "Fri, 26 Dec 2025 10:00:00 +0800"

            attachments = [
                Attachment(
                    file_path=str(invoice_file),
                    original_name="invoice.pdf",
                    content_type="application/pdf",
                    size=invoice_file.stat().st_size,
                    is_invoice=True,
                    is_extra=False,
                ),
                Attachment(
                    file_path=str(extra_file),
                    original_name="江苏高速通行费电子票据行程单_20251226.pdf",
                    content_type="application/pdf",
                    size=extra_file.stat().st_size,
                    is_invoice=False,
                    is_extra=True,
                ),
            ]

            parser_info = InvoiceInfo(
                invoice_number="32020125",
                invoice_code="0052691226",
                invoice_date="2025-12-26",
                total_amount="30.00",
                seller_name="科技有限公司",
                buyer_name="",
                invoice_type="电子发票",
                parse_success=True,
                parse_note="ok",
            )

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                db.insert_invoice({
                    "invoice_number": "32020125",
                    "invoice_code": "0052691226",
                    "invoice_date": "2025-12-26",
                    "total_amount": "30.00",
                    "seller_name": "科技有限公司",
                    "invoice_type": "电子发票",
                    "category": "过路费",
                    "has_extra": False,
                    "missing_extra": True,
                    "attachment_path": "",
                    "extra_paths": [],
                    "mail_uid": 4050,
                    "mail_subject": "江苏省收费公路通行费电子票据",
                    "mail_date": "2025-12-26",
                    "mail_sender": "jsgs96777fapiao@mail.fapiao.js96777.com",
                    "parse_success": True,
                    "parse_note": "旧记录",
                    "review_status": "to_review",
                })

                recorded = cli._process_email(
                    cli.MailMessage(uid=4050, raw_msg=msg),
                    StaticAttachmentHandler(attachments_root, attachments),
                    StaticParser(parser_info),
                    NoopLinkDownloader(),
                    db,
                    {"toll": {"keywords": ["通行费"], "extra_name": "行程单"}},
                )
                rows = db.get_all_invoices()

            self.assertEqual(recorded, 1)
            self.assertEqual(len(rows), 1)
            row = rows[0]
            self.assertTrue((runtime / row["attachment_path"]).exists())
            extra_paths = json.loads(row["extra_paths"]) if isinstance(row["extra_paths"], str) else row["extra_paths"]
            self.assertEqual(len(extra_paths), 1)
            self.assertTrue((runtime / extra_paths[0]).exists())
            self.assertTrue(row["has_extra"])
            self.assertFalse(row["missing_extra"])

    def test_process_email_updates_existing_invoice_by_number_and_amount(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            src = base / "invoice.pdf"
            src.write_bytes(b"%PDF- invoice")
            msg = email.message.EmailMessage()
            msg["Subject"] = "Invoice reminder"
            msg["From"] = "billing@example.com"
            msg["Date"] = "Fri, 16 Oct 2025 10:00:00 +0800"
            attachment = Attachment(
                file_path=str(src),
                original_name="invoice.pdf",
                content_type="application/pdf",
                size=src.stat().st_size,
                is_invoice=True,
                is_extra=False,
            )
            db = InvoiceDB(base / "invoices.db")
            db.insert_invoice({
                "invoice_number": "EMAIL001",
                "total_amount": "88.00",
                "seller_name": "旧销售方",
                "invoice_date": "2025-10-16",
                "review_status": "to_review",
            })

            recorded = cli._process_email(
                cli.MailMessage(uid=101, raw_msg=msg),
                StaticAttachmentHandler(base, [attachment]),
                StaticParser(InvoiceInfo(
                    invoice_number="EMAIL001",
                    invoice_code="31001519300050031029",
                    invoice_date="2025-10-16",
                    total_amount="88.00",
                    seller_name="新销售方",
                    buyer_name="购买方",
                    invoice_type="电子发票",
                    parse_success=True,
                )),
                NoopLinkDownloader(),
                db,
                {},
            )
            rows = db.get_all_invoices()
            db.close()

            self.assertEqual(recorded, 1)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["seller_name"], "旧销售方")
            self.assertEqual(rows[0]["buyer_name"], "购买方")
            self.assertEqual(rows[0]["invoice_code"], "31001519300050031029")

    def test_standalone_hotel_folio_attachment_is_recorded_as_receipt(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            src = base / "hotel_folio.pdf"
            src.write_bytes(b"%PDF- folio")
            msg = email.message.EmailMessage()
            msg["Subject"] = "Your hotel folio"
            msg["From"] = "hotel@example.sg"
            msg["Date"] = "Fri, 16 Oct 2025 10:00:00 +0800"
            attachment = Attachment(
                file_path=str(src),
                original_name="hotel_folio.pdf",
                content_type="application/pdf",
                size=src.stat().st_size,
                is_invoice=False,
                is_extra=True,
            )
            db = InvoiceDB(base / "invoices.db")

            recorded = cli._process_email(
                cli.MailMessage(uid=100, raw_msg=msg),
                StaticAttachmentHandler(base, [attachment]),
                StaticParser(InvoiceInfo(parse_success=False, parse_note="内容不像发票")),
                NoopLinkDownloader(),
                db,
                {"hotel": {"keywords": ["hotel"], "extra_name": "水单"}},
            )
            rows = db.get_all_invoices()
            db.close()

            self.assertEqual(recorded, 1)
            self.assertEqual(rows[0]["category"], "酒店住宿")
            self.assertEqual(rows[0]["invoice_type"], "海外凭证/收据")
            self.assertFalse(rows[0]["missing_extra"])

    def test_unbalanced_parentheses_repair_seller(self):
        """测试 1：缺中文右括号时自动补全"""
        text = "\n".join([
            "电子发票（普通发票）",
            "销售方名称：南京市钾程水饺店（有限合伙",
            "统一社会信用代码：91320000123456789A",
            "购买方名称：远景智能零碳（江苏）科技有限公司",
            "发票号码：2632200002260046326",
            "开票日期：2026年03月24日",
            "价税合计（小写）¥30.00",
        ])
        parser = InvoiceParser()
        fake_plumber = _FakePdfPlumber(text)
        with tempfile.TemporaryDirectory() as td, patch.object(parser, "_plumber", return_value=fake_plumber):
            pdf_path = Path(td) / "invoice.pdf"
            pdf_path.write_bytes(b"%PDF-1.4 synthetic")
            info = parser.parse_pdf(str(pdf_path))
        self.assertTrue(info.parse_success)
        self.assertEqual(info.seller_name, "南京市钾程水饺店（有限合伙）")
        self.assertEqual(info.buyer_name, "远景智能零碳（江苏）科技有限公司")

    def test_balanced_parentheses_no_double_repair(self):
        """测试 2：完整括号不重复补"""
        text = "\n".join([
            "电子发票（普通发票）",
            "销售方名称：南京市钾程水饺店（有限合伙）",
            "统一社会信用代码：91320000123456789A",
            "购买方名称：远景智能零碳（江苏）科技有限公司",
            "发票号码：2632200002260046326",
            "开票日期：2026年03月24日",
            "价税合计（小写）¥30.00",
        ])
        parser = InvoiceParser()
        fake_plumber = _FakePdfPlumber(text)
        with tempfile.TemporaryDirectory() as td, patch.object(parser, "_plumber", return_value=fake_plumber):
            pdf_path = Path(td) / "invoice.pdf"
            pdf_path.write_bytes(b"%PDF-1.4 synthetic")
            info = parser.parse_pdf(str(pdf_path))
        self.assertTrue(info.parse_success)
        self.assertEqual(info.seller_name, "南京市钾程水饺店（有限合伙）")

    def test_unbalanced_english_parentheses_repair(self):
        """测试 3：英文括号缺失时也能补"""
        text = "\n".join([
            "电子发票（普通发票）",
            "销售方名称：南京市钾程水饺店(有限合伙",
            "统一社会信用代码：91320000123456789A",
            "购买方名称：远景智能零碳（江苏）科技有限公司",
            "发票号码：2632200002260046326",
            "开票日期：2026年03月24日",
            "价税合计（小写）¥30.00",
        ])
        parser = InvoiceParser()
        fake_plumber = _FakePdfPlumber(text)
        with tempfile.TemporaryDirectory() as td, patch.object(parser, "_plumber", return_value=fake_plumber):
            pdf_path = Path(td) / "invoice.pdf"
            pdf_path.write_bytes(b"%PDF-1.4 synthetic")
            info = parser.parse_pdf(str(pdf_path))
        self.assertTrue(info.parse_success)
        self.assertEqual(info.seller_name, "南京市钾程水饺店(有限合伙)")

    def test_unbalanced_parentheses_no_repair_for_general_remarks(self):
        """测试 4：不应乱补普通文本"""
        text = "\n".join([
            "电子发票（普通发票）",
            "销售方名称：南京市钾程水饺店（测试数据",
            "统一社会信用代码：91320000123456789A",
            "购买方名称：远景智能零碳（江苏）科技有限公司",
            "发票号码：2632200002260046326",
            "开票日期：2026年03月24日",
            "价税合计（小写）¥30.00",
        ])
        parser = InvoiceParser()
        fake_plumber = _FakePdfPlumber(text)
        with tempfile.TemporaryDirectory() as td, patch.object(parser, "_plumber", return_value=fake_plumber):
            pdf_path = Path(td) / "invoice.pdf"
            pdf_path.write_bytes(b"%PDF-1.4 synthetic")
            info = parser.parse_pdf(str(pdf_path))
        self.assertTrue(info.parse_success)
        # 括号内不是组织后缀，不应补右括号
        self.assertEqual(info.seller_name, "南京市钾程水饺店（测试数据")

    def test_didi_evidence_matching_by_invoice_number(self):
        """测试 1：滴滴行程单按发票号匹配仍然有效"""
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            import_dir = base / "local_evidence"
            import_dir.mkdir(parents=True)
            evidence = import_dir / "滴滴行程单_2532700001694376933.pdf"
            evidence.write_bytes(b"%PDF- trip evidence")

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                invoice_id = db.insert_invoice({
                    "invoice_number": "2532700001694376933",
                    "invoice_date": "2025-12-20",
                    "total_amount": "1142.81",
                    "seller_name": "南京滴滴出行科技有限公司",
                    "category": "出租车",
                    "extra_paths": [],
                })
                # Mock parse_pdf returning a non-success parse (evidence typically fails standard invoice parse)
                parser = StaticParser(InvoiceInfo(parse_success=False, parse_note="行程单无标准发票号"))
                stats = cli._import_local_directory(
                    import_dir=import_dir,
                    db=db,
                    parser=parser,
                    categories={},
                    att_dir=runtime / "attachments",
                )
                rows = db.get_all_invoices()
                updated = db.get_invoice(invoice_id)

            self.assertEqual(stats["added"], 1)
            self.assertEqual(len(rows), 1)
            extra_paths = json.loads(updated["extra_paths"])
            self.assertEqual(len(extra_paths), 1)

    def test_same_mail_extra_file_attaches_to_multiple_invoices(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            runtime.mkdir()
            attachments_dir = runtime / "attachments"
            attachments_dir.mkdir()

            evidence_file = attachments_dir / "trip.pdf"
            evidence_file.write_bytes(b"%PDF-evidence")

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                invoice_id_1 = db.insert_invoice({
                    "invoice_number": "SHARED001",
                    "invoice_code": "001",
                    "invoice_date": "2026-05-18",
                    "total_amount": "88.00",
                    "seller_name": "Shared Seller A",
                    "invoice_type": "电子发票",
                    "extra_paths": [],
                    "has_extra": False,
                    "missing_extra": True,
                })
                invoice_id_2 = db.insert_invoice({
                    "invoice_number": "SHARED002",
                    "invoice_code": "002",
                    "invoice_date": "2026-05-18",
                    "total_amount": "188.00",
                    "seller_name": "Shared Seller B",
                    "invoice_type": "电子发票",
                    "extra_paths": [],
                    "has_extra": False,
                    "missing_extra": True,
                })

                attachment = Attachment(
                    file_path=str(evidence_file),
                    original_name="trip.pdf",
                    content_type="application/pdf",
                    size=evidence_file.stat().st_size,
                    is_invoice=False,
                    is_extra=True,
                )
                cli._attach_email_extras_to_invoice(
                    db=db,
                    invoice_id=invoice_id_1,
                    extra_files=[attachment],
                    code="001",
                    inv_date="2026-05-18",
                    att_base=attachments_dir,
                    category="餐饮",
                    total_amount="88.00",
                    invoice_number="SHARED001",
                    kept_paths=set(),
                )
                cli._attach_email_extras_to_invoice(
                    db=db,
                    invoice_id=invoice_id_2,
                    extra_files=[attachment],
                    code="002",
                    inv_date="2026-05-18",
                    att_base=attachments_dir,
                    category="餐饮",
                    total_amount="188.00",
                    invoice_number="SHARED002",
                    kept_paths=set(),
                )
                rows = sorted(db.get_all_invoices(), key=lambda row: row["id"])

            self.assertEqual(len(rows), 2)
            for row in rows:
                extra_paths = json.loads(row["extra_paths"])
                self.assertEqual(len(extra_paths), 1)
                self.assertTrue((runtime / extra_paths[0]).exists())
                self.assertTrue(row["has_extra"])
                self.assertFalse(row["missing_extra"])

    def test_email_image_attachment_is_recorded_as_pending_manual(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            runtime.mkdir()
            attachments_dir = runtime / "attachments"
            attachments_dir.mkdir()

            image_file = attachments_dir / "IMG_001.jpg"
            image_file.write_bytes(b"\xff\xd8\xff synthetic image")
            attachments = [
                Attachment(file_path=str(image_file), original_name="IMG_001.jpg", content_type="image/jpeg", size=image_file.stat().st_size, is_invoice=False, is_extra=False),
            ]

            msg = email.message.EmailMessage()
            msg["Subject"] = "plain photo receipt"
            msg["From"] = "sender@example.com"
            msg["Date"] = "Mon, 18 May 2026 10:00:00 +0800"
            mail_msg = cli.MailMessage(uid=778, raw_msg=msg)

            parser = StaticParser(InvoiceInfo(parse_success=False, parse_note="not invoice"))

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                handler = StaticAttachmentHandler(attachments_dir, attachments)
                recorded = cli._process_email(
                    msg=mail_msg,
                    att_handler=handler,
                    parser=parser,
                    link_dl=NoopLinkDownloader(),
                    db=db,
                    categories={},
                )
                rows = db.get_all_invoices()

            self.assertEqual(recorded, 1)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["invoice_type"], "图片待识别")
            self.assertIn("图片待识别", rows[0]["parse_note"])
            self.assertFalse(rows[0]["parse_success"])
            stored_path = runtime / rows[0]["attachment_path"]
            self.assertTrue(stored_path.exists())
            self.assertEqual(stored_path.read_bytes(), b"\xff\xd8\xff synthetic image")

    def test_receipt_duplicate_scan_is_deduped_by_hash_and_source(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            runtime.mkdir()
            attachments_dir = runtime / "attachments"
            attachments_dir.mkdir()

            receipt_file1 = attachments_dir / "receipt1.pdf"
            receipt_file1.write_bytes(b"%PDF-receipt-same")
            receipt_file2 = attachments_dir / "receipt2.pdf"
            receipt_file2.write_bytes(b"%PDF-receipt-same")

            msg1 = email.message.EmailMessage()
            msg1["Subject"] = "hotel receipt"
            msg1["From"] = "hotel@example.com"
            msg1["Date"] = "Mon, 18 May 2026 10:00:00 +0800"
            msg2 = email.message.EmailMessage()
            msg2["Subject"] = "hotel receipt"
            msg2["From"] = "hotel@example.com"
            msg2["Date"] = "Mon, 18 May 2026 10:05:00 +0800"

            attachment1 = Attachment(
                file_path=str(receipt_file1),
                original_name="receipt.pdf",
                content_type="application/pdf",
                size=receipt_file1.stat().st_size,
                is_invoice=False,
                is_extra=True,
            )
            attachment2 = Attachment(
                file_path=str(receipt_file2),
                original_name="receipt.pdf",
                content_type="application/pdf",
                size=receipt_file2.stat().st_size,
                is_invoice=False,
                is_extra=True,
            )

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                handler1 = StaticAttachmentHandler(attachments_dir, [attachment1])
                handler2 = StaticAttachmentHandler(attachments_dir, [attachment2])
                cli._process_email(
                    msg=cli.MailMessage(uid=900, raw_msg=msg1),
                    att_handler=handler1,
                    parser=StaticParser(InvoiceInfo(parse_success=False, parse_note="not invoice")),
                    link_dl=NoopLinkDownloader(),
                    db=db,
                    categories={},
                    mailbox_key="account_a",
                )
                cli._process_email(
                    msg=cli.MailMessage(uid=900, raw_msg=msg2),
                    att_handler=handler2,
                    parser=StaticParser(InvoiceInfo(parse_success=False, parse_note="not invoice")),
                    link_dl=NoopLinkDownloader(),
                    db=db,
                    categories={},
                    mailbox_key="account_a",
                )
                rows = db.get_all_invoices()

            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["invoice_type"], "\u6d77\u5916\u51ed\u8bc1/\u6536\u636e")
            self.assertTrue(rows[0]["file_hash"])
            self.assertTrue((runtime / rows[0]["attachment_path"]).exists())

    def test_didi_evidence_matching_by_date_and_amount(self):
        """测试 2：滴滴行程单无发票号时，用日期 + 金额 + 滴滴关键词唯一匹配"""
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            import_dir = base / "local_evidence"
            import_dir.mkdir(parents=True)
            evidence = import_dir / "滴滴行程单.pdf"
            # Simulate parsed raw_text containing dates and amounts
            raw_text = "滴滴出行行程单 日期：2025-12-20 金额：1142.81"
            evidence.write_bytes(b"%PDF- trip evidence text")

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                invoice_id = db.insert_invoice({
                    "invoice_number": "2532700001694376933",
                    "invoice_date": "2025-12-20",
                    "total_amount": "1142.81",
                    "seller_name": "南京滴滴出行科技有限公司",
                    "category": "出租车",
                    "extra_paths": [],
                })
                parser = StaticParser(InvoiceInfo(
                    parse_success=False,
                    parse_note="行程单无标准发票号",
                    raw_text=raw_text
                ))
                stats = cli._import_local_directory(
                    import_dir=import_dir,
                    db=db,
                    parser=parser,
                    categories={},
                    att_dir=runtime / "attachments",
                )
                rows = db.get_all_invoices()
                updated = db.get_invoice(invoice_id)

            self.assertEqual(stats["added"], 1)
            self.assertEqual(len(rows), 1)
            extra_paths = json.loads(updated["extra_paths"])
            self.assertEqual(len(extra_paths), 1)
            self.assertIn("滴滴行程单.pdf", extra_paths[0])

    def test_evidence_amount_candidates_support_integer_and_currency_forms(self):
        samples = {
            "滴滴行程单 2026-06-01 实付 ¥35": "35.00",
            "高德打车 2026年6月1日 合计35元": "35.00",
            "T3出行 金额 RMB 35.0": "35.00",
            "出租车 总计 CNY 35": "35.00",
        }
        for text, expected in samples.items():
            with self.subTest(text=text):
                self.assertIn(
                    expected,
                    cli._extract_amount_candidates_for_evidence(text),
                )

    def test_evidence_amount_candidates_ignore_long_order_numbers(self):
        self.assertEqual(
            cli._extract_amount_candidates_for_evidence(
                "滴滴行程单 2026-06-01 订单号 2026060112345678"
            ),
            [],
        )

    def test_transport_evidence_integer_amount_matches_by_date_and_context(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            evidence = base / "stored.pdf"
            evidence.write_bytes(b"%PDF- evidence")
            parsed = InvoiceInfo(
                parse_success=False,
                raw_text="滴滴行程单 2026-06-01 实付 ¥35",
            )
            with InvoiceDB(base / "invoices.db") as db:
                invoice_id = db.insert_invoice({
                    "invoice_number": "INTEGER-AMOUNT-001",
                    "invoice_date": "2026-06-01",
                    "total_amount": "35.00",
                    "seller_name": "Synthetic Taxi",
                    "category": "出租车",
                })
                matched, status = cli._find_matching_invoice_for_evidence(
                    db,
                    parsed,
                    evidence,
                )

            self.assertIsNone(status)
            self.assertEqual(matched["id"], invoice_id)

    def test_evidence_matching_uses_original_source_name_hints(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            stored_file = base / "stored.pdf"
            stored_file.write_bytes(b"%PDF- evidence")
            parsed = InvoiceInfo(parse_success=False, parse_note="", raw_text="")

            with InvoiceDB(runtime / "invoices.db") as db:
                invoice_id = db.insert_invoice({
                    "invoice_number": "DIDI-SOURCE-001",
                    "invoice_date": "2025-12-20",
                    "total_amount": "1142.81",
                    "seller_name": "南京滴滴出行科技有限公司",
                    "category": "出租车",
                })
                matched, status = cli._find_matching_invoice_for_evidence(
                    db,
                    parsed,
                    stored_file,
                    source_name="滴滴行程单_2025-12-20_1142.81.pdf",
                )

            self.assertIsNone(status)
            self.assertEqual(matched["id"], invoice_id)

    def test_didi_evidence_matching_multiple_candidates_fails(self):
        """测试 3：候选多张时不自动关联"""
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            import_dir = base / "local_evidence"
            import_dir.mkdir(parents=True)
            evidence = import_dir / "滴滴行程单.pdf"
            raw_text = "滴滴出行行程单 日期：2025-12-20 金额：1142.81"
            evidence.write_bytes(b"%PDF- trip evidence text")

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                db.insert_invoice({
                    "invoice_number": "11111111",
                    "invoice_date": "2025-12-20",
                    "total_amount": "1142.81",
                    "seller_name": "南京滴滴出行科技有限公司",
                    "category": "出租车",
                    "extra_paths": [],
                })
                db.insert_invoice({
                    "invoice_number": "22222222",
                    "invoice_date": "2025-12-20",
                    "total_amount": "1142.81",
                    "seller_name": "南京滴滴出行科技有限公司",
                    "category": "出租车",
                    "extra_paths": [],
                })
                parser = StaticParser(InvoiceInfo(
                    parse_success=False,
                    parse_note="行程单无标准发票号",
                    raw_text=raw_text
                ))
                stats = cli._import_local_directory(
                    import_dir=import_dir,
                    db=db,
                    parser=parser,
                    categories={},
                    att_dir=runtime / "attachments",
                )
                rows = db.get_all_invoices()

            self.assertEqual(stats["pending_manual"], 1)
            self.assertEqual(len(rows), 3)  # Two standard invoices + one pending manual evidence
            evidence_row = [r for r in rows if r["invoice_type"] == "待关联证明材料"][0]
            self.assertIn("疑似滴滴/出租车证明材料，但匹配到多张候选发票，请人工关联", evidence_row["parse_note"])

    def test_non_evidence_failed_pdf_is_not_associated(self):
        """测试 4：普通解析失败 PDF 不应误判为行程单"""
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            import_dir = base / "local_invoices"
            import_dir.mkdir(parents=True)
            unknown_pdf = import_dir / "unknown.pdf"
            unknown_pdf.write_bytes(b"%PDF- bad file")

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                # Target candidates with date and amount matching to try to trigger it
                db.insert_invoice({
                    "invoice_number": "2532700001694376933",
                    "invoice_date": "2025-12-20",
                    "total_amount": "1142.81",
                    "seller_name": "南京滴滴出行科技有限公司",
                    "category": "出租车",
                    "extra_paths": [],
                })
                parser = StaticParser(InvoiceInfo(
                    parse_success=False,
                    parse_note="无法解析发票内容",
                    raw_text="Random text here dated 2025-12-20 with amount 1142.81"
                ))
                stats = cli._import_local_directory(
                    import_dir=import_dir,
                    db=db,
                    parser=parser,
                    categories={},
                    att_dir=runtime / "attachments",
                )
                rows = db.get_all_invoices()

            self.assertEqual(stats["pending_manual"], 1)
            self.assertEqual(len(rows), 2)
            evidence_row = [r for r in rows if r["invoice_type"] == "本地导入待处理"][0]
            self.assertNotIn("待关联证明材料", evidence_row["invoice_type"])

    def test_duplicate_evidence_import_is_idempotent(self):
        """测试 5：重复导入同一行程单不重复写入 extra_paths"""
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            import_dir = base / "local_evidence"
            import_dir.mkdir(parents=True)
            evidence = import_dir / "滴滴行程单_2532700001694376933.pdf"
            evidence.write_bytes(b"%PDF- trip evidence content")

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                invoice_id = db.insert_invoice({
                    "invoice_number": "2532700001694376933",
                    "invoice_date": "2025-12-20",
                    "total_amount": "1142.81",
                    "seller_name": "南京滴滴出行科技有限公司",
                    "category": "出租车",
                    "extra_paths": [],
                })
                parser = StaticParser(InvoiceInfo(parse_success=False, parse_note="行程单"))
                # Import once
                stats1 = cli._import_local_directory(
                    import_dir=import_dir,
                    db=db,
                    parser=parser,
                    categories={},
                    att_dir=runtime / "attachments",
                )
                # Import twice
                stats2 = cli._import_local_directory(
                    import_dir=import_dir,
                    db=db,
                    parser=parser,
                    categories={},
                    att_dir=runtime / "attachments",
                )
                updated = db.get_invoice(invoice_id)

            self.assertEqual(stats1["added"], 1)
            self.assertEqual(stats2["duplicates"], 1)
            extra_paths = json.loads(updated["extra_paths"])
            self.assertEqual(len(extra_paths), 1)

    def test_email_attachment_classification_for_extra_files(self):
        """测试 1：邮箱附件分类"""
        from scripts.invoice_fetch.attachment_handler import AttachmentHandler
        import email.message
        msg = email.message.EmailMessage()
        msg["Subject"] = "发票行程单邮件"
        msg["From"] = "user@example.com"
        msg["Date"] = "Mon, 18 May 2026 10:00:00 +0800"

        attachments_to_add = [
            ("用车明细.pdf", b"%PDF-detail"),
            ("行程记录.pdf", b"%PDF-record"),
            ("ride_detail.pdf", b"%PDF-ride"),
            ("invoice.pdf", b"%PDF-invoice"),
            ("发票_行程单.pdf", b"%PDF-both"),
        ]
        for filename, data in attachments_to_add:
            msg.add_attachment(data, maintype="application", subtype="pdf", filename=filename)

        with tempfile.TemporaryDirectory() as td:
            handler = AttachmentHandler(td)
            extracted = handler.extract(msg, mail_uid=123, date_str="2026-06-06")

            by_name = {a.original_name: a for a in extracted}

            self.assertIn("用车明细.pdf", by_name)
            self.assertTrue(by_name["用车明细.pdf"].is_extra)
            self.assertFalse(by_name["用车明细.pdf"].is_invoice)

            self.assertIn("行程记录.pdf", by_name)
            self.assertTrue(by_name["行程记录.pdf"].is_extra)
            self.assertFalse(by_name["行程记录.pdf"].is_invoice)

            self.assertIn("ride_detail.pdf", by_name)
            self.assertTrue(by_name["ride_detail.pdf"].is_extra)
            self.assertFalse(by_name["ride_detail.pdf"].is_invoice)

            self.assertIn("invoice.pdf", by_name)
            self.assertFalse(by_name["invoice.pdf"].is_extra)
            self.assertTrue(by_name["invoice.pdf"].is_invoice)

            self.assertIn("发票_行程单.pdf", by_name)
            self.assertTrue(by_name["发票_行程单.pdf"].is_extra)
            self.assertFalse(by_name["发票_行程单.pdf"].is_invoice)

            invoice_pdfs = [a for a in extracted if a.is_invoice and a.file_path.lower().endswith(".pdf")]
            self.assertEqual(len(invoice_pdfs), 1)
            self.assertEqual(invoice_pdfs[0].original_name, "invoice.pdf")

    def test_same_email_invoice_and_itinerary_import(self):
        """测试 2：同一封邮件发票 + 行程单入库"""
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            runtime.mkdir()
            attachments_dir = runtime / "attachments"
            attachments_dir.mkdir()

            inv_file = attachments_dir / "invoice.pdf"
            inv_file.write_bytes(b"%PDF-invoice")
            iti_file = attachments_dir / "用车明细.pdf"
            iti_file.write_bytes(b"%PDF-iti")

            attachments = [
                Attachment(file_path=str(inv_file), original_name="invoice.pdf", content_type="application/pdf", size=1024, is_invoice=True, is_extra=False),
                Attachment(file_path=str(iti_file), original_name="用车明细.pdf", content_type="application/pdf", size=1024, is_invoice=False, is_extra=True),
            ]

            msg = email.message.EmailMessage()
            msg["Subject"] = "滴滴打车发票及明细"
            msg["From"] = "didi@example.com"
            msg["Date"] = "Mon, 18 May 2026 10:00:00 +0800"
            mail_msg = cli.MailMessage(uid=101, raw_msg=msg)

            parser = StaticParser(InvoiceInfo(
                invoice_number="DIDI999888",
                invoice_code="11002233",
                invoice_date="2025-12-20",
                total_amount="123.45",
                seller_name="南京滴滴出行科技有限公司",
                invoice_type="电子发票",
                parse_success=True,
            ))
            handler = StaticAttachmentHandler(attachments_dir, attachments)

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                recorded = cli._process_email(
                    msg=mail_msg,
                    att_handler=handler,
                    parser=parser,
                    link_dl=NoopLinkDownloader(),
                    db=db,
                    categories={},
                )
                rows = db.get_all_invoices()

            self.assertEqual(recorded, 1)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["invoice_number"], "DIDI999888")
            self.assertEqual(rows[0]["category"], "出租车")

            extra_paths = json.loads(rows[0]["extra_paths"])
            self.assertEqual(len(extra_paths), 1)
            self.assertIn("DIDI999888_ex.pdf", extra_paths[0].replace("\\", "/"))
            self.assertTrue(rows[0]["has_extra"])
            self.assertFalse(rows[0]["missing_extra"])

    def test_multi_invoice_email_keeps_unmatched_evidence_pending(self):
        class MappingParser:
            def parse_pdf(self, path):
                name = Path(path).name
                if name == "invoice-a.pdf":
                    return InvoiceInfo(
                        invoice_number="MULTI-A-001",
                        invoice_date="2026-06-01",
                        total_amount="35.00",
                        seller_name="Synthetic Seller A",
                        invoice_type="电子发票",
                        parse_success=True,
                    )
                return InvoiceInfo(
                    invoice_number="MULTI-B-002",
                    invoice_date="2026-06-02",
                    total_amount="45.00",
                    seller_name="Synthetic Seller B",
                    invoice_type="电子发票",
                    parse_success=True,
                )

        with tempfile.TemporaryDirectory() as td:
            runtime = Path(td) / "runtime"
            attachments_dir = runtime / "attachments"
            attachments_dir.mkdir(parents=True)
            invoice_a = attachments_dir / "invoice-a.pdf"
            invoice_b = attachments_dir / "invoice-b.pdf"
            evidence = attachments_dir / "用车明细.pdf"
            invoice_a.write_bytes(b"%PDF-invoice-a")
            invoice_b.write_bytes(b"%PDF-invoice-b")
            evidence.write_bytes(b"%PDF-evidence")
            attachments = [
                Attachment(
                    file_path=str(invoice_a),
                    original_name=invoice_a.name,
                    content_type="application/pdf",
                    size=invoice_a.stat().st_size,
                    is_invoice=True,
                ),
                Attachment(
                    file_path=str(invoice_b),
                    original_name=invoice_b.name,
                    content_type="application/pdf",
                    size=invoice_b.stat().st_size,
                    is_invoice=True,
                ),
                Attachment(
                    file_path=str(evidence),
                    original_name=evidence.name,
                    content_type="application/pdf",
                    size=evidence.stat().st_size,
                    is_extra=True,
                ),
            ]
            msg = email.message.EmailMessage()
            msg["Subject"] = "两张电子发票及用车明细"
            msg["From"] = "billing@example.com"
            msg["Date"] = "Mon, 1 Jun 2026 10:00:00 +0800"

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(
                cli, "RUNTIME_DIR", runtime
            ):
                recorded = cli._process_email(
                    msg=cli.MailMessage(uid=1901, raw_msg=msg),
                    att_handler=StaticAttachmentHandler(attachments_dir, attachments),
                    parser=MappingParser(),
                    link_dl=NoopLinkDownloader(),
                    db=db,
                    categories={},
                    mailbox_key="account_a",
                )
                rows = db.get_all_invoices()

            standard_rows = [
                row for row in rows if row["invoice_type"] == "电子发票"
            ]
            pending_rows = [
                row for row in rows if row["invoice_type"] == "待关联证明材料"
            ]
            self.assertEqual(recorded, 3)
            self.assertEqual(len(standard_rows), 2)
            self.assertEqual(len(pending_rows), 1)
            self.assertTrue(all(json.loads(row["extra_paths"]) == [] for row in standard_rows))
            self.assertTrue((runtime / pending_rows[0]["attachment_path"]).exists())

    def test_multi_invoice_email_attaches_only_uniquely_named_evidence(self):
        class MappingParser:
            def parse_pdf(self, path):
                if Path(path).name == "invoice-a.pdf":
                    return InvoiceInfo(
                        invoice_number="MATCH-A-001",
                        invoice_date="2026-06-01",
                        total_amount="35.00",
                        seller_name="Synthetic Seller A",
                        invoice_type="电子发票",
                        parse_success=True,
                    )
                return InvoiceInfo(
                    invoice_number="MATCH-B-002",
                    invoice_date="2026-06-02",
                    total_amount="45.00",
                    seller_name="Synthetic Seller B",
                    invoice_type="电子发票",
                    parse_success=True,
                )

        with tempfile.TemporaryDirectory() as td:
            runtime = Path(td) / "runtime"
            attachments_dir = runtime / "attachments"
            attachments_dir.mkdir(parents=True)
            files = {
                "invoice-a.pdf": b"%PDF-invoice-a",
                "invoice-b.pdf": b"%PDF-invoice-b",
                "行程单_MATCH-A-001.pdf": b"%PDF-evidence-a",
                "行程单_MATCH-B-002.pdf": b"%PDF-evidence-b",
            }
            attachments = []
            for name, content in files.items():
                path = attachments_dir / name
                path.write_bytes(content)
                attachments.append(
                    Attachment(
                        file_path=str(path),
                        original_name=name,
                        content_type="application/pdf",
                        size=path.stat().st_size,
                        is_invoice=name.startswith("invoice-"),
                        is_extra=name.startswith("行程单_"),
                    )
                )
            msg = email.message.EmailMessage()
            msg["Subject"] = "两张电子发票及各自行程单"
            msg["From"] = "billing@example.com"
            msg["Date"] = "Mon, 1 Jun 2026 10:00:00 +0800"

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(
                cli, "RUNTIME_DIR", runtime
            ):
                cli._process_email(
                    msg=cli.MailMessage(uid=1902, raw_msg=msg),
                    att_handler=StaticAttachmentHandler(attachments_dir, attachments),
                    parser=MappingParser(),
                    link_dl=NoopLinkDownloader(),
                    db=db,
                    categories={},
                    mailbox_key="account_a",
                )
                rows = sorted(db.get_all_invoices(), key=lambda row: row["invoice_number"])

            self.assertEqual(len(rows), 2)
            for row in rows:
                extra_paths = json.loads(row["extra_paths"])
                self.assertEqual(len(extra_paths), 1)
                self.assertIn(row["invoice_number"], extra_paths[0])
                self.assertTrue((runtime / extra_paths[0]).exists())

    def test_multi_invoice_email_matches_evidence_by_unique_date_and_amount(self):
        invoice_a = InvoiceInfo(
            invoice_number="DATE-A-001",
            invoice_date="2026-06-01",
            total_amount="35.00",
            parse_success=True,
        )
        invoice_b = InvoiceInfo(
            invoice_number="DATE-B-002",
            invoice_date="2026-06-02",
            total_amount="45.00",
            parse_success=True,
        )
        extra = Attachment(
            file_path="滴滴行程单_2026-06-01_实付35元.pdf",
            original_name="滴滴行程单_2026-06-01_实付35元.pdf",
            content_type="application/pdf",
            size=1,
            is_extra=True,
        )

        matched, unmatched = cli._match_email_extras_to_invoices(
            [extra],
            [invoice_a, invoice_b],
        )

        self.assertEqual(matched[id(invoice_a)], [extra])
        self.assertNotIn(id(invoice_b), matched)
        self.assertEqual(unmatched, [])

    def test_multi_invoice_email_does_not_match_ambiguous_date_and_amount(self):
        invoice_a = InvoiceInfo(
            invoice_number="AMB-A-001",
            invoice_date="2026-06-01",
            total_amount="35.00",
            parse_success=True,
        )
        invoice_b = InvoiceInfo(
            invoice_number="AMB-B-002",
            invoice_date="2026-06-01",
            total_amount="35.00",
            parse_success=True,
        )
        extra = Attachment(
            file_path="滴滴行程单_2026-06-01_实付35元.pdf",
            original_name="滴滴行程单_2026-06-01_实付35元.pdf",
            content_type="application/pdf",
            size=1,
            is_extra=True,
        )

        matched, unmatched = cli._match_email_extras_to_invoices(
            [extra],
            [invoice_a, invoice_b],
        )

        self.assertEqual(matched, {})
        self.assertEqual(unmatched, [extra])

    def test_link_downloaded_invoice_with_email_extra_updates_extra_flags(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            attachments_dir = runtime / "attachments"
            attachments_dir.mkdir(parents=True)

            downloaded_file = base / "downloaded-invoice.pdf"
            downloaded_file.write_bytes(b"%PDF-downloaded-invoice")
            extra_file = base / "用车明细.pdf"
            extra_file.write_bytes(b"%PDF-trip-detail")

            class StaticLinkDownloader:
                def download_from_email(self, *args, **kwargs):
                    return [
                        DownloadedFile(
                            url="https://example.invalid/synthetic-invoice",
                            file_path=str(downloaded_file),
                            filename=downloaded_file.name,
                            size=downloaded_file.stat().st_size,
                            is_invoice=True,
                        )
                    ]

            msg = email.message.EmailMessage()
            msg["Subject"] = "电子发票及用车明细"
            msg["From"] = "billing@example.com"
            msg["Date"] = "Mon, 18 May 2026 10:00:00 +0800"
            attachment = Attachment(
                file_path=str(extra_file),
                original_name="用车明细.pdf",
                content_type="application/pdf",
                size=extra_file.stat().st_size,
                is_invoice=False,
                is_extra=True,
            )
            parser = StaticParser(InvoiceInfo(
                invoice_number="LINK-EXTRA-001",
                invoice_code="LINKCODE001",
                invoice_date="2026-05-18",
                total_amount="88.00",
                seller_name="Synthetic Mobility Seller",
                invoice_type="电子发票",
                parse_success=True,
            ))

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                recorded = cli._process_email(
                    msg=cli.MailMessage(uid=102, raw_msg=msg),
                    att_handler=StaticAttachmentHandler(attachments_dir, [attachment]),
                    parser=parser,
                    link_dl=StaticLinkDownloader(),
                    db=db,
                    categories={},
                )
                rows = db.get_all_invoices()

            self.assertEqual(recorded, 1)
            self.assertEqual(len(rows), 1)
            self.assertEqual(len(json.loads(rows[0]["extra_paths"])), 1)
            self.assertTrue(rows[0]["has_extra"])
            self.assertFalse(rows[0]["missing_extra"])

    def test_duplicate_email_scanning_does_not_duplicate_extra_paths(self):
        """测试 3：重复扫描同一封邮件不重复 extra_paths"""
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            runtime.mkdir()
            attachments_dir = runtime / "attachments"
            attachments_dir.mkdir()

            inv_file = attachments_dir / "invoice.pdf"
            inv_file.write_bytes(b"%PDF-invoice")
            iti_file = attachments_dir / "用车明细.pdf"
            iti_file.write_bytes(b"%PDF-iti")

            attachments = [
                Attachment(file_path=str(inv_file), original_name="invoice.pdf", content_type="application/pdf", size=1024, is_invoice=True, is_extra=False),
                Attachment(file_path=str(iti_file), original_name="用车明细.pdf", content_type="application/pdf", size=1024, is_invoice=False, is_extra=True),
            ]

            msg = email.message.EmailMessage()
            msg["Subject"] = "滴滴打车发票及明细"
            msg["From"] = "didi@example.com"
            msg["Date"] = "Mon, 18 May 2026 10:00:00 +0800"
            mail_msg = cli.MailMessage(uid=101, raw_msg=msg)

            parser = StaticParser(InvoiceInfo(
                invoice_number="DIDI999888",
                invoice_code="11002233",
                invoice_date="2025-12-20",
                total_amount="123.45",
                seller_name="南京滴滴出行科技有限公司",
                invoice_type="电子发票",
                parse_success=True,
            ))

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                handler1 = StaticAttachmentHandler(attachments_dir, attachments)
                cli._process_email(
                    msg=mail_msg,
                    att_handler=handler1,
                    parser=parser,
                    link_dl=NoopLinkDownloader(),
                    db=db,
                    categories={},
                )

                inv_file2 = attachments_dir / "invoice2.pdf"
                inv_file2.write_bytes(b"%PDF-invoice")
                iti_file2 = attachments_dir / "用车明细2.pdf"
                iti_file2.write_bytes(b"%PDF-iti")
                attachments2 = [
                    Attachment(file_path=str(inv_file2), original_name="invoice2.pdf", content_type="application/pdf", size=1024, is_invoice=True, is_extra=False),
                    Attachment(file_path=str(iti_file2), original_name="用车明细2.pdf", content_type="application/pdf", size=1024, is_invoice=False, is_extra=True),
                ]
                handler2 = StaticAttachmentHandler(attachments_dir, attachments2)
                cli._process_email(
                    msg=mail_msg,
                    att_handler=handler2,
                    parser=parser,
                    link_dl=NoopLinkDownloader(),
                    db=db,
                    categories={},
                )

                rows = db.get_all_invoices()

            self.assertEqual(len(rows), 1)
            extra_paths = json.loads(rows[0]["extra_paths"])
            self.assertEqual(len(extra_paths), 1)

    def test_itinerary_only_email_matches_existing_invoice(self):
        """测试 4：只有行程单邮件，能匹配已有发票"""
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            runtime.mkdir()
            attachments_dir = runtime / "attachments"
            attachments_dir.mkdir()

            iti_file = attachments_dir / "滴滴行程单_2025-12-20_1142.81.pdf"
            iti_file.write_bytes(b"%PDF-iti-content")

            attachments = [
                Attachment(file_path=str(iti_file), original_name="滴滴行程单_2025-12-20_1142.81.pdf", content_type="application/pdf", size=1024, is_invoice=False, is_extra=True),
            ]

            msg = email.message.EmailMessage()
            msg["Subject"] = "行程单"
            msg["From"] = "didi@example.com"
            msg["Date"] = "Mon, 18 May 2026 10:00:00 +0800"
            mail_msg = cli.MailMessage(uid=102, raw_msg=msg)

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                db.insert_invoice({
                    "invoice_number": "DIDI111222",
                    "invoice_date": "2025-12-20",
                    "total_amount": "1142.81",
                    "seller_name": "南京滴滴出行科技有限公司",
                    "category": "出租车",
                    "extra_paths": [],
                })

                handler = StaticAttachmentHandler(attachments_dir, attachments)
                parser = StaticParser(InvoiceInfo(parse_success=False, parse_note="滴滴行程单"))

                recorded = cli._process_email(
                    msg=mail_msg,
                    att_handler=handler,
                    parser=parser,
                    link_dl=NoopLinkDownloader(),
                    db=db,
                    categories={},
                )

                rows = db.get_all_invoices()

            self.assertEqual(recorded, 1)
            self.assertEqual(len(rows), 1)
            extra_paths = json.loads(rows[0]["extra_paths"])
            self.assertEqual(len(extra_paths), 1)
            self.assertIn("滴滴行程单", extra_paths[0])

    def test_unknown_failed_pdf_is_ignored(self):
        """测试 5：只有普通未知 PDF 邮件，不误判"""
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            runtime.mkdir()
            attachments_dir = runtime / "attachments"
            attachments_dir.mkdir()

            unknown_file = attachments_dir / "unknown.pdf"
            unknown_file.write_bytes(b"%PDF-unknown")

            attachments = [
                Attachment(file_path=str(unknown_file), original_name="unknown.pdf", content_type="application/pdf", size=1024, is_invoice=True, is_extra=False),
            ]

            msg = email.message.EmailMessage()
            msg["Subject"] = "未知邮件"
            msg["From"] = "stranger@example.com"
            msg["Date"] = "Mon, 18 May 2026 10:00:00 +0800"
            mail_msg = cli.MailMessage(uid=103, raw_msg=msg)

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                db.insert_invoice({
                    "invoice_number": "DIDI111222",
                    "invoice_date": "2025-12-20",
                    "total_amount": "1142.81",
                    "seller_name": "南京滴滴出行科技有限公司",
                    "category": "出租车",
                    "extra_paths": [],
                })

                handler = StaticAttachmentHandler(attachments_dir, attachments)
                parser = StaticParser(InvoiceInfo(parse_success=False, parse_note="解析失败"))

                recorded = cli._process_email(
                    msg=mail_msg,
                    att_handler=handler,
                    parser=parser,
                    link_dl=NoopLinkDownloader(),
                    db=db,
                    categories={},
                )

                rows = db.get_all_invoices()

            self.assertEqual(recorded, 0)
            self.assertEqual(len(rows), 1)
            extra_paths = json.loads(rows[0]["extra_paths"])
            self.assertEqual(len(extra_paths), 0)

    def test_failed_invoice_pdf_and_itinerary_extra(self):
        """单个解析失败 invoice_pdf + 一个行程单 extra"""
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            runtime.mkdir()
            attachments_dir = runtime / "attachments"
            attachments_dir.mkdir()

            inv_file = attachments_dir / "failed_invoice.pdf"
            inv_file.write_bytes(b"%PDF-failed")
            iti_file = attachments_dir / "滴滴行程单.pdf"
            iti_file.write_bytes(b"%PDF-iti")

            attachments = [
                Attachment(file_path=str(inv_file), original_name="failed_invoice.pdf", content_type="application/pdf", size=1024, is_invoice=True, is_extra=False),
                Attachment(file_path=str(iti_file), original_name="滴滴行程单.pdf", content_type="application/pdf", size=1024, is_invoice=False, is_extra=True),
            ]

            msg = email.message.EmailMessage()
            msg["Subject"] = "滴滴打车发票"
            msg["From"] = "didi@example.com"
            msg["Date"] = "Mon, 18 May 2026 10:00:00 +0800"
            mail_msg = cli.MailMessage(uid=201, raw_msg=msg)

            parser = StaticParser(InvoiceInfo(
                parse_success=False,
                parse_note="无法解析发票内容",
            ))
            handler = StaticAttachmentHandler(attachments_dir, attachments)

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                recorded = cli._process_email(
                    msg=mail_msg,
                    att_handler=handler,
                    parser=parser,
                    link_dl=NoopLinkDownloader(),
                    db=db,
                    categories={},
                )
                rows = db.get_all_invoices()

            self.assertEqual(recorded, 1)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["invoice_type"], "待关联证明材料")
            self.assertTrue((runtime / rows[0]["attachment_path"]).exists())

    def test_single_success_invoice_pdf_and_multiple_extras(self):
        """单个解析成功 invoice_pdf + 多个 extra，全部挂到该发票"""
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            runtime.mkdir()
            attachments_dir = runtime / "attachments"
            attachments_dir.mkdir()

            inv_file = attachments_dir / "invoice.pdf"
            inv_file.write_bytes(b"%PDF-invoice")
            iti_file1 = attachments_dir / "行程单1.pdf"
            iti_file1.write_bytes(b"%PDF-iti1")
            iti_file2 = attachments_dir / "用车明细2.pdf"
            iti_file2.write_bytes(b"%PDF-iti2")

            attachments = [
                Attachment(file_path=str(inv_file), original_name="invoice.pdf", content_type="application/pdf", size=1024, is_invoice=True, is_extra=False),
                Attachment(file_path=str(iti_file1), original_name="行程单1.pdf", content_type="application/pdf", size=1024, is_invoice=False, is_extra=True),
                Attachment(file_path=str(iti_file2), original_name="用车明细2.pdf", content_type="application/pdf", size=1024, is_invoice=False, is_extra=True),
            ]

            msg = email.message.EmailMessage()
            msg["Subject"] = "发票行程单"
            msg["From"] = "didi@example.com"
            msg["Date"] = "Mon, 18 May 2026 10:00:00 +0800"
            mail_msg = cli.MailMessage(uid=202, raw_msg=msg)

            parser = StaticParser(InvoiceInfo(
                invoice_number="DIDI777",
                invoice_code="220033",
                invoice_date="2025-12-20",
                total_amount="100.00",
                seller_name="滴滴出行",
                invoice_type="电子发票",
                parse_success=True,
            ))
            handler = StaticAttachmentHandler(attachments_dir, attachments)

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                recorded = cli._process_email(
                    msg=mail_msg,
                    att_handler=handler,
                    parser=parser,
                    link_dl=NoopLinkDownloader(),
                    db=db,
                    categories={},
                )
                rows = db.get_all_invoices()

            self.assertEqual(recorded, 1)
            self.assertEqual(len(rows), 1)
            extra_paths = json.loads(rows[0]["extra_paths"])
            self.assertEqual(len(extra_paths), 2)
            self.assertIn("DIDI777_ex.pdf", extra_paths[0].replace("\\", "/"))
            self.assertIn("DIDI777_ex_1.pdf", extra_paths[1].replace("\\", "/"))

    def test_multiple_success_invoice_pdfs_and_multiple_extras(self):
        """多个解析成功 invoice_pdf + 多个 extra，按发票号/日期金额唯一匹配，不能全量互挂"""
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            runtime.mkdir()
            attachments_dir = runtime / "attachments"
            attachments_dir.mkdir()

            inv_file1 = attachments_dir / "invoice1.pdf"
            inv_file1.write_bytes(b"%PDF-invoice1")
            inv_file2 = attachments_dir / "invoice2.pdf"
            inv_file2.write_bytes(b"%PDF-invoice2")

            iti_file1 = attachments_dir / "行程单_DIDI001.pdf"
            iti_file1.write_bytes(b"%PDF-iti1")

            iti_file2 = attachments_dir / "行程单_DIDI002.pdf"
            iti_file2.write_bytes(b"%PDF-iti2")

            attachments = [
                Attachment(file_path=str(inv_file1), original_name="invoice1.pdf", content_type="application/pdf", size=1024, is_invoice=True, is_extra=False),
                Attachment(file_path=str(inv_file2), original_name="invoice2.pdf", content_type="application/pdf", size=1024, is_invoice=True, is_extra=False),
                Attachment(file_path=str(iti_file1), original_name="行程单_DIDI001.pdf", content_type="application/pdf", size=1024, is_invoice=False, is_extra=True),
                Attachment(file_path=str(iti_file2), original_name="行程单_DIDI002.pdf", content_type="application/pdf", size=1024, is_invoice=False, is_extra=True),
            ]

            msg = email.message.EmailMessage()
            msg["Subject"] = "发票行程单"
            msg["From"] = "didi@example.com"
            msg["Date"] = "Mon, 18 May 2026 10:00:00 +0800"
            mail_msg = cli.MailMessage(uid=203, raw_msg=msg)

            class MultiStaticParser:
                def parse_pdf(self, path):
                    if "invoice1.pdf" in path:
                        return InvoiceInfo(
                            invoice_number="DIDI001",
                            invoice_code="111",
                            invoice_date="2025-12-20",
                            total_amount="100.00",
                            seller_name="滴滴出行",
                            invoice_type="电子发票",
                            parse_success=True,
                        )
                    else:
                        return InvoiceInfo(
                            invoice_number="DIDI002",
                            invoice_code="222",
                            invoice_date="2025-12-20",
                            total_amount="200.00",
                            seller_name="滴滴出行",
                            invoice_type="电子发票",
                            parse_success=True,
                        )

            handler = StaticAttachmentHandler(attachments_dir, attachments)

            with InvoiceDB(runtime / "invoices.db") as db, patch.object(cli, "RUNTIME_DIR", runtime):
                recorded = cli._process_email(
                    msg=mail_msg,
                    att_handler=handler,
                    parser=MultiStaticParser(),
                    link_dl=NoopLinkDownloader(),
                    db=db,
                    categories={},
                )
                rows = db.get_all_invoices()

            self.assertEqual(recorded, 2)
            self.assertEqual(len(rows), 2)

            rows = sorted(rows, key=lambda r: r["invoice_number"])
            self.assertEqual(rows[0]["invoice_number"], "DIDI001")
            self.assertEqual(rows[1]["invoice_number"], "DIDI002")

            extra_paths1 = json.loads(rows[0]["extra_paths"])
            extra_paths2 = json.loads(rows[1]["extra_paths"])

            self.assertEqual(len(extra_paths1), 1)
            self.assertEqual(len(extra_paths2), 1)

            self.assertIn("DIDI001_ex.pdf", extra_paths1[0].replace("\\", "/"))
            self.assertIn("DIDI002_ex.pdf", extra_paths2[0].replace("\\", "/"))

    def test_extract_item_names_cjk_asterisk(self):
        """测试从发票文本中提取前 3 个明细项目名称"""
        from scripts.invoice_fetch.invoice_parser import InvoiceParser
        parser = InvoiceParser()

        # Case 1: Standard header-based item extraction
        text1 = (
            "货物或应税劳务、服务名称   规格型号   单位   数量   单价   金额   税率   税额\n"
            "*餐饮服务*48元炒饭      无        份     1     48.00  48.00  6%     2.88\n"
            "*餐饮服务*可乐         无        瓶     1     10.00  10.00  6%     0.60\n"
            "合计                                                 58.00         3.48\n"
        )
        item_names1 = parser._extract_item_names(text1)
        self.assertEqual(item_names1, "*餐饮服务*48元炒饭, *餐饮服务*可乐")

        # Case 2: Max 3 items
        text2 = (
            "项目名称   金额\n"
            "项目A\n"
            "项目B\n"
            "项目C\n"
            "项目D\n"
            "合计\n"
        )
        item_names2 = parser._extract_item_names(text2)
        self.assertEqual(item_names2, "项目A, 项目B, 项目C")

        # Case 3: Whole-text scan fallback with asterisk patterns
        text3 = (
            "Some unrelated headers\n"
            "*餐饮*炒饭 其他干扰信息\n"
            "其他文本\n"
        )
        item_names3 = parser._extract_item_names(text3)
        self.assertEqual(item_names3, "*餐饮*炒饭")

    def test_classify_prioritizes_item_name(self):
        """测试消费类型分类优先级：项目名称强餐饮/交通关键词 > 销售方/主题"""
        # Define categories dict
        categories = {
            "transport": {"keywords": ["客运", "铁路"], "extra_name": ""},
            "dining": {"keywords": ["餐费", "餐饮"], "extra_name": ""},
        }

        # Case 1: Railway dining invoice (seller is railway, item is dining) -> "餐饮"
        cat, _, _ = cli._classify(
            subject="电子发票",
            sender="invoice@info.nuonuo.com",
            seller="北京京铁列车服务有限公司",
            categories=categories,
            item_name="*餐饮服务*炒饭",
            invoice_type="电子发票",
        )
        self.assertEqual(cat, "餐饮")

        # Case 2: Railway ticket invoice (seller is railway, item is transport) -> "交通"
        cat, _, _ = cli._classify(
            subject="电子发票",
            sender="12306@railway.com.cn",
            seller="中国国家铁路集团有限公司",
            categories=categories,
            item_name="*旅客运输服务*火车票",
            invoice_type="铁路电子客票",
        )
        self.assertEqual(cat, "交通")

        # Case 3: Empty item name fallback (based on seller keyword) -> "交通"
        cat_rail, _, _ = cli._classify(
            subject="电子发票",
            sender="12306@railway.com.cn",
            seller="中国国家铁路集团有限公司",
            categories=categories,
            item_name="",
        )
        self.assertEqual(cat_rail, "交通")

    def test_duplicate_backfill_dining_safe(self):
        """测试重复发票安全回填消费分类的逻辑"""
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            runtime = base / "runtime"
            runtime.mkdir()

            with InvoiceDB(runtime / "invoices.db") as db:
                # Insert V5 migration to ensure table is upgraded
                from scripts.invoice_fetch.migrations import check_and_migrate
                check_and_migrate(db._conn)

                # Case 1: Existing invoice is "其他", review_status is "to_review", should backfill to "餐饮"
                inv_id1 = db.insert_invoice({
                    "invoice_number": "FP001",
                    "total_amount": "100.00",
                    "seller_name": "列车服务公司",
                    "category": "其他",
                    "review_status": "to_review",
                    "item_name": "",
                })
                existing1 = db.get_invoice(inv_id1)
                refreshed1 = cli._refresh_invoice_from_parse(
                    db=db,
                    existing=existing1,
                    invoice_number="FP001",
                    invoice_code="111",
                    invoice_date="2026-05-18",
                    amount="100.00",
                    total_amount="100.00",
                    seller_name="列车服务公司",
                    buyer_name="公司",
                    invoice_type="电子发票",
                    category="餐饮",
                    has_extra=False,
                    extra_type="",
                    missing_extra=False,
                    parse_note="重复发票",
                    item_name="*餐饮服务*炒饭",
                )
                self.assertTrue(refreshed1)
                updated1 = db.get_invoice(inv_id1)
                self.assertEqual(updated1["category"], "餐饮")
                self.assertEqual(updated1["item_name"], "*餐饮服务*炒饭")

                # Case 2: Existing invoice has non-empty valid category (e.g. "交通"), should NOT backfill to "餐饮"
                inv_id2 = db.insert_invoice({
                    "invoice_number": "FP002",
                    "total_amount": "200.00",
                    "seller_name": "列车服务公司",
                    "category": "交通",
                    "review_status": "to_review",
                    "item_name": "",
                })
                existing2 = db.get_invoice(inv_id2)
                refreshed2 = cli._refresh_invoice_from_parse(
                    db=db,
                    existing=existing2,
                    invoice_number="FP002",
                    invoice_code="222",
                    invoice_date="2026-05-18",
                    amount="200.00",
                    total_amount="200.00",
                    seller_name="列车服务公司",
                    buyer_name="公司",
                    invoice_type="电子发票",
                    category="餐饮",
                    has_extra=False,
                    extra_type="",
                    missing_extra=False,
                    parse_note="重复发票",
                    item_name="*餐饮服务*餐费",
                )
                self.assertTrue(refreshed2)
                updated2 = db.get_invoice(inv_id2)
                self.assertEqual(updated2["category"], "交通")  # Preserved

                # Case 3: Existing invoice is approved/claimed, should NOT backfill to "餐饮"
                inv_id3 = db.insert_invoice({
                    "invoice_number": "FP003",
                    "total_amount": "300.00",
                    "seller_name": "列车服务公司",
                    "category": "未分类",
                    "review_status": "approved",
                    "item_name": "",
                })
                existing3 = db.get_invoice(inv_id3)
                refreshed3 = cli._refresh_invoice_from_parse(
                    db=db,
                    existing=existing3,
                    invoice_number="FP003",
                    invoice_code="333",
                    invoice_date="2026-05-18",
                    amount="300.00",
                    total_amount="300.00",
                    seller_name="列车服务公司",
                    buyer_name="公司",
                    invoice_type="电子发票",
                    category="餐饮",
                    has_extra=False,
                    extra_type="",
                    missing_extra=False,
                    parse_note="重复发票",
                    item_name="*餐饮服务*盒饭",
                )
                self.assertTrue(refreshed3)
                updated3 = db.get_invoice(inv_id3)
                self.assertEqual(updated3["category"], "未分类")  # Preserved due to approved review_status


if __name__ == "__main__":
    unittest.main()
