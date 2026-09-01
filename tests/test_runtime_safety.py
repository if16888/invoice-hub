from __future__ import annotations

import os
import threading
import tempfile
import unittest
from pathlib import Path
from unittest.mock import Mock, patch

from openpyxl import load_workbook

from scripts.invoice_fetch.db import InvoiceDB
from scripts.invoice_fetch.excel_export import export_excel


class ExcelFormulaInjectionTests(unittest.TestCase):
    def test_external_text_is_sanitized_on_every_export_sheet(self):
        row = {
            "invoice_number": "=1+1",
            "invoice_code": "normal subject",
            "expense_date": "2026-08-30",
            "invoice_date": "2026-08-29",
            "amount": "10.00",
            "total_amount": "20.00",
            "seller_name": "+SUM(A1:A2)",
            "buyer_name": "-1+2",
            "invoice_type": "普通发票",
            "category": "@SUM(A1:A2)",
            "has_extra": False,
            "missing_extra": True,
            "parse_success": 0,
            "parse_note": "=parse note",
            "mail_subject": "@mail subject",
            "mail_date": "2026-08-30",
            "attachment_path": "",
            "extra_paths": ["@evidence.pdf"],
            "download_url": "https://example.test/download/1",
            "confirmed_note": "+confirmed note",
            "warning": "中文普通文本",
            "review_status": "to_review",
        }

        with tempfile.TemporaryDirectory() as td:
            destination = Path(td) / "safe.xlsx"
            export_excel([row], destination)
            # Pass an explicitly scoped file handle so older openpyxl releases
            # cannot keep the Windows temporary file locked after loading.
            with destination.open("rb") as workbook_file:
                workbook = load_workbook(workbook_file, data_only=False)
            workbook.close()

        main = workbook["发票汇总"]
        summary = workbook["分类汇总"]
        exceptions = workbook["异常待处理"]

        # Main invoice sheet: every formula-looking source string is stored as text.
        self.assertEqual(main["A2"].value, "'=1+1")
        self.assertEqual(main["G2"].value, "'+SUM(A1:A2)")
        self.assertEqual(main["H2"].value, "'-1+2")
        self.assertEqual(main["J2"].value, "'@SUM(A1:A2)")
        self.assertEqual(main["M2"].value, "'=parse note")
        self.assertEqual(main["N2"].value, "'@mail subject")
        self.assertEqual(main["Q2"].value, "'@evidence.pdf")
        self.assertEqual(main["S2"].value, "'+confirmed note")
        self.assertEqual(main["B2"].value, "normal subject")
        self.assertEqual(main["T2"].value, "中文普通文本")

        # The category is also an external string on both the summary and exception sheets.
        self.assertEqual(summary["A2"].value, "'@SUM(A1:A2)")
        self.assertEqual(summary["B2"].value, 1)
        self.assertEqual(summary["C2"].value, 20.0)
        self.assertEqual(exceptions["A2"].value, "'=1+1")
        self.assertEqual(exceptions["E2"].value, "'@SUM(A1:A2)")
        self.assertEqual(exceptions["F2"].value, "'=parse note")
        self.assertEqual(exceptions["G2"].value, "'+confirmed note")
        self.assertEqual(exceptions["H2"].value, "'@evidence.pdf")
        self.assertEqual(exceptions["I2"].value, "'@mail subject")

        # A workbook-level check protects future sheets from bypassing the common sink.
        for sheet in workbook.worksheets:
            for cells in sheet.iter_rows():
                for cell in cells:
                    self.assertNotEqual(cell.data_type, "f", f"formula leaked in {sheet.title}!{cell.coordinate}")


class IntegrityTransactionTests(unittest.TestCase):
    @staticmethod
    def _invoice(number: str, total: str = "10.00", seller: str = "Seller") -> dict:
        return {
            "invoice_number": number,
            "total_amount": total,
            "seller_name": seller,
            "buyer_name": "Buyer",
            "invoice_date": "2026-08-30",
        }

    def test_duplicate_email_then_later_insert_succeeds(self):
        with tempfile.TemporaryDirectory() as td, InvoiceDB(Path(td) / "emails.db") as db:
            self.assertTrue(db.upsert_email(1, "first", "a@example.com", "2026-08-30"))
            self.assertFalse(db.upsert_email(1, "duplicate", "a@example.com", "2026-08-30"))
            self.assertTrue(db.upsert_email(2, "second", "b@example.com", "2026-08-30"))
            self.assertEqual(db._conn.execute("SELECT COUNT(*) FROM emails").fetchone()[0], 2)

    def test_duplicate_invoice_then_later_insert_succeeds(self):
        with tempfile.TemporaryDirectory() as td, InvoiceDB(Path(td) / "invoices.db") as db:
            first_id = db.insert_invoice(self._invoice("DUP-1"))
            self.assertIsNotNone(first_id)
            self.assertIsNone(db.insert_invoice(self._invoice("DUP-1")))
            next_id = db.insert_invoice(self._invoice("DUP-2"))
            self.assertIsNotNone(next_id)
            self.assertNotEqual(first_id, next_id)

    def test_update_unique_conflict_rolls_back_before_next_update(self):
        with tempfile.TemporaryDirectory() as td, InvoiceDB(Path(td) / "updates.db") as db:
            first_id = db.insert_invoice(self._invoice("UPDATE-1"))
            second_id = db.insert_invoice(self._invoice("UPDATE-2", total="20.00"))

            self.assertFalse(
                db.update_invoice_fields(
                    second_id,
                    invoice_number="UPDATE-1",
                    expense_date="2026-08-30",
                    seller_name="Seller",
                    total_amount="10.00",
                    category="其他",
                )
            )
            self.assertEqual(db.last_error, "unique_conflict")
            self.assertTrue(
                db.update_invoice_fields(
                    second_id,
                    invoice_number="UPDATE-2-OK",
                    expense_date="2026-08-30",
                    seller_name="Seller",
                    total_amount="20.00",
                    category="其他",
                )
            )
            self.assertEqual(db.get_invoice(first_id)["invoice_number"], "UPDATE-1")
            self.assertEqual(db.get_invoice(second_id)["invoice_number"], "UPDATE-2-OK")

    def test_parsed_metadata_conflict_rolls_back_before_next_update(self):
        def parsed_kwargs(number: str, total: str, seller: str) -> dict:
            return {
                "invoice_number": number,
                "invoice_code": "CODE",
                "invoice_date": "2026-08-30",
                "amount": total,
                "total_amount": total,
                "seller_name": seller,
                "buyer_name": "Buyer",
                "invoice_type": "普通发票",
                "category": "其他",
                "has_extra": False,
                "extra_type": "",
                "missing_extra": False,
                "parse_success": True,
                "parse_note": "",
                "item_name": "",
            }

        with tempfile.TemporaryDirectory() as td, InvoiceDB(Path(td) / "parsed.db") as db:
            first_id = db.insert_invoice(self._invoice("PARSED-1", total="30.00"))
            second_id = db.insert_invoice(self._invoice("PARSED-2", total="40.00"))

            self.assertFalse(
                db.update_invoice_parsed_metadata(
                    second_id,
                    **parsed_kwargs("PARSED-1", "30.00", "Seller"),
                )
            )
            self.assertEqual(db.last_error, "unique_conflict")
            self.assertTrue(
                db.update_invoice_parsed_metadata(
                    second_id,
                    **parsed_kwargs("PARSED-2-OK", "40.00", "Seller"),
                )
            )
            self.assertEqual(db.get_invoice(first_id)["invoice_number"], "PARSED-1")
            self.assertEqual(db.get_invoice(second_id)["invoice_number"], "PARSED-2-OK")

    def test_duplicate_claim_mapping_then_other_mapping_succeeds(self):
        with tempfile.TemporaryDirectory() as td, InvoiceDB(Path(td) / "claims.db") as db:
            claim_id = db.create_claim_group("Trip")
            first_id = db.insert_invoice(self._invoice("CLAIM-1"))
            second_id = db.insert_invoice(self._invoice("CLAIM-2", total="20.00"))

            self.assertTrue(db.add_invoice_to_claim(claim_id, first_id, "first"))
            self.assertFalse(db.add_invoice_to_claim(claim_id, first_id, "duplicate"))
            self.assertEqual(db.last_error, "integrity_error")
            self.assertTrue(db.add_invoice_to_claim(claim_id, second_id, "second"))
            self.assertEqual(
                {row["invoice_number"] for row in db.get_claim_invoices(claim_id)},
                {"CLAIM-1", "CLAIM-2"},
            )


class _ControlledWorker:
    """Small deterministic worker double for close-order tests."""

    def __init__(self, window, *, running: bool, cancellable: bool):
        self.window = window
        self._running = running
        self.cancellable = cancellable
        self.wait_calls = 0
        self.cancel_calls = 0
        self.db_open_during_wait: list[bool] = []
        self.blocked = []
        if cancellable:
            self.request_cancel = self._request_cancel

    def isRunning(self):
        return self._running

    def blockSignals(self, blocked: bool):
        self.blocked.append(bool(blocked))

    def _request_cancel(self):
        self.cancel_calls += 1

    def wait(self):
        self.wait_calls += 1
        self.db_open_during_wait.append(self.window.db.is_open)
        self._running = False
        return True


class WorkerShutdownTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
        try:
            from PySide6.QtCore import QCoreApplication
            from PySide6.QtCore import Qt
            from PySide6.QtGui import QCloseEvent
            from PySide6.QtWidgets import QApplication
            from scripts.invoice_fetch.gui.app import InvoiceReviewApp
            from scripts.invoice_fetch.gui.workers import EmailScanWorker, LocalImportWorker
        except Exception as exc:  # pragma: no cover - depends on host GUI libraries
            raise unittest.SkipTest(f"Qt GUI unavailable: {exc}")
        cls._QApplication = QApplication
        cls._QCoreApplication = QCoreApplication
        cls._QCloseEvent = QCloseEvent
        cls._Qt = Qt
        cls._InvoiceReviewApp = InvoiceReviewApp
        cls._EmailScanWorker = EmailScanWorker
        cls._LocalImportWorker = LocalImportWorker
        cls._qt_app = QApplication.instance() or QApplication([])

        class QueuedLocalImportWorker(LocalImportWorker):
            def __init__(self, import_dir, db_path, payload):
                super().__init__(import_dir, db_path)
                self.payload = dict(payload)
                self.emitted = threading.Event()
                self.release = threading.Event()
                self.wait_calls = 0
                self.db_open_during_wait = []
                self.window = None

            def run(self):
                self.finished.emit(self.payload)
                self.emitted.set()
                self.release.wait()

            def wait(self, *args, **kwargs):
                self.wait_calls += 1
                if self.window is not None:
                    self.db_open_during_wait.append(self.window.db.is_open)
                self.release.set()
                return super().wait(*args, **kwargs)

        class QueuedEmailScanWorker(EmailScanWorker):
            def __init__(self, db_path, signal_name, payload):
                super().__init__(db_path)
                self.signal_name = signal_name
                self.payload = payload
                self.emitted = threading.Event()
                self.release = threading.Event()
                self.wait_calls = 0
                self.db_open_during_wait = []
                self.window = None

            def run(self):
                getattr(self, self.signal_name).emit(self.payload)
                self.emitted.set()
                self.release.wait()

            def wait(self, *args, **kwargs):
                self.wait_calls += 1
                if self.window is not None:
                    self.db_open_during_wait.append(self.window.db.is_open)
                self.release.set()
                return super().wait(*args, **kwargs)

        cls._QueuedLocalImportWorker = QueuedLocalImportWorker
        cls._QueuedEmailScanWorker = QueuedEmailScanWorker

    def _window(self, td: str):
        window = self._InvoiceReviewApp(Path(td) / "shutdown.db", startup_probe=True)
        # startup_probe intentionally skips the full UI; these are the settings
        # attributes closeEvent normally receives from the full constructor.
        window._nav_collapsed_manual = None
        window._close_pending = False
        return window

    def _close(self, window):
        event = self._QCloseEvent()
        window.closeEvent(event)
        self.assertTrue(event.isAccepted())
        self._QCoreApplication.processEvents()
        window.deleteLater()
        self._QCoreApplication.processEvents()

    def _close_without_processing(self, window):
        event = self._QCloseEvent()
        window.closeEvent(event)
        self.assertTrue(event.isAccepted())

    @staticmethod
    def _install_callback_probe(window):
        callback_checks = []
        business_calls = []
        end_calls = []
        original_end = window._end_data_operation

        def callback_allowed():
            callback_checks.append(bool(window._shutdown_requested))
            return not window._shutdown_requested

        def end_operation(operation):
            end_calls.append(operation)
            original_end(operation)

        window._worker_callback_allowed = callback_allowed
        window._end_data_operation = end_operation
        for name in (
            "_clear_action_busy",
            "_record_import_activity",
            "_performance_completion_call",
            "_refresh_visible_completion_page",
            "_refresh_imports_page",
            "_load_invoices",
            "write_log",
        ):
            setattr(window, name, Mock(side_effect=lambda *args, _name=name, **kwargs: business_calls.append(_name)))
        window.btn_import_local = Mock()
        window.btn_scan_email = Mock()
        return callback_checks, business_calls, end_calls

    def _assert_queued_callback_is_ignored(self, window, worker, signal, handler, operation):
        callback_checks, business_calls, end_calls = self._install_callback_probe(window)
        worker.window = window
        self.assertTrue(window._try_begin_data_operation(operation, notify=False))
        if operation == "本地导入":
            window.import_worker = worker
        else:
            window.scan_worker = worker
        signal.connect(handler, self._Qt.ConnectionType.QueuedConnection)
        worker.start()
        self.assertTrue(worker.emitted.wait(5), "worker did not enqueue its callback")

        self._close_without_processing(window)
        self.assertFalse(window.db.is_open)
        self.assertEqual(worker.db_open_during_wait, [True])
        self.assertEqual(worker.wait_calls, 1)
        self.assertEqual(end_calls, [operation])

        # The signal was emitted before closeEvent and is delivered only now.
        self._QCoreApplication.processEvents()
        self.assertEqual(callback_checks, [True])
        self.assertEqual(business_calls, [])
        self.assertEqual(window._data_operation_gate.owner, "")

        worker.deleteLater()
        window.deleteLater()
        self._QCoreApplication.processEvents()

    def test_queued_local_import_finished_is_ignored_after_shutdown(self):
        with tempfile.TemporaryDirectory() as td:
            window = self._window(td)
            worker = self._QueuedLocalImportWorker(
                Path(td),
                Path(td) / "shutdown.db",
                {"added": 1, "duplicates": 0, "conflicts": 0, "pending_manual": 0, "failed": 0},
            )
            self._assert_queued_callback_is_ignored(
                window,
                worker,
                worker.finished,
                window._import_local_finished,
                "本地导入",
            )

    def test_queued_email_finished_is_ignored_after_shutdown(self):
        with tempfile.TemporaryDirectory() as td:
            window = self._window(td)
            worker = self._QueuedEmailScanWorker(
                Path(td) / "shutdown.db",
                "finished",
                {"cancelled": True},
            )
            self._assert_queued_callback_is_ignored(
                window,
                worker,
                worker.finished,
                window._scan_email_finished,
                "邮箱扫描",
            )

    def test_queued_email_error_is_ignored_after_shutdown(self):
        with tempfile.TemporaryDirectory() as td:
            window = self._window(td)
            worker = self._QueuedEmailScanWorker(
                Path(td) / "shutdown.db",
                "error",
                "synthetic scan failure",
            )
            with patch("scripts.invoice_fetch.gui.app.QMessageBox.critical"):
                self._assert_queued_callback_is_ignored(
                    window,
                    worker,
                    worker.error,
                    window._scan_email_error,
                    "邮箱扫描",
                )

    def test_local_import_waits_before_database_close(self):
        with tempfile.TemporaryDirectory() as td:
            window = self._window(td)
            worker = _ControlledWorker(window, running=True, cancellable=False)
            self.assertTrue(window._try_begin_data_operation("本地导入", notify=False))
            window.import_worker = worker
            self._close(window)
            self.assertEqual(worker.cancel_calls, 0)
            self.assertEqual(worker.wait_calls, 1)
            self.assertEqual(worker.db_open_during_wait, [True])
            self.assertFalse(window.db.is_open)
            self.assertEqual(window._data_operation_gate.owner, "")
            self.assertFalse(window._try_begin_data_operation("本地导入", notify=False))

    def test_local_import_shutdown_cancels_at_source_boundary(self):
        with tempfile.TemporaryDirectory() as td:
            base = Path(td)
            source_dir = base / "source"
            source_dir.mkdir()
            for name in ("A.pdf", "B.pdf", "C.pdf"):
                (source_dir / name).write_bytes(f"%PDF-{name}".encode("ascii"))

            window = self._window(td)
            started = threading.Event()
            release = threading.Event()
            processed = []
            observed = {}
            db_open_during_wait = []

            from scripts.invoice_fetch import services

            real_import = services.import_local_directory

            def wrapped_import(import_dir, db_path, *, scan_control):
                observed["control"] = scan_control
                return real_import(import_dir, db_path, scan_control=scan_control)

            def fake_pdf(source_name, *_args, **_kwargs):
                processed.append(source_name)
                if source_name == "A.pdf":
                    started.set()
                    release.wait()
                return services.LocalImportItemResult(
                    status="added",
                    invoice_id=len(processed),
                    created=True,
                    reviewable=True,
                )

            class ControlledLocalImportWorker(self._LocalImportWorker):
                def request_cancel(self):
                    super().request_cancel()
                    release.set()

                def wait(self, *args, **kwargs):
                    db_open_during_wait.append(window.db.is_open)
                    return super().wait(*args, **kwargs)

            worker = ControlledLocalImportWorker(source_dir, base / "shutdown.db")
            self.assertTrue(window._try_begin_data_operation("本地导入", notify=False))
            window.import_worker = worker

            with patch.object(services, "RUNTIME_DIR", base / "runtime"), patch.object(
                services, "InvoiceParser", return_value=Mock()
            ), patch.object(services, "_import_local_pdf", side_effect=fake_pdf), patch.object(
                services, "import_local_directory", side_effect=wrapped_import
            ):
                worker.start()
                try:
                    self.assertTrue(started.wait(5), "local import did not enter the controlled source")
                    self.assertTrue(window.db.is_open)

                    event = self._QCloseEvent()
                    window.closeEvent(event)

                    self.assertTrue(event.isAccepted())
                    self.assertTrue(release.is_set())
                    self.assertFalse(worker.isRunning())
                    self.assertEqual(processed, ["A.pdf"])
                    self.assertIs(observed["control"], worker.control)
                    self.assertTrue(worker.control.cancelled)
                    self.assertEqual(db_open_during_wait, [True])
                    self.assertFalse(window.db.is_open)
                    self.assertEqual(window._data_operation_gate.owner, "")
                finally:
                    release.set()
                    if worker.isRunning():
                        worker.wait(5000)

            window.deleteLater()
            self._QCoreApplication.processEvents()

    def test_scan_requests_cancel_then_waits_before_database_close(self):
        with tempfile.TemporaryDirectory() as td:
            window = self._window(td)
            worker = _ControlledWorker(window, running=True, cancellable=True)
            self.assertTrue(window._try_begin_data_operation("邮箱扫描", notify=False))
            window.scan_worker = worker
            self._close(window)
            self.assertEqual(worker.cancel_calls, 1)
            self.assertEqual(worker.wait_calls, 1)
            self.assertEqual(worker.db_open_during_wait, [True])
            self.assertFalse(window.db.is_open)

    def test_no_active_worker_closes_immediately(self):
        with tempfile.TemporaryDirectory() as td:
            window = self._window(td)
            self._close(window)
            self.assertFalse(window.db.is_open)

    def test_finished_worker_is_not_waited_again(self):
        with tempfile.TemporaryDirectory() as td:
            window = self._window(td)
            worker = _ControlledWorker(window, running=False, cancellable=False)
            self.assertTrue(window._try_begin_data_operation("本地导入", notify=False))
            window.import_worker = worker
            self._close(window)
            self.assertEqual(worker.wait_calls, 0)
            self.assertFalse(window.db.is_open)
            self.assertEqual(window._data_operation_gate.owner, "")


if __name__ == "__main__":
    unittest.main()
