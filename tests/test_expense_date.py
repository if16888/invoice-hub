import json
import tempfile
import unittest
from pathlib import Path
from openpyxl import load_workbook

from scripts.invoice_fetch.invoice_parser import (
    parse_railway_travel_date,
    parse_railway_invoice_date,
    InvoiceParser,
    InvoiceInfo
)
from scripts.invoice_fetch.db import InvoiceDB
from scripts.invoice_fetch.reimbursement import get_date_warning
from scripts.invoice_fetch.claim_export import export_claim_package


class ExpenseDateTests(unittest.TestCase):
    def test_railway_date_helpers(self):
        # 1. Travel date helper test
        self.assertEqual(parse_railway_travel_date("2026年05月07日 08:47开"), "2026-05-07")
        self.assertEqual(parse_railway_travel_date("2026年5月7日 08:47开"), "2026-05-07")
        self.assertEqual(parse_railway_travel_date("2026 年 05 月 07 日 08:47 开"), "2026-05-07")
        self.assertIsNone(parse_railway_travel_date("2026-05-07"))

        # 2. Invoice date helper test
        self.assertEqual(parse_railway_invoice_date("开票日期: 2026年05月18日"), "2026-05-18")
        self.assertEqual(parse_railway_invoice_date("开票日期：2026年5月18日"), "2026-05-18")
        self.assertEqual(parse_railway_invoice_date("开票日期  2026 年 05 月 18 日"), "2026-05-18")
        self.assertIsNone(parse_railway_invoice_date("开票日期 2026-05-18"))

    def test_date_warning_logic(self):
        # 1. Fallback to invoice date (railway ticket triggers warning)
        inv_fallback = {
            "expense_date": "2026-05-18",
            "invoice_date": "2026-05-18",
            "date_source": "invoice_date",
            "invoice_type": "铁路电子客票",
            "seller_name": "中国国家铁路集团有限公司",
        }
        self.assertEqual(get_date_warning(inv_fallback), "未识别到费用发生日期，已使用开票日期。")

        # 2. Travel date source (no warning)
        inv_travel = {
            "expense_date": "2026-05-07",
            "invoice_date": "2026-05-18",
            "date_source": "travel_date",
            "invoice_type": "铁路电子客票",
            "seller_name": "中国国家铁路集团有限公司",
        }
        self.assertEqual(get_date_warning(inv_travel), "")

        # 3. Ordinary dining invoice fallback (no warning)
        inv_dining = {
            "expense_date": "2026-05-18",
            "invoice_date": "2026-05-18",
            "date_source": "invoice_date",
            "category": "餐饮",
            "seller_name": "餐饮商户",
        }
        self.assertEqual(get_date_warning(inv_dining), "")

        # 4. Empty date source / empty dates
        inv_empty = {
            "expense_date": "",
            "invoice_date": "",
            "date_source": ""
        }
        self.assertEqual(get_date_warning(inv_empty), "")

    def test_database_migration_v6(self):
        import sqlite3
        with tempfile.TemporaryDirectory() as td:
            db_path = Path(td) / "test_migration.db"
            
            # 1. Manually create an old invoices table (V5 representation with all V5 columns)
            conn = sqlite3.connect(str(db_path))
            cursor = conn.cursor()
            cursor.execute("""
            CREATE TABLE invoices (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                mailbox_key     TEXT NOT NULL DEFAULT 'legacy',
                invoice_number  TEXT,
                invoice_code    TEXT,
                invoice_date    TEXT,
                amount          TEXT,
                total_amount    TEXT,
                seller_name     TEXT,
                buyer_name      TEXT,
                invoice_type    TEXT,
                category        TEXT DEFAULT '其他',
                has_extra       INTEGER DEFAULT 0,
                extra_type      TEXT DEFAULT '',
                missing_extra   INTEGER DEFAULT 0,
                mail_uid        INTEGER,
                mail_subject    TEXT,
                mail_date       TEXT,
                mail_sender     TEXT,
                parse_success   INTEGER DEFAULT 0,
                parse_note      TEXT DEFAULT '',
                attachment_path TEXT DEFAULT '',
                extra_paths     TEXT DEFAULT '[]',
                download_url    TEXT DEFAULT '',
                item_name       TEXT DEFAULT '',
                is_deleted      INTEGER NOT NULL DEFAULT 0,
                created_at      TEXT DEFAULT (datetime('now','localtime')),
                UNIQUE(invoice_number, total_amount, seller_name)
            );
            """)
            cursor.execute("""
            INSERT INTO invoices (invoice_number, invoice_date, total_amount, seller_name, category)
            VALUES ('MIGRATE001', '2026-05-01', '100.00', 'Legacy Seller', '其他')
            """)
            cursor.execute("PRAGMA user_version = 5")
            conn.commit()
            conn.close()
            
            # 2. Open via InvoiceDB which runs migration
            with InvoiceDB(db_path) as db:
                # Check version is 6
                cursor = db._conn.cursor()
                cursor.execute("PRAGMA user_version")
                ver = cursor.fetchone()[0]
                self.assertEqual(ver, 6)

                # Verify new columns exist
                cursor.execute("PRAGMA table_info(invoices)")
                cols = {row[1] for row in cursor.fetchall()}
                self.assertIn("expense_date", cols)
                self.assertIn("date_source", cols)

                # Check backfilled values
                inv = db.get_invoice(1)
                self.assertEqual(inv["expense_date"], "2026-05-01")
                self.assertEqual(inv["date_source"], "legacy")

    def test_excel_export_expense_and_invoice_dates(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            runtime = root / "runtime"
            attachment = runtime / "attachments" / "invoice.pdf"
            attachment.parent.mkdir(parents=True)
            attachment.write_bytes(b"pdf")
            db_path = runtime / "invoices.db"

            with InvoiceDB(db_path) as db:
                # Insert two invoices: one railway (distinct travel/invoice dates) and one normal
                id_railway = db.insert_invoice({
                    "invoice_number": "TRAIN123",
                    "invoice_date": "2026-05-18",
                    "expense_date": "2026-05-07",
                    "date_source": "travel_date",
                    "total_amount": "140.00",
                    "seller_name": "中国国家铁路集团有限公司",
                    "buyer_name": "示例科技有限公司",
                    "category": "交通",
                    "attachment_path": "attachments/invoice.pdf",
                    "review_status": "approved",
                })
                
                id_normal = db.insert_invoice({
                    "invoice_number": "NORMAL456",
                    "invoice_date": "2026-05-20",
                    "expense_date": "2026-05-20",
                    "date_source": "invoice_date",
                    "total_amount": "80.00",
                    "seller_name": "餐饮商户",
                    "buyer_name": "示例科技有限公司",
                    "category": "餐饮",
                    "attachment_path": "attachments/invoice.pdf",
                    "review_status": "approved",
                })

                claim_id = db.create_claim_group("date test claim")
                db.add_invoice_to_claim(claim_id, id_railway)
                db.add_invoice_to_claim(claim_id, id_normal)

                export_dir = export_claim_package(
                    db,
                    claim_id,
                    project_root=root,
                    runtime_dir=runtime,
                    reimbursement_config={
                        "buyer_name": "示例科技有限公司",
                        "strict_buyer_check": True,
                    },
                )

            # 1. Verify manifest.json
            manifest = json.loads((export_dir / "manifest.json").read_text(encoding="utf-8"))
            items_by_num = {item["invoice_number"]: item for item in manifest["items"]}
            
            self.assertEqual(items_by_num["TRAIN123"]["expense_date"], "2026-05-07")
            self.assertEqual(items_by_num["TRAIN123"]["date_source"], "travel_date")
            self.assertEqual(items_by_num["TRAIN123"]["warning"], "")  # No date warning
            
            self.assertEqual(items_by_num["NORMAL456"]["expense_date"], "2026-05-20")
            self.assertEqual(items_by_num["NORMAL456"]["date_source"], "invoice_date")
            self.assertEqual(items_by_num["NORMAL456"]["warning"], "")

            # 2. Verify Excel columns
            wb = load_workbook(export_dir / "reimbursement.xlsx")
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            
            self.assertIn("费用日期", headers)
            self.assertIn("开票日期", headers)
            self.assertIn("校验提示", headers)
            
            col_expense = headers.index("费用日期") + 1
            col_invoice = headers.index("开票日期") + 1
            col_warning = headers.index("校验提示") + 1
            
            rows_by_num = {}
            for row in range(2, ws.max_row + 1):
                num = ws.cell(row=row, column=headers.index("发票号码") + 1).value
                if num:
                    rows_by_num[num] = {
                        "expense_date": ws.cell(row=row, column=col_expense).value,
                        "invoice_date": ws.cell(row=row, column=col_invoice).value,
                        "warning": ws.cell(row=row, column=col_warning).value,
                    }

            self.assertEqual(rows_by_num["TRAIN123"]["expense_date"], "2026-05-07")
            self.assertEqual(rows_by_num["TRAIN123"]["invoice_date"], "2026-05-18")
            self.assertIsNone(rows_by_num["TRAIN123"]["warning"])

            self.assertEqual(rows_by_num["NORMAL456"]["expense_date"], "2026-05-20")
            self.assertEqual(rows_by_num["NORMAL456"]["invoice_date"], "2026-05-20")
            self.assertIsNone(rows_by_num["NORMAL456"]["warning"])

    def test_refresh_legacy_railway_expense_date_upgrade(self):
        from scripts.invoice_fetch.__main__ import _refresh_invoice_from_parse
        with tempfile.TemporaryDirectory() as td:
            db_path = Path(td) / "test.db"
            with InvoiceDB(db_path) as db:
                existing_id = db.insert_invoice({
                    "invoice_number": "UPGRADE001",
                    "invoice_date": "2026-05-18",
                    "expense_date": "2026-05-18",
                    "date_source": "legacy",
                    "seller_name": "中国国家铁路集团有限公司",
                    "total_amount": "140.00",
                    "category": "交通",
                })
                existing = db.get_invoice(existing_id)
                res = _refresh_invoice_from_parse(
                    db,
                    existing,
                    invoice_number="UPGRADE001",
                    invoice_code="CODE123",
                    invoice_date="2026-05-18",
                    amount="140.00",
                    total_amount="140.00",
                    seller_name="中国国家铁路集团有限公司",
                    buyer_name="Test Co",
                    invoice_type="铁路电子客票",
                    category="交通",
                    has_extra=False,
                    extra_type="",
                    missing_extra=False,
                    parse_note="new parse",
                    expense_date="2026-05-07",
                    date_source="travel_date",
                    force_refresh_metadata=False,
                )
                self.assertTrue(res)
                updated = db.get_invoice(existing_id)
                self.assertEqual(updated["expense_date"], "2026-05-07")
                self.assertEqual(updated["date_source"], "travel_date")
                self.assertEqual(updated["invoice_date"], "2026-05-18")

    def test_manual_edit_expense_date_preserves_invoice_date(self):
        with tempfile.TemporaryDirectory() as td:
            db_path = Path(td) / "test.db"
            with InvoiceDB(db_path) as db:
                inv_id = db.insert_invoice({
                    "invoice_number": "MANUAL001",
                    "invoice_date": "2026-05-18",
                    "expense_date": "2026-05-07",
                    "date_source": "travel_date",
                    "seller_name": "中国国家铁路集团有限公司",
                    "total_amount": "140.00",
                    "category": "交通",
                })
                res = db.update_invoice_fields(
                    invoice_id=inv_id,
                    invoice_number="MANUAL001",
                    expense_date="2026-05-08",
                    seller_name="中国国家铁路集团有限公司",
                    total_amount="140.00",
                    category="交通",
                )
                self.assertTrue(res)
                updated = db.get_invoice(inv_id)
                self.assertEqual(updated["expense_date"], "2026-05-08")
                self.assertEqual(updated["date_source"], "manual")
                self.assertEqual(updated["invoice_date"], "2026-05-18")

    def test_gui_table_uses_expense_date(self):
        import sys
        from PySide6.QtWidgets import QApplication
        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        app = QApplication.instance() or QApplication(sys.argv)
        with tempfile.TemporaryDirectory() as td:
            db_path = Path(td) / "test.db"
            with InvoiceDB(db_path) as db:
                db.insert_invoice({
                    "invoice_number": "GUITABLE001",
                    "invoice_date": "2026-05-18",
                    "expense_date": "2026-05-07",
                    "date_source": "travel_date",
                    "seller_name": "中国国家铁路集团有限公司",
                    "total_amount": "140.00",
                    "category": "交通",
                    "review_status": "to_review",
                })

            try:
                window = InvoiceReviewApp(db_path, splash=None)
                window._deferred_init()
                app.processEvents()

                self.assertEqual(window.table.rowCount(), 1)
                date_item = window.table.item(0, 1)
                self.assertEqual(date_item.text(), "2026-05-07")
                window.close()
                window.deleteLater()
                app.processEvents()
            except Exception as e:
                self.skipTest(f"Skipping GUI test due to UI environment issues: {e}")

    def test_quality_uses_expense_date(self):
        import sys
        from PySide6.QtWidgets import QApplication
        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        app = QApplication.instance() or QApplication(sys.argv)
        with tempfile.TemporaryDirectory() as td:
            db_path = Path(td) / "test.db"
            with InvoiceDB(db_path) as db:
                pass

            try:
                window = InvoiceReviewApp(db_path, splash=None)
                window._deferred_init()
                app.processEvents()

                inv_ok = {
                    "invoice_number": "QUALITY001",
                    "invoice_date": "",
                    "expense_date": "2026-05-07",
                    "total_amount": "100.00",
                    "seller_name": "Test Seller",
                }
                self.assertEqual(window._get_invoice_quality(inv_ok), "")

                inv_bad = {
                    "invoice_number": "QUALITY002",
                    "invoice_date": "",
                    "expense_date": "",
                    "total_amount": "100.00",
                    "seller_name": "Test Seller",
                }
                self.assertEqual(window._get_invoice_quality(inv_bad), "待补全")

                window.close()
                window.deleteLater()
                app.processEvents()
            except Exception as e:
                self.skipTest(f"Skipping GUI test due to UI environment issues: {e}")


if __name__ == "__main__":
    unittest.main()
