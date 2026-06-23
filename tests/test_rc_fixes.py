# -*- coding: utf-8 -*-
"""Unit tests covering the 4 P1 fixes in Invoice Hub v0.1.3-rc2."""

import os
import sys
import tempfile
import unittest
import json
import gc
from pathlib import Path
from unittest.mock import MagicMock, patch

try:
    from PySide6.QtWidgets import QApplication, QMessageBox, QAbstractItemView
    from PySide6.QtCore import Qt, QItemSelectionModel
    from PySide6.QtTest import QTest
    _HAS_PYSIDE = True
except ImportError:
    _HAS_PYSIDE = False

from scripts.invoice_fetch.attachment_handler import build_managed_attachment_name
from scripts.invoice_fetch.excel_export import export_excel
from scripts.invoice_fetch.db import InvoiceDB

_QAPP = None


def _get_or_create_app():
    global _QAPP
    if not _HAS_PYSIDE:
        return None
    _QAPP = QApplication.instance() or QApplication(sys.argv)
    return _QAPP


class TestNamingRules(unittest.TestCase):
    """Tests for Fix 3: full-format naming with fallbacks for missing fields."""

    def test_naming_with_full_details(self):
        # YYYY-MM-DD_消费类型_金额_发票号_原件/证明材料.ext
        name = build_managed_attachment_name(
            original_name="my_invoice.pdf",
            invoice_date="2026-06-13",
            expense_date="2026-06-12",
            fallback_date="2026-06-11",
            category="餐饮",
            total_amount="123.45",
            invoice_number="98765432",
            role="原件"
        )
        self.assertEqual(name, "2026-06-12_餐饮_123.45_98765432_原件.pdf")

    def test_naming_with_missing_details_fallbacks(self):
        # Empty category -> 未分类, empty amount -> 金额待补全, empty number -> 待补全, empty role -> 原件
        name = build_managed_attachment_name(
            original_name="test.ofd",
            invoice_date="2026-06-13",
            expense_date=None,
            category=None,
            total_amount=None,
            invoice_number=None,
            role=None
        )
        self.assertEqual(name, "2026-06-13_test.ofd")

    def test_naming_partial_fields(self):
        # If at least one of category, amount, number is present, it uses full format
        name = build_managed_attachment_name(
            original_name="receipt.png",
            invoice_date="2026-06-13",
            category="交通",
            role="证明材料"
        )
        self.assertEqual(name, "2026-06-13_交通_金额待补全_待补全_证明材料.png")


class TestExcelSorting(unittest.TestCase):
    """Tests for Fix 4: Excel export stable sort (empty dates last)."""

    def test_excel_export_sorting_logic(self):
        rows = [
            {"id": 1, "expense_date": "", "invoice_date": "", "mail_date": "", "created_at": "", "invoice_number": "1"},
            {"id": 2, "expense_date": "2026-06-12", "invoice_number": "2"},
            {"id": 3, "expense_date": "", "invoice_date": "2026-06-10", "invoice_number": "3"},
            {"id": 4, "expense_date": "", "invoice_date": "", "mail_date": "2026-06-08", "invoice_number": "4"},
            {"id": 5, "expense_date": "", "invoice_date": "", "mail_date": "", "created_at": "2026-06-05", "invoice_number": "5"},
        ]
        # Copy to check non-mutation
        orig_rows = list(rows)

        with tempfile.TemporaryDirectory() as td:
            dest = Path(td) / "test.xlsx"
            export_excel(rows, dest)

            # Ensure original list not mutated
            self.assertEqual(rows, orig_rows)

            # Read back using openpyxl to verify order
            from openpyxl import load_workbook
            wb = load_workbook(dest)
            ws = wb.active

            # Row 1 is header. Rows 2-6 should be data
            # Target sorted order:
            # 1. 2026-06-05 (ID 5, created_at)
            # 2. 2026-06-08 (ID 4, mail_date)
            # 3. 2026-06-10 (ID 3, invoice_date)
            # 4. 2026-06-12 (ID 2, expense_date)
            # 5. Empty date (ID 1, sort sentinel 9999-12-31)
            row_vals = []
            for r in range(2, 7):
                inv_num = ws.cell(row=r, column=1).value  # col 1 is 发票号码
                row_vals.append(inv_num)

            # Assert order: "5", "4", "3", "2", "1"
            self.assertEqual(row_vals, ["5", "4", "3", "2", "1"])


class TestGUIFixes(unittest.TestCase):
    """Tests for Fix 1 (toolbar busy transitions) and Fix 2 (multi-row delete selection & de-dup)."""

    def setUp(self):
        if not _HAS_PYSIDE:
            self.skipTest("PySide6 not available")
        self.app = _get_or_create_app()
        self.temp_dir = tempfile.TemporaryDirectory()
        self.db_path = Path(self.temp_dir.name) / "test.db"
        self.db = InvoiceDB(self.db_path)

        # Seed mock invoices
        self.db._conn.execute("""
            INSERT INTO invoices (invoice_number, invoice_date, total_amount, seller_name, is_deleted, review_status)
            VALUES ('111', '2026-06-11', '10.00', 'SellerA', 0, 'to_review'),
                   ('222', '2026-06-12', '20.00', 'SellerB', 0, 'to_review'),
                   ('333', '2026-06-13', '30.00', 'SellerC', 0, 'to_review')
        """)
        self.db._conn.commit()

        # Patch QMessageBox dialogs to prevent blocking popup dialogs
        self.msgbox_patcher = patch.object(QMessageBox, "question", return_value=QMessageBox.Yes)
        self.msgbox_patcher.start()

        from scripts.invoice_fetch.gui.app import InvoiceReviewApp
        with patch("scripts.invoice_fetch.gui.app.load_config_safe", return_value={}):
            self.review_app = InvoiceReviewApp(db_path=self.db_path)
            self.review_app._deferred_init()  # Initialize DB, categories, and load table rows
        self.review_app.show()
        self.app.processEvents()

    def tearDown(self):
        self.msgbox_patcher.stop()
        if hasattr(self, "review_app") and self.review_app is not None:
            if hasattr(self.review_app, "db") and self.review_app.db is not None:
                self.review_app.db.close()
            self.review_app.close()
            self.review_app.deleteLater()
        self.db.close()
        gc.collect()
        if self.app:
            self.app.processEvents()
        try:
            self.temp_dir.cleanup()
        except Exception:
            pass

    def test_toolbar_buttons_busy_state(self):
        app = self.review_app
        # Initial states
        self.assertTrue(app.btn_import_local.isEnabled())
        self.assertTrue(app.btn_mobile_upload.isEnabled())
        self.assertTrue(app.btn_scan_email.isEnabled())
        self.assertTrue(app.btn_toolbar_export.isEnabled())

        # Set import local busy
        app._set_action_busy(app.btn_import_local, "导入中...")
        self.assertEqual(app.btn_import_local.text(), "导入中...")
        self.assertEqual(app.btn_import_local.property("busy"), "true")

        # Others must be disabled
        self.assertFalse(app.btn_import_local.isEnabled())
        self.assertFalse(app.btn_mobile_upload.isEnabled())
        self.assertFalse(app.btn_scan_email.isEnabled())
        self.assertFalse(app.btn_toolbar_export.isEnabled())

        # Clear busy
        app._clear_action_busy(app.btn_import_local, "导入发票")
        self.assertEqual(app.btn_import_local.text(), "导入发票")
        self.assertEqual(app.btn_import_local.property("busy"), "false")

        # Restored to enabled
        self.assertTrue(app.btn_import_local.isEnabled())
        self.assertTrue(app.btn_mobile_upload.isEnabled())
        self.assertTrue(app.btn_scan_email.isEnabled())
        self.assertTrue(app.btn_toolbar_export.isEnabled())

    def test_delete_multiple_rows_deduplication_and_selection_hint(self):
        app = self.review_app
        self.assertEqual(app.table.rowCount(), 3)

        # Explicitly select rows 0 and 1 using QItemSelectionModel flags
        app.table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        model = app.table.selectionModel()
        model.clearSelection()

        # Select row 0
        idx0 = app.table.model().index(0, 0)
        model.select(idx0, QItemSelectionModel.Select | QItemSelectionModel.Rows)

        # Select row 1
        idx1 = app.table.model().index(1, 0)
        model.select(idx1, QItemSelectionModel.Select | QItemSelectionModel.Rows)

        # Confirm selected rows count
        self.assertEqual(len(model.selectedRows()), 2)

        # Trigger delete selected
        app._delete_selected_invoices()

        # Database should reflect deletion (only 1 undeleted remains)
        remaining = [r for r in app.db.get_all_invoices() if r.get("is_deleted") != 1]
        self.assertEqual(len(remaining), 1)
        self.assertEqual(remaining[0]["invoice_number"], "111")

        # UI table should reload and select the remaining row
        # min_selected_row was 0. Remaining row index in new list is 0.
        # So selection should be auto-restored to row 0.
        self.assertEqual(app.table.currentRow(), 0)
        self.assertEqual(app.invoices_list[0]["invoice_number"], "111")

    def test_link_to_claim_keeps_original_invoice_selected_when_still_visible(self):
        app = self.review_app
        claim_id = app.db.create_claim_group("Selection claim")
        app._load_claims()
        claim_idx = app.combo_claims.findData(claim_id)
        self.assertGreaterEqual(claim_idx, 0)
        app.combo_claims.setCurrentIndex(claim_idx)

        app.table.selectRow(1)
        self.app.processEvents()
        selected_id = app.invoices_list[1]["id"]

        with patch.object(QMessageBox, "information", return_value=QMessageBox.Ok):
            app._link_invoices_to_claim()
        self.app.processEvents()

        self.assertEqual(app.invoices_list[app.table.currentRow()]["id"], selected_id)

    def test_link_to_claim_under_unlinked_filter_keeps_same_row_position(self):
        app = self.review_app
        claim_id = app.db.create_claim_group("Filtered selection claim")
        app._load_claims()
        claim_idx = app.combo_claims.findData(claim_id)
        self.assertGreaterEqual(claim_idx, 0)
        app.combo_claims.setCurrentIndex(claim_idx)

        app.chk_unlinked.setChecked(True)
        app.search_reload_timer.stop()
        app._load_invoices()
        self.assertEqual(app.table.rowCount(), 3)

        app.table.selectRow(1)
        self.app.processEvents()
        removed_id = app.invoices_list[1]["id"]
        expected_id = app.invoices_list[2]["id"]

        with patch.object(QMessageBox, "information", return_value=QMessageBox.Ok):
            app._link_invoices_to_claim()
        self.app.processEvents()

        self.assertEqual(app.table.currentRow(), 1)
        self.assertEqual(app.invoices_list[1]["id"], expected_id)
        self.assertNotIn(removed_id, [inv["id"] for inv in app.invoices_list])

    def test_capture_selection_row_hint_uses_valid_minimum_row(self):
        app = self.review_app
        original_table = app.table
        selection_model = MagicMock()
        fake_table = MagicMock()
        fake_table.selectionModel.return_value = selection_model

        row_one = MagicMock()
        row_one.row.return_value = 1
        row_two = MagicMock()
        row_two.row.return_value = 2
        invalid_row = MagicMock()
        invalid_row.row.return_value = 99

        try:
            app.table = fake_table
            selection_model.selectedRows.return_value = [row_two]
            selection_model.selectedIndexes.return_value = [row_one]
            self.assertEqual(app._capture_selection_row_hint(), 2)

            selection_model.selectedRows.return_value = []
            selection_model.selectedIndexes.return_value = [
                row_two,
                row_one,
                row_one,
                invalid_row,
            ]
            self.assertEqual(app._capture_selection_row_hint(), 1)

            selection_model.selectedIndexes.return_value = [invalid_row]
            self.assertEqual(app._capture_selection_row_hint(), -1)
        finally:
            app.table = original_table

    def test_unlink_from_claim_keeps_original_invoice_selected(self):
        app = self.review_app
        claim_id = app.db.create_claim_group("Unlink selection claim")
        selected_id = app.invoices_list[1]["id"]
        self.assertTrue(app.db.add_invoice_to_claim(claim_id, selected_id))
        app._load_claims()
        claim_idx = app.combo_claims.findData(claim_id)
        self.assertGreaterEqual(claim_idx, 0)
        app.combo_claims.setCurrentIndex(claim_idx)
        app._load_invoices()

        selected_row = next(
            idx for idx, inv in enumerate(app.invoices_list)
            if inv["id"] == selected_id
        )
        app.table.selectRow(selected_row)
        self.app.processEvents()

        with patch.object(QMessageBox, "information", return_value=QMessageBox.Ok):
            app._unlink_selected_invoices()
        self.app.processEvents()

        self.assertEqual(app.invoices_list[app.table.currentRow()]["id"], selected_id)

    def test_restore_invoice_keeps_original_invoice_selected(self):
        app = self.review_app
        selected_id = app.invoices_list[1]["id"]
        self.assertTrue(app.db.soft_delete_invoice(selected_id))
        app.chk_show_deleted.setChecked(True)
        app.search_reload_timer.stop()
        app._load_invoices()

        selected_row = next(
            idx for idx, inv in enumerate(app.invoices_list)
            if inv["id"] == selected_id
        )
        app.table.selectRow(selected_row)
        self.app.processEvents()
        app._restore_selected_invoices()
        self.app.processEvents()

        self.assertEqual(app.invoices_list[app.table.currentRow()]["id"], selected_id)
        self.assertEqual(app.db.get_invoice(selected_id)["is_deleted"], 0)

    def test_context_actions_capture_selection_row_hint_before_reload(self):
        app = self.review_app
        claim_id = app.db.create_claim_group("Context action claim")
        app._load_claims()
        claim_idx = app.combo_claims.findData(claim_id)
        self.assertGreaterEqual(claim_idx, 0)
        app.combo_claims.setCurrentIndex(claim_idx)

        app.table.selectRow(1)
        self.app.processEvents()

        captured_hints = []

        def record_reload():
            captured_hints.append(app._select_row_hint)

        with patch.object(app, "_load_invoices", side_effect=record_reload), \
                patch.object(app, "_load_claims"), \
                patch.object(QMessageBox, "information", return_value=QMessageBox.Ok):
            app._link_invoices_to_claim()
        self.assertEqual(captured_hints[-1], 1)

        captured_hints.clear()
        with patch.object(app, "_load_invoices", side_effect=record_reload), \
                patch.object(app, "_load_claims"), \
                patch.object(QMessageBox, "information", return_value=QMessageBox.Ok):
            app._unlink_selected_invoices()
        self.assertEqual(captured_hints[-1], 1)

        selected_id = app.invoices_list[1]["id"]
        app.db.soft_delete_invoice(selected_id)
        captured_hints.clear()
        with patch.object(app, "_load_invoices", side_effect=record_reload), \
                patch.object(app, "_load_claims"):
            app._restore_selected_invoices()
        self.assertEqual(captured_hints[-1], 1)

        captured_hints.clear()
        with patch.object(app, "_load_invoices", side_effect=record_reload), \
                patch.object(app, "_load_claims"), \
                patch.object(QMessageBox, "information", return_value=QMessageBox.Ok):
            app._reparse_selected_invoices()
        self.assertEqual(captured_hints[-1], 1)

        class FakeDownloader:
            def __init__(self, download_dir):
                self.download_dir = download_dir

            def close(self):
                return None

        captured_hints.clear()
        with patch("scripts.invoice_fetch.link_downloader.LinkDownloader", FakeDownloader), \
                patch.object(app, "_load_invoices", side_effect=record_reload), \
                patch.object(app, "_load_claims"), \
                patch.object(QMessageBox, "information", return_value=QMessageBox.Ok):
            app._redownload_selected_invoices()
        self.assertEqual(captured_hints[-1], 1)
