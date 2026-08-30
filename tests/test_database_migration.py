import gc
import os
import sqlite3
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from scripts.invoice_fetch.db import InvoiceDB, SQLITE_BUSY_TIMEOUT_MS
from scripts.invoice_fetch.migrations import (
    LATEST_INDEX_CONTRACT,
    LATEST_SCHEMA_VERSION,
    check_and_migrate,
    validate_latest_schema,
)


class DatabaseMigrationTests(unittest.TestCase):
    def test_latest_database_uses_wal_busy_timeout_and_v8_indexes(self):
        with tempfile.TemporaryDirectory() as td:
            with InvoiceDB(Path(td) / "settings.db") as db:
                self.assertEqual(
                    db._conn.execute("PRAGMA journal_mode").fetchone()[0],
                    "wal",
                )
                self.assertEqual(
                    db._conn.execute("PRAGMA busy_timeout").fetchone()[0],
                    SQLITE_BUSY_TIMEOUT_MS,
                )
                self.assertEqual(
                    db._conn.execute("PRAGMA user_version").fetchone()[0],
                    LATEST_SCHEMA_VERSION,
                )
                indexes = {
                    row[0]
                    for row in db._conn.execute(
                        "SELECT name FROM sqlite_master WHERE type = 'index'"
                    )
                }
                self.assertTrue(LATEST_INDEX_CONTRACT <= indexes)
                validate_latest_schema(db._conn)

    def test_v8_indexes_match_production_query_plans(self):
        with tempfile.TemporaryDirectory() as td:
            with InvoiceDB(Path(td) / "plans.db") as db:
                plans = {
                    "idx_emails_pending": db._conn.execute(
                        "EXPLAIN QUERY PLAN SELECT uid FROM emails "
                        "WHERE mailbox_key = ? AND is_invoice = 1 AND downloaded = 0",
                        ("primary",),
                    ).fetchall(),
                    "idx_invoices_file_hash": db._conn.execute(
                        "EXPLAIN QUERY PLAN SELECT id FROM invoices WHERE file_hash = ?",
                        ("hash",),
                    ).fetchall(),
                    "idx_invoices_review_order": db._conn.execute(
                        "EXPLAIN QUERY PLAN SELECT id FROM invoices "
                        "WHERE is_deleted = 0 AND review_status = ? "
                        "ORDER BY expense_date DESC, id DESC",
                        ("to_review",),
                    ).fetchall(),
                    "idx_claim_items_invoice": db._conn.execute(
                        "EXPLAIN QUERY PLAN SELECT COUNT(*) FROM claim_group_items "
                        "WHERE invoice_id = ?",
                        (1,),
                    ).fetchall(),
                }
                for index_name, rows in plans.items():
                    detail = " ".join(str(row[3]) for row in rows)
                    self.assertIn(index_name, detail, detail)

    def test_wal_recovers_committed_data_after_forced_process_exit(self):
        with tempfile.TemporaryDirectory() as td:
            db_path = Path(td) / "forced-exit.db"
            code = "\n".join(
                (
                    "import os",
                    "from pathlib import Path",
                    "from scripts.invoice_fetch.db import InvoiceDB",
                    f"db = InvoiceDB(Path({str(db_path)!r}))",
                    "db._conn.execute('PRAGMA wal_autocheckpoint = 0')",
                    "db._conn.execute('CREATE TABLE crash_marker(value TEXT NOT NULL)')",
                    "db._conn.execute(\"INSERT INTO crash_marker VALUES ('committed')\")",
                    "db._conn.commit()",
                    "os._exit(0)",
                )
            )
            env = os.environ.copy()
            env["PYTHONPATH"] = str(Path(__file__).resolve().parents[1])
            completed = subprocess.run(
                [sys.executable, "-c", code],
                cwd=Path(__file__).resolve().parents[1],
                env=env,
                check=False,
                timeout=20,
            )
            self.assertEqual(completed.returncode, 0)

            with InvoiceDB(db_path) as recovered:
                value = recovered._conn.execute(
                    "SELECT value FROM crash_marker"
                ).fetchone()[0]
                self.assertEqual(value, "committed")
                self.assertEqual(
                    recovered._conn.execute("PRAGMA quick_check").fetchone()[0],
                    "ok",
                )

    def test_bulk_email_writes_preserve_namespace_duplicates_and_classification(self):
        with tempfile.TemporaryDirectory() as td:
            with InvoiceDB(Path(td) / "bulk.db") as db:
                rows = [
                    {
                        "uid": uid,
                        "subject": f"Subject {uid}",
                        "sender": "sender@example.com",
                        "date": "2026-08-30",
                    }
                    for uid in range(1, 101)
                ]
                self.assertEqual(db.bulk_upsert_emails(rows, "primary"), 100)
                self.assertEqual(db.bulk_upsert_emails(rows, "primary"), 0)
                self.assertEqual(db.bulk_upsert_emails(rows, "secondary"), 100)

                db.bulk_classify(
                    [
                        {
                            "mailbox_key": "primary",
                            "uid": uid,
                            "is_invoice": uid % 2 == 0,
                            "by": "batch-test",
                            "reason": "synthetic",
                        }
                        for uid in range(1, 101)
                    ]
                )
                classified = db._conn.execute(
                    "SELECT COUNT(*) FROM emails WHERE mailbox_key = 'primary' "
                    "AND classify_by = 'batch-test'"
                ).fetchone()[0]
                untouched = db._conn.execute(
                    "SELECT COUNT(*) FROM emails WHERE mailbox_key = 'secondary' "
                    "AND is_invoice = -1"
                ).fetchone()[0]
                self.assertEqual(classified, 100)
                self.assertEqual(untouched, 100)

    def test_migration_can_run_twice_without_failure(self):
        with tempfile.TemporaryDirectory() as td:
            db_path = Path(td) / "test.db"

            # 1. Initialize and perform migration
            db = InvoiceDB(db_path)
            db.close()

            # 2. Run migrations again manually
            conn = sqlite3.connect(str(db_path))
            try:
                check_and_migrate(conn)
            except Exception as e:
                self.fail(f"Re-running migration failed: {e}")
            finally:
                conn.close()

    def test_old_minimal_invoices_table_can_be_upgraded(self):
        with tempfile.TemporaryDirectory() as td:
            db_path = Path(td) / "test_old.db"

            # 1. Manually create an old minimal invoices table (V0)
            conn = sqlite3.connect(str(db_path))
            cursor = conn.cursor()
            cursor.execute("""
            CREATE TABLE invoices (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_number  TEXT,
                invoice_date    TEXT,
                total_amount    TEXT,
                seller_name     TEXT,
                category        TEXT DEFAULT '其他',
                UNIQUE(invoice_number, total_amount, seller_name)
            );
            """)
            conn.commit()
            conn.close()

            # 2. Open via InvoiceDB which runs migration
            db = InvoiceDB(db_path)

            # 3. Check that the new V1 columns exist
            conn = sqlite3.connect(str(db_path))
            cursor = conn.cursor()
            cursor.execute("PRAGMA table_info(invoices)")
            cols = {row[1] for row in cursor.fetchall()}
            conn.close()

            new_columns = [
                "review_status", "processing_status", "currency",
                "exchange_rate", "amount_home", "file_hash",
                "confirmed_at", "confirmed_note"
            ]
            for col in new_columns:
                self.assertIn(col, cols)

            db.close()

    def test_new_columns_exist_after_opening_invoicedb(self):
        with tempfile.TemporaryDirectory() as td:
            db_path = Path(td) / "test_new.db"
            db = InvoiceDB(db_path)

            # Check PRAGMA table_info
            conn = sqlite3.connect(str(db_path))
            cursor = conn.cursor()
            cursor.execute("PRAGMA table_info(invoices)")
            cols = {row[1] for row in cursor.fetchall()}
            conn.close()

            new_columns = [
                "review_status", "processing_status", "currency",
                "exchange_rate", "amount_home", "file_hash",
                "confirmed_at", "confirmed_note"
            ]
            for col in new_columns:
                self.assertIn(col, cols)

            db.close()

    def test_existing_insert_invoice_and_get_all_invoices_still_work(self):
        with tempfile.TemporaryDirectory() as td:
            db_path = Path(td) / "test_ops.db"
            db = InvoiceDB(db_path)

            invoice_rec = {
                "invoice_number": "TEST1234",
                "invoice_code": "CODE5678",
                "invoice_date": "2026-05-24",
                "amount": "100.00",
                "total_amount": "106.00",
                "seller_name": "Test Seller",
                "buyer_name": "Test Buyer",
                "invoice_type": "电子发票",
                "category": "住宿费",
                "has_extra": False,
                "extra_type": "",
                "missing_extra": 0,
                "mail_uid": 100,
                "mail_subject": "Invoice email",
                "mail_date": "2026-05-24",
                "mail_sender": "billing@example.com",
                "parse_success": 1,
                "parse_note": "Parsed successfully",
                "attachment_path": "attachments/invoice.pdf",
                "extra_paths": [],
                "download_url": "http://example.com/pdf",
                # New fields
                "review_status": "approved",
                "processing_status": "done",
                "currency": "USD",
                "exchange_rate": "7.2345",
                "amount_home": "723.45",
                "file_hash": "abcde12345",
                "confirmed_at": "2026-05-24 13:00:00",
                "confirmed_note": "All good"
            }

            row_id = db.insert_invoice(invoice_rec)
            self.assertIsNotNone(row_id)

            all_invs = db.get_all_invoices()
            self.assertEqual(len(all_invs), 1)

            retrieved = all_invs[0]
            self.assertEqual(retrieved["invoice_number"], "TEST1234")
            self.assertEqual(retrieved["review_status"], "approved")
            self.assertEqual(retrieved["currency"], "USD")
            self.assertEqual(retrieved["exchange_rate"], "7.2345")
            self.assertEqual(retrieved["amount_home"], "723.45")
            self.assertEqual(retrieved["file_hash"], "abcde12345")

            db.close()

    def test_count_invoices_respects_soft_deleted_flag(self):
        with tempfile.TemporaryDirectory() as td:
            db_path = Path(td) / "test_count.db"
            db = InvoiceDB(db_path)

            first_id = db.insert_invoice({
                "invoice_number": "COUNT001",
                "total_amount": "10.00",
                "seller_name": "Seller A",
                "invoice_date": "2026-06-01",
            })
            db.insert_invoice({
                "invoice_number": "COUNT002",
                "total_amount": "20.00",
                "seller_name": "Seller B",
                "invoice_date": "2026-06-02",
            })

            self.assertEqual(db.count_invoices(), 2)
            self.assertEqual(db.count_invoices(include_deleted=True), 2)

            self.assertTrue(db.soft_delete_invoice(first_id))
            self.assertEqual(db.count_invoices(), 1)
            self.assertEqual(db.count_invoices(include_deleted=True), 2)

            db.close()

    def test_mailbox_key_namespaces_email_and_processed_uid_tables(self):
        with tempfile.TemporaryDirectory() as td:
            db_path = Path(td) / "test_mailbox_namespace.db"
            db = InvoiceDB(db_path)

            self.assertTrue(db.upsert_email(100, "Subject A", "sender@example.com", "2026-06-01", mailbox_key="primary@qq.com"))
            self.assertTrue(db.upsert_email(100, "Subject B", "sender@example.com", "2026-06-01", mailbox_key="secondary@example.com"))
            self.assertFalse(db.upsert_email(100, "Subject C", "sender@example.com", "2026-06-01", mailbox_key="primary@qq.com"))

            self.assertEqual(db.get_all_email_uids("primary@qq.com"), {100})
            self.assertEqual(db.get_all_email_uids("secondary@example.com"), {100})
            self.assertEqual(db.get_all_email_uids("legacy"), set())

            db.mark_email_processed(100, subject="Processed A", sender="sender@example.com", mail_date="2026-06-01", mailbox_key="primary@qq.com")
            db.mark_email_processed(100, subject="Processed B", sender="sender@example.com", mail_date="2026-06-01", mailbox_key="secondary@example.com")

            self.assertTrue(db.is_email_processed(100, mailbox_key="primary@qq.com"))
            self.assertTrue(db.is_email_processed(100, mailbox_key="secondary@example.com"))
            self.assertEqual(db.get_processed_uids("primary@qq.com"), {100})
            self.assertEqual(db.get_processed_uids("secondary@example.com"), {100})

            db.close()

    def test_legacy_email_tables_are_migrated_with_legacy_mailbox_key(self):
        with tempfile.TemporaryDirectory() as td:
            db_path = Path(td) / "test_legacy_mailboxes.db"
            conn = sqlite3.connect(str(db_path))
            cursor = conn.cursor()
            cursor.execute("""
            CREATE TABLE emails (
                uid INTEGER PRIMARY KEY,
                subject TEXT NOT NULL DEFAULT '',
                sender TEXT NOT NULL DEFAULT '',
                mail_date TEXT NOT NULL DEFAULT '',
                is_invoice INTEGER NOT NULL DEFAULT -1,
                classify_by TEXT NOT NULL DEFAULT '',
                classify_reason TEXT NOT NULL DEFAULT '',
                downloaded INTEGER NOT NULL DEFAULT 0,
                scanned_at TEXT DEFAULT (datetime('now', 'localtime')),
                processed_at TEXT
            );
            """)
            cursor.execute("""
            CREATE TABLE processed_emails (
                uid INTEGER PRIMARY KEY,
                subject TEXT,
                sender TEXT,
                mail_date TEXT,
                processed_at TEXT DEFAULT (datetime('now', 'localtime'))
            );
            """)
            cursor.execute("INSERT INTO emails (uid, subject, sender, mail_date) VALUES (101, 'Legacy mail', 'legacy@example.com', '2026-06-01')")
            cursor.execute("INSERT INTO processed_emails (uid, subject, sender, mail_date) VALUES (101, 'Legacy mail', 'legacy@example.com', '2026-06-01')")
            conn.commit()
            conn.close()

            db = InvoiceDB(db_path)
            try:
                self.assertEqual(db.get_all_email_uids("legacy"), {101})
                self.assertEqual(db.get_processed_uids("legacy"), {101})
                self.assertTrue(db.is_email_processed(101, mailbox_key="legacy"))
                self.assertEqual(db.get_all_email_uids("primary@qq.com"), set())
            finally:
                db.close()

    def test_excel_export_includes_review_status_if_present(self):
        from scripts.invoice_fetch.excel_export import export_excel
        import openpyxl

        with tempfile.TemporaryDirectory() as td:
            # 1. Row with review_status
            rows_with_status = [{
                "invoice_number": "NUM123",
                "invoice_date": "2026-05-24",
                "total_amount": "100.00",
                "review_status": "approved",
            }]
            dest_with = Path(td) / "summary_with.xlsx"
            export_excel(rows_with_status, dest_with)

            wb_with = openpyxl.load_workbook(dest_with)
            ws_with = wb_with["发票汇总"]
            headers_with = [ws_with.cell(row=1, column=col).value for col in range(1, ws_with.max_column + 1)]
            self.assertIn("审核状态", headers_with)
            wb_with.close()
            del ws_with, wb_with
            gc.collect()

            # 2. Row without review_status
            rows_without_status = [{
                "invoice_number": "NUM123",
                "invoice_date": "2026-05-24",
                "total_amount": "100.00",
            }]
            dest_without = Path(td) / "summary_without.xlsx"
            export_excel(rows_without_status, dest_without)

            wb_without = openpyxl.load_workbook(dest_without)
            ws_without = wb_without["发票汇总"]
            headers_without = [ws_without.cell(row=1, column=col).value for col in range(1, ws_without.max_column + 1)]
            self.assertNotIn("审核状态", headers_without)
            wb_without.close()
            del ws_without, wb_without
            gc.collect()

    def test_excel_export_includes_confirmed_note_column(self):
        from scripts.invoice_fetch.excel_export import export_excel
        import openpyxl

        with tempfile.TemporaryDirectory() as td:
            rows = [{
                "invoice_number": "NUM123",
                "invoice_date": "2026-05-24",
                "total_amount": "100.00",
                "parse_note": "Parsed from PDF",
                "confirmed_note": "Lunch with client",
            }]
            dest = Path(td) / "summary_notes.xlsx"
            export_excel(rows, dest)

            wb = openpyxl.load_workbook(dest)
            ws = wb["发票汇总"]
            headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
            self.assertIn("解析备注", headers)
            self.assertIn("个人备注", headers)
            self.assertEqual(ws.cell(row=2, column=headers.index("解析备注") + 1).value, "Parsed from PDF")
            self.assertEqual(ws.cell(row=2, column=headers.index("个人备注") + 1).value, "Lunch with client")
            wb.close()
            del ws, wb
            gc.collect()

    def test_old_database_migration_preserves_data_and_insert_works(self):
        with tempfile.TemporaryDirectory() as td:
            db_path = Path(td) / "test_migration_preserves.db"

            # 1. Create a V0 SQLite database manually
            conn = sqlite3.connect(str(db_path))
            cursor = conn.cursor()
            cursor.execute("""
            CREATE TABLE invoices (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_number  TEXT,
                invoice_date    TEXT,
                total_amount    TEXT,
                seller_name     TEXT,
                category        TEXT DEFAULT '其他',
                UNIQUE(invoice_number, total_amount, seller_name)
            );
            """)
            # Insert V0 dummy invoice data
            cursor.execute("""
            INSERT INTO invoices (invoice_number, invoice_date, total_amount, seller_name, category)
            VALUES ('INV_V0_ABC', '2026-05-20', '200.00', 'V0 Seller', '住宿费')
            """)
            conn.commit()
            conn.close()

            # 2. Open via InvoiceDB to trigger check_and_migrate() (V0 -> V1 -> V2)
            with InvoiceDB(db_path) as db:
                # 3. Verify old data is preserved and has migrated default values
                all_invs = db.get_all_invoices()
                self.assertEqual(len(all_invs), 1)

                v0_inv = all_invs[0]
                self.assertEqual(v0_inv["invoice_number"], "INV_V0_ABC")
                self.assertEqual(v0_inv["total_amount"], "200.00")
                self.assertEqual(v0_inv["seller_name"], "V0 Seller")
                self.assertEqual(v0_inv["category"], "住宿费")
                self.assertEqual(v0_inv["review_status"], "to_review") # V1 column default

                # 4. Prove that inserting a new V1/V2 invoice still works perfectly
                new_inv_rec = {
                    "invoice_number": "INV_V2_DEF",
                    "total_amount": "300.00",
                    "seller_name": "V2 Seller",
                    "category": "交通",
                    "review_status": "approved",
                    "currency": "EUR",
                    "exchange_rate": "7.8500",
                    "amount_home": "2355.00",
                    "file_hash": "hashv2",
                    "confirmed_at": "2026-05-24 14:00:00"
                }
                new_row_id = db.insert_invoice(new_inv_rec)
                self.assertIsNotNone(new_row_id)

                # Query and check both records
                updated_invs = db.get_all_invoices()
                self.assertEqual(len(updated_invs), 2)

                # Check newly inserted V2 invoice details
                new_inv = next(i for i in updated_invs if i["invoice_number"] == "INV_V2_DEF")
                self.assertEqual(new_inv["review_status"], "approved")
                self.assertEqual(new_inv["currency"], "EUR")
                self.assertEqual(new_inv["exchange_rate"], "7.8500")
                self.assertEqual(new_inv["amount_home"], "2355.00")
                self.assertEqual(new_inv["file_hash"], "hashv2")

    def test_excel_export_masks_download_url(self):
        from scripts.invoice_fetch.excel_export import export_excel
        import openpyxl

        with tempfile.TemporaryDirectory() as td:
            rows = [{
                "invoice_number": "NUM123",
                "invoice_date": "2026-05-24",
                "total_amount": "100.00",
                "download_url": "https://example.com/pdf?token=secret&invoice=NUM123#frag",
            }]
            dest = Path(td) / "summary_url.xlsx"
            export_excel(rows, dest)

            wb = openpyxl.load_workbook(dest)
            ws = wb.worksheets[0]

            self.assertEqual(
                ws.cell(row=2, column=18).value,
                "https://example.com/pdf?token=%2A%2A%2A&invoice=%2A%2A%2A",
            )
            wb.close()
            del ws, wb
            gc.collect()

    def test_excel_export_escapes_formula_like_user_values(self):
        from scripts.invoice_fetch.excel_export import export_excel
        import openpyxl

        with tempfile.TemporaryDirectory() as td:
            rows = [{
                "invoice_number": "NUM123",
                "invoice_date": "2026-05-24",
                "total_amount": "100.00",
                "seller_name": "=HYPERLINK(\"http://evil.example\",\"click\")",
                "mail_subject": "@SUM(1,1)",
            }]
            dest = Path(td) / "summary_formula.xlsx"
            export_excel(rows, dest)

            wb = openpyxl.load_workbook(dest, data_only=False)
            ws = wb.worksheets[0]

            self.assertEqual(ws.cell(row=2, column=7).value, "'=HYPERLINK(\"http://evil.example\",\"click\")")
            self.assertEqual(ws.cell(row=2, column=14).value, "'@SUM(1,1)")
            wb.close()
            del ws, wb
            gc.collect()


if __name__ == "__main__":
    unittest.main()
