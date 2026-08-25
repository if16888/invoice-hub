import gc
import json
import tempfile
import unittest
from pathlib import Path

import openpyxl

from scripts.invoice_fetch import review_status
from scripts.invoice_fetch.claim_export import export_claim_package
from scripts.invoice_fetch.db import InvoiceDB
from scripts.invoice_fetch.excel_export import export_excel


class MvpExportEvidenceTests(unittest.TestCase):
    def test_excel_exports_user_note_and_keeps_parse_note_separate(self):
        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "out.xlsx"
            export_excel([
                {
                    "invoice_number": "INV-NOTE-001",
                    "invoice_date": "2026-05-20",
                    "total_amount": "128.00",
                    "seller_name": "Restaurant A",
                    "category": "项目餐饮",
                    "confirmed_note": "项目A现场调试，和张三、李四午餐",
                    "parse_note": "PDF解析成功",
                    "review_status": review_status.APPROVED,
                }
            ], xlsx_path)

            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb["发票汇总"]
            headers = [cell.value for cell in ws[1]]

            self.assertIn("个人备注", headers)
            self.assertIn("解析备注", headers)

            user_note_col = headers.index("个人备注") + 1
            parse_note_col = headers.index("解析备注") + 1
            self.assertEqual(ws.cell(row=2, column=user_note_col).value, "项目A现场调试，和张三、李四午餐")
            self.assertEqual(ws.cell(row=2, column=parse_note_col).value, "PDF解析成功")
            wb.close()
            del ws, wb
            gc.collect()

    def test_excel_exports_extra_materials_and_warnings(self):
        with tempfile.TemporaryDirectory() as td:
            xlsx_path = Path(td) / "out.xlsx"
            export_excel([
                {
                    "invoice_number": "INV-002",
                    "invoice_date": "2026-05-22",
                    "total_amount": "500.00",
                    "seller_name": "Vendor B",
                    "has_extra": True,
                    "missing_extra": True,
                    "confirmed_note": "Test Note",
                    "warning": "购方抬头不匹配，可能导致退单",
                }
            ], xlsx_path)

            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb["发票汇总"]
            headers = [cell.value for cell in ws[1]]

            self.assertIn("附加材料", headers)
            self.assertIn("缺少附件", headers)
            self.assertIn("个人备注", headers)
            self.assertIn("校验提示", headers)

            has_extra_col = headers.index("附加材料") + 1
            missing_extra_col = headers.index("缺少附件") + 1
            note_col = headers.index("个人备注") + 1
            warn_col = headers.index("校验提示") + 1

            self.assertEqual(ws.cell(row=2, column=has_extra_col).value, "有")
            self.assertEqual(ws.cell(row=2, column=missing_extra_col).value, "缺少")
            self.assertEqual(ws.cell(row=2, column=note_col).value, "Test Note")
            self.assertEqual(ws.cell(row=2, column=warn_col).value, "购方抬头不匹配，可能导致退单")
            wb.close()
            del ws, wb
            gc.collect()

    def test_claim_export_copies_extra_paths_and_records_manifest(self):
        with tempfile.TemporaryDirectory() as td:
            project_root = Path(td) / "project"
            runtime_dir = project_root / "runtime"
            db_path = runtime_dir / "invoices.db"

            attachment_dir = runtime_dir / "attachments" / "2026-05-20"
            attachment_dir.mkdir(parents=True, exist_ok=True)
            invoice_pdf = attachment_dir / "hotel_invoice.pdf"
            folio_pdf = attachment_dir / "hotel_folio.pdf"
            invoice_pdf.write_bytes(b"%PDF-1.4 invoice")
            folio_pdf.write_bytes(b"%PDF-1.4 folio")

            with InvoiceDB(db_path) as db:
                claim_id = db.create_claim_group("2026-05 Project A")
                invoice_id = db.insert_invoice({
                    "invoice_number": "HOTEL-001",
                    "total_amount": "980.00",
                    "seller_name": "Hotel A",
                    "invoice_date": "2026-05-20",
                    "category": "酒店住宿",
                    "attachment_path": "attachments/2026-05-20/hotel_invoice.pdf",
                    "extra_type": "水单",
                    "extra_paths": ["attachments/2026-05-20/hotel_folio.pdf"],
                    "has_extra": True,
                    "missing_extra": False,
                    "confirmed_note": "项目A现场支持住宿",
                    "review_status": review_status.APPROVED,
                })
                db.add_invoice_to_claim(claim_id, invoice_id)

                export_dir = export_claim_package(
                    db,
                    claim_id,
                    project_root,
                    runtime_dir,
                    export_root=project_root / "exports",
                )

            self.assertTrue((export_dir / "attachments" / "2026-05-20_hotel_invoice.pdf").exists())
            self.assertTrue((export_dir / "attachments" / "2026-05-20_hotel_folio.pdf").exists())

            with open(export_dir / "manifest.json", "r", encoding="utf-8") as f:
                manifest = json.load(f)

            item = manifest["items"][0]
            self.assertEqual(item["invoice_number"], "HOTEL-001")
            self.assertEqual(item["extra_type"], "水单")
            self.assertEqual(item["attachment_path"], "attachments/2026-05-20_hotel_invoice.pdf")
            self.assertEqual(item["extra_paths"], ["attachments/2026-05-20_hotel_folio.pdf"])
            self.assertEqual(item["copied_extra_paths"], ["attachments/2026-05-20_hotel_folio.pdf"])
            self.assertNotIn("file_hash", item)
            self.assertTrue(item["has_extra"])
            self.assertFalse(item["missing_extra"])
            self.assertEqual(item["confirmed_note"], "项目A现场支持住宿")

            wb = openpyxl.load_workbook(export_dir / "reimbursement.xlsx")
            ws = wb["发票汇总"]
            headers = [cell.value for cell in ws[1]]
            self.assertIn("个人备注", headers)
            note_col = headers.index("个人备注") + 1
            self.assertEqual(ws.cell(row=2, column=note_col).value, "项目A现场支持住宿")
            wb.close()
            del ws, wb
            gc.collect()


    def test_claim_export_prefixes_attachment_names_by_date_and_fallback(self):
        with tempfile.TemporaryDirectory() as td:
            project_root = Path(td) / "project"
            runtime_dir = project_root / "runtime"
            db_path = runtime_dir / "invoices.db"

            invoice_dir = runtime_dir / "attachments" / "2026-05-21"
            invoice_dir.mkdir(parents=True, exist_ok=True)
            invoice_pdf = invoice_dir / "2026-05-21_meal_invoice.pdf"
            invoice_pdf.write_bytes(b"%PDF-1.4 invoice")

            fallback_dir = runtime_dir / "attachments" / "unknown_date"
            fallback_dir.mkdir(parents=True, exist_ok=True)
            fallback_pdf = fallback_dir / "receipt.pdf"
            fallback_pdf.write_bytes(b"%PDF-1.4 receipt")
            fallback_extra = fallback_dir / "receipt-support.pdf"
            fallback_extra.write_bytes(b"%PDF-1.4 receipt extra")
            unknown_pdf = fallback_dir / "plain.pdf"
            unknown_pdf.write_bytes(b"%PDF-1.4 unknown")

            with InvoiceDB(db_path) as db:
                claim_id = db.create_claim_group("Date Prefix Claim")
                invoice_id = db.insert_invoice({
                    "invoice_number": "DATE-001",
                    "total_amount": "128.00",
                    "seller_name": "Vendor A",
                    "invoice_date": "2026-05-21",
                    "expense_date": "2026-05-21",
                    "category": "餐饮",
                    "attachment_path": "attachments/2026-05-21/2026-05-21_meal_invoice.pdf",
                    "extra_paths": [],
                    "has_extra": False,
                    "missing_extra": False,
                    "review_status": review_status.APPROVED,
                })
                fallback_id = db.insert_invoice({
                    "invoice_number": "DATE-002",
                    "total_amount": "64.00",
                    "seller_name": "Vendor B",
                    "invoice_date": "",
                    "expense_date": "2026-05-22",
                    "category": "餐饮",
                    "attachment_path": "attachments/unknown_date/receipt.pdf",
                    "extra_paths": ["attachments/unknown_date/receipt-support.pdf"],
                    "has_extra": True,
                    "missing_extra": False,
                    "review_status": review_status.APPROVED,
                })
                unknown_id = db.insert_invoice({
                    "invoice_number": "DATE-003",
                    "total_amount": "16.00",
                    "seller_name": "Vendor C",
                    "invoice_date": "",
                    "expense_date": "",
                    "category": "餐饮",
                    "attachment_path": "attachments/unknown_date/plain.pdf",
                    "extra_paths": [],
                    "has_extra": False,
                    "missing_extra": False,
                    "review_status": review_status.APPROVED,
                })
                db.add_invoice_to_claim(claim_id, invoice_id)
                db.add_invoice_to_claim(claim_id, fallback_id)
                db.add_invoice_to_claim(claim_id, unknown_id)

                export_dir = export_claim_package(
                    db,
                    claim_id,
                    project_root,
                    runtime_dir,
                    export_root=project_root / "exports",
                )

            self.assertTrue((export_dir / "attachments" / "2026-05-21_meal_invoice.pdf").exists())
            self.assertTrue((export_dir / "attachments" / "2026-05-22_receipt.pdf").exists())
            self.assertTrue((export_dir / "attachments" / "2026-05-22_receipt-support.pdf").exists())
            self.assertTrue((export_dir / "attachments" / "unknown-date_plain.pdf").exists())

            with open(export_dir / "manifest.json", "r", encoding="utf-8") as f:
                manifest = json.load(f)

            items = {item["invoice_number"]: item for item in manifest["items"]}
            self.assertEqual(items["DATE-001"]["attachment_path"], "attachments/2026-05-21_meal_invoice.pdf")
            self.assertEqual(items["DATE-001"]["extra_paths"], [])
            self.assertEqual(items["DATE-002"]["attachment_path"], "attachments/2026-05-22_receipt.pdf")
            self.assertEqual(items["DATE-002"]["extra_paths"], ["attachments/2026-05-22_receipt-support.pdf"])
            self.assertEqual(items["DATE-003"]["attachment_path"], "attachments/unknown-date_plain.pdf")
            self.assertEqual(items["DATE-003"]["extra_paths"], [])


if __name__ == "__main__":
    unittest.main()
