"""Excel export — generate a styled reimbursement summary from the SQLite DB."""

from __future__ import annotations

import logging
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .url_utils import _mask_url

_log = logging.getLogger(__name__)

_COLUMNS = [
    ("invoice_number",  "发票号码",   14),
    ("invoice_code",    "发票代码",   16),
    ("invoice_date",    "开票日期",   13),
    ("amount",          "金额(税前)", 12),
    ("total_amount",    "价税合计",   12),
    ("seller_name",     "销售方",     28),
    ("buyer_name",      "购买方",     28),
    ("invoice_type",    "发票类型",   18),
    ("category",        "分类",       10),
    ("has_extra",       "附加材料",   10),
    ("missing_extra",   "缺少附件",   10),
    ("parse_note",      "解析备注",    16),
    ("mail_subject",    "邮件主题",   40),
    ("mail_date",       "邮件日期",   13),
    ("attachment_path", "文件路径",   35),
    ("extra_paths",     "证明材料",   30),
    ("download_url",    "下载链接",   30),
    ("confirmed_note",  "用户备注",    24),
    ("warning",         "校验提示",    24),
]

_HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_FONT = Font(name="微软雅黑", size=10)
_LINK_FONT = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
_CELL_ALIGN = Alignment(vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
_ALT_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
_FORMULA_PREFIXES = ("=", "+", "-", "@")


def _safe_excel_value(value):
    """Prevent user-controlled text from being interpreted as an Excel formula."""
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return "'" + value
    return value


def export_excel(rows: list[dict], dest: str | Path) -> Path:
    """Write *rows* (from ``InvoiceDB.get_all_invoices()``) to an Excel file."""
    dest = Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)

    columns = list(_COLUMNS)
    if any(r.get("review_status") for r in rows):
        columns.append(("review_status", "审核状态", 12))

    wb = Workbook()
    ws = wb.active
    ws.title = "发票汇总"

    # ── Header row ───────────────────────────────────────────────────
    for col_idx, (_, label, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 28
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    ws.freeze_panes = "A2"

    # ── Data rows ────────────────────────────────────────────────────
    for row_idx, row in enumerate(rows, 2):
        is_alt = row_idx % 2 == 0
        for col_idx, (key, _, _) in enumerate(columns, 1):
            val = row.get(key, "")
            if key == "has_extra":
                val = "有" if val else ""
            elif key == "missing_extra":
                val = "缺少" if val else ""
            elif key == "extra_paths":
                if isinstance(val, list):
                    val = "\n".join(str(p) for p in val if str(p).strip())
                elif not val:
                    val = ""
            elif key == "download_url":
                val = _mask_url(str(val or ""))
            cell = ws.cell(row=row_idx, column=col_idx, value=_safe_excel_value(val))
            cell.font = _CELL_FONT
            cell.alignment = _CELL_ALIGN
            cell.border = _THIN_BORDER
            if is_alt:
                cell.fill = _ALT_FILL
            if key == "attachment_path" and val:
                cell.hyperlink = str(val)
                cell.font = _LINK_FONT

    _add_summary_sheet(wb, rows)
    _add_exception_sheet(wb, rows)

    wb.save(str(dest))
    _log.info("Excel 已导出: %s (%d 条记录)", dest.name, len(rows))
    return dest


def _amount(value) -> Decimal:
    try:
        return Decimal(str(value or "0").replace(",", ""))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _style_header(ws, labels: list[str], widths: list[int]):
    for idx, label in enumerate(labels, 1):
        cell = ws.cell(row=1, column=idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = widths[idx - 1]
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def _style_cell(cell):
    cell.font = _CELL_FONT
    cell.alignment = _CELL_ALIGN
    cell.border = _THIN_BORDER


def _add_summary_sheet(wb: Workbook, rows: list[dict]):
    ws = wb.create_sheet("分类汇总")
    _style_header(ws, ["分类", "张数", "价税合计"], [16, 10, 14])

    summary = defaultdict(lambda: {"count": 0, "amount": Decimal("0")})
    for row in rows:
        category = row.get("category") or "未分类"
        summary[category]["count"] += 1
        summary[category]["amount"] += _amount(row.get("total_amount"))

    for row_idx, (category, item) in enumerate(summary.items(), 2):
        values = [category, item["count"], float(item["amount"])]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_safe_excel_value(val))
            _style_cell(cell)
        ws.cell(row=row_idx, column=3).number_format = "#,##0.00"

    total_row = len(summary) + 2
    for col_idx, val in enumerate(["总计", len(rows), float(sum((_amount(r.get("total_amount")) for r in rows), Decimal("0")))], 1):
        cell = ws.cell(row=total_row, column=col_idx, value=val)
        _style_cell(cell)
        cell.font = Font(name="微软雅黑", bold=True, size=10)
    ws.cell(row=total_row, column=3).number_format = "#,##0.00"


def _is_exception(row: dict) -> bool:
    return bool(row.get("missing_extra")) or not row.get("attachment_path") or row.get("parse_success") == 0


def _add_exception_sheet(wb: Workbook, rows: list[dict]):
    ws = wb.create_sheet("异常待处理")
    exc_columns = [
        ("invoice_number", "发票号码", 16),
        ("invoice_date", "开票日期", 13),
        ("total_amount", "价税合计", 12),
        ("category", "分类", 12),
        ("parse_note", "解析备注", 24),
        ("confirmed_note", "用户备注", 24),
        ("extra_paths", "证明材料", 30),
        ("mail_subject", "邮件主题", 42),
        ("attachment_path", "文件路径", 35),
    ]
    if any(r.get("review_status") for r in rows):
        exc_columns.append(("review_status", "审核状态", 12))

    _style_header(ws, [c[1] for c in exc_columns], [c[2] for c in exc_columns])

    for row_idx, row in enumerate([r for r in rows if _is_exception(r)], 2):
        for col_idx, (key, _, _) in enumerate(exc_columns, 1):
            val = row.get(key, "")
            if key == "extra_paths":
                if isinstance(val, list):
                    val = "\n".join(str(p) for p in val if str(p).strip())
                elif not val:
                    val = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _style_cell(cell)
            if key == "attachment_path" and val:
                cell.hyperlink = str(val)
                cell.font = _LINK_FONT
